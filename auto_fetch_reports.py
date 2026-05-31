# -*- coding: utf-8 -*-
"""Auto-fetch work reports from MES (v2)."""
import os, sys, time, json, sqlite3
from datetime import datetime

MES_URL = "https://web.ycmes.cn/"
FACTORY_CODE = "606999"
USERNAME = "GS-001"
PASSWORD = "675726"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
DB_PATH = os.path.join(BASE_DIR, "data", "production.db")


def get_last_import_time():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("SELECT MAX(report_time) FROM work_reports")
        result = c.fetchone()[0]
    except:
        result = None
    conn.close()
    return result[:10] if result else None


def import_work_reports(filepath):
    import openpyxl
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    # Row 1 is merged title, Row 2 is headers
    headers = []
    for cell in ws[2]:
        headers.append(str(cell.value or "").strip())

    # Field mapping (Chinese header name -> db field)
    fm = {
        "报工数": "quantity",
        "报工良品数": "good_quantity",
        "报工不良数": "bad_quantity",
        "报工单位": "unit",
        "报工良品率": "good_rate",
        "生产人员": "reporter",
        "报工开始时间": "report_start_time",
        "报工结束时间": "report_end_time",
        "审批状态": "approval_status",
        "审批人": "approver",
        "审批时间": "approval_time",
        "创建人": "creator",
        "报工创建时间": "report_time",
        "工序名称": "process_name",
        "工单编号": "work_order_no",
        "产品编号": "product_code",
        "产品名称": "product_name",
        "关联单据号": "related_doc_no",
        "设备机台": "equipment",
        "报工时长": "duration",
        "焊点数量": "weld_count",
        "出勤人员备注": "attendance_note",
    }

    # Build column index map: db_field -> column_index
    col_map = {}
    for i, h in enumerate(headers):
        if h in fm:
            col_map[fm[h]] = i

    def sget(row, field):
        idx = col_map.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def fget(row, field):
        v = sget(row, field)
        try:
            return float(v) if v else 0
        except:
            return 0

    inserted = skipped = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue
        try:
            rpt = sget(row, "report_time")
            won = sget(row, "work_order_no")
            pn = sget(row, "process_name")
            rep = sget(row, "reporter")
            if not rpt or rpt == "None":
                skipped += 1
                continue
            c.execute("SELECT 1 FROM work_reports WHERE work_order_no=? AND process_name=? AND reporter=? AND report_time=?", (won, pn, rep, rpt))
            if c.fetchone():
                skipped += 1
                continue
            dur = fget(row, "duration")
            c.execute("INSERT INTO work_reports (work_order_no,product_code,product_name,process_name,quantity,good_quantity,bad_quantity,unit,good_rate,reporter,report_start_time,report_end_time,approval_status,approver,approval_time,creator,report_time,related_doc_no,equipment,duration,weld_count,attendance_note,is_overtime,import_time) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (won, sget(row,"product_code"), sget(row,"product_name"), pn,
                 fget(row,"quantity"), fget(row,"good_quantity"), fget(row,"bad_quantity"),
                 sget(row,"unit"), fget(row,"good_rate"), rep,
                 sget(row,"report_start_time"), sget(row,"report_end_time"),
                 sget(row,"approval_status"), sget(row,"approver"),
                 sget(row,"approval_time"), sget(row,"creator"),
                 rpt, sget(row,"related_doc_no"), sget(row,"equipment"),
                 dur, fget(row,"weld_count"), sget(row,"attendance_note"),
                 1 if dur > 9 else 0,
                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            inserted += 1
        except:
            pass
    conn.commit()
    conn.close()
    wb.close()
    return inserted, skipped


def run_fetch():
    from playwright.sync_api import sync_playwright

    last_date = get_last_import_time()
    now = datetime.now().strftime("%H:%M:%S")
    print("[%s] Start (last: %s)" % (now, last_date or "none"))
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()

        try:
            # Login
            page.goto(MES_URL, wait_until="networkidle", timeout=30000)
            time.sleep(2)
            page.locator(u"text=\u8d26\u53f7\u767b\u5f55").first.click()
            time.sleep(1)
            page.locator(u"input[placeholder*=\u5de5\u5382]").first.fill(FACTORY_CODE)
            page.locator(u"input[placeholder*=\u7528\u6237\u540d]").first.fill(USERNAME)
            page.locator("input[type=password]").first.fill(PASSWORD)
            page.locator("button").filter(has_text=u"\u767b\u5f55").first.click()
            page.wait_for_load_state("networkidle", timeout=15000)
            time.sleep(3)
            print("  Login OK")

            # Go to report page
            page.goto("https://web.ycmes.cn/production/output", wait_until="networkidle", timeout=30000)
            time.sleep(5)
            print("  Page loaded")

            # Use expect_download to capture the export file
            print("  Exporting...")
            with page.expect_download(timeout=180000) as dl_info:
                btns = page.locator("button").all()
                for btn in btns:
                    try:
                        txt = btn.text_content() or ""
                        if u"\u5bfc\u51fa" in txt:
                            btn.click()
                            print("  Clicked export")
                            break
                    except:
                        pass
            download = dl_info.value
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(DOWNLOAD_DIR, "work_reports_%s.xlsx" % ts)
            download.save_as(filepath)
            print("  Downloaded: %s" % filepath)

            # Import
            inserted, skipped = import_work_reports(filepath)
            print("  Imported: %d, Skipped: %d" % (inserted, skipped))

            # Cleanup old files
            try:
                files = sorted([f for f in os.listdir(DOWNLOAD_DIR)
                    if f.startswith("work_reports_") and f.endswith(".xlsx")])
                for old in files[:-3]:
                    os.remove(os.path.join(DOWNLOAD_DIR, old))
            except:
                pass

        except Exception as e:
            print("Error: %s" % e)
            import traceback
            traceback.print_exc()
            try:
                page.screenshot(path=os.path.join(DOWNLOAD_DIR, "error.png"))
            except:
                pass
        finally:
            browser.close()



# Also run work order import
try:
    from auto_fetch_orders import run_fetch_orders
    run_fetch_orders()
except Exception as e:
    print("Work order fetch error: %s" % e)

if __name__ == "__main__":
    run_fetch()