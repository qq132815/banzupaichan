# -*- coding: utf-8 -*-
import os
import sys
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)

from utils.db import get_connection

# ===== 1. Create Equipment Data =====
print('Creating equipment data...')
conn = get_connection()
c = conn.cursor()

# Get team IDs
c.execute('SELECT id, name FROM teams')
teams = {row[1]: row[0] for row in c.fetchall()}
print('Teams:', teams)

# Equipment data for each team
equipment_data = [
    # 前段 equipment
    ('QD-001', '自动下料机1号', teams.get('前段', 1), '自动', 100, 'A区'),
    ('QD-002', '自动下料机2号', teams.get('前段', 1), '自动', 100, 'A区'),
    ('QD-003', '手动倒角机1号', teams.get('前段', 1), '手动', 50, 'A区'),
    ('QD-004', '手动倒角机2号', teams.get('前段', 1), '手动', 50, 'A区'),
    ('QD-005', '自动弯管机1号', teams.get('前段', 1), '自动', 80, 'B区'),
    ('QD-006', '自动弯管机2号', teams.get('前段', 1), '自动', 80, 'B区'),
    ('QD-007', '手动弯管机1号', teams.get('前段', 1), '手动', 40, 'B区'),
    ('QD-008', '冲孔机1号', teams.get('前段', 1), '自动', 120, 'C区'),
    ('QD-009', '去毛刺机1号', teams.get('前段', 1), '自动', 150, 'C区'),
    ('QD-010', '清洗机1号', teams.get('前段', 1), '自动', 200, 'C区'),
    
    # 焊接 equipment
    ('HJ-001', '自动焊接机1号', teams.get('焊接', 2), '自动', 60, 'D区'),
    ('HJ-002', '自动焊接机2号', teams.get('焊接', 2), '自动', 60, 'D区'),
    ('HJ-003', '手动焊接机1号', teams.get('焊接', 2), '手动', 30, 'D区'),
    ('HJ-004', '手动焊接机2号', teams.get('焊接', 2), '手动', 30, 'D区'),
    ('HJ-005', '氩弧焊机1号', teams.get('焊接', 2), '手动', 40, 'E区'),
    
    # 扣压 equipment
    ('KY-001', '自动扣压机1号', teams.get('扣压', 3), '自动', 80, 'F区'),
    ('KY-002', '自动扣压机2号', teams.get('扣压', 3), '自动', 80, 'F区'),
    ('KY-003', '手动扣压机1号', teams.get('扣压', 3), '手动', 40, 'F区'),
    
    # 装配包装 equipment
    ('ZB-001', '装配线1号', teams.get('装配包装', 4), '手动', 50, 'G区'),
    ('ZB-002', '装配线2号', teams.get('装配包装', 4), '手动', 50, 'G区'),
    ('ZB-003', '自动涂装线1号', teams.get('装配包装', 4), '自动', 100, 'H区'),
    ('ZB-004', '包装机1号', teams.get('装配包装', 4), '自动', 200, 'H区'),
    ('ZB-005', '成品检验台1号', teams.get('装配包装', 4), '手动', 100, 'H区'),
]

c.execute('DELETE FROM equipments')
for eq in equipment_data:
    c.execute('INSERT INTO equipments (equipment_code, equipment_name, team_id, equipment_type, capacity_per_hour, location) VALUES (?, ?, ?, ?, ?, ?)', eq)

print(f'Created {len(equipment_data)} equipment records')

# ===== 2. Import Process Routes =====
print('Importing process routes...')
wb = openpyxl.load_workbook(os.path.join(BASE, '工艺路线.xlsx'))
ws = wb.active

c.execute('DELETE FROM process_routes')
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    route_code = str(row[0]).strip()
    route_name = str(row[1]).strip() if row[1] else ''
    process_list = str(row[2]).strip() if row[2] else ''
    # Skip header row
    if route_code == '工艺路线编号':
        continue
    c.execute('INSERT INTO process_routes (route_code, route_name, product_code, process_list) VALUES (?, ?, ?, ?)', 
              (route_code, route_name, route_code.split('-')[0] if '-' in route_code else route_code, process_list))
    count += 1

print(f'Imported {count} process routes')

# ===== 3. Connect Processes to Teams =====
print('Connecting processes to teams...')
wb2 = openpyxl.load_workbook(os.path.join(BASE, '工序_全部工序 (1).xlsx'))
ws2 = wb2.active

c.execute('DELETE FROM processes')
count = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    process_code = str(row[0]).strip()
    process_name = str(row[1]).strip() if row[1] else ''
    team_name = str(row[2]).strip() if row[2] else ''
    # Skip header
    if process_code == '工序编号':
        continue
    c.execute('INSERT OR REPLACE INTO processes (process_code, process_name, team_name) VALUES (?, ?, ?)', 
              (process_code, process_name, team_name))
    count += 1

print(f'Imported {count} processes')

# ===== 4. Create process-equipment mapping =====
print('Creating process-equipment mapping...')
c.execute('DELETE FROM process_equipment')

# Map processes to equipment based on team and process type
c.execute('SELECT id, process_code, process_name, team_name FROM processes')
all_processes = c.fetchall()

c.execute('SELECT id, equipment_code, equipment_name, team_id FROM equipments')
all_equipment = c.fetchall()

# Get team name to id mapping
team_name_to_id = {}
for t_name, t_id in teams.items():
    team_name_to_id[t_name] = t_id

mapped = 0
for proc in all_processes:
    proc_id = proc[0]
    proc_code = proc[1]
    proc_name = proc[2]
    proc_team = proc[3]
    
    # Find equipment for this team
    for eq in all_equipment:
        eq_id = eq[0]
        eq_team_id = eq[2]
        
        # Check if this equipment's team matches the process team
        for t_name, t_id in teams.items():
            if t_id == eq_team_id and t_name in proc_team:
                c.execute('INSERT OR IGNORE INTO process_equipment (process_code, equipment_id, is_primary) VALUES (?, ?, ?)',
                         (proc_code, eq_id, 1))
                mapped += 1

print(f'Created {mapped} process-equipment mappings')

conn.commit()
conn.close()
print('Rebuild complete!')