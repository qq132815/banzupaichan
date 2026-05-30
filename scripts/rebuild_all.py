# -*- coding: utf-8 -*-
import os, sys, sqlite3

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)
DB_PATH = os.path.join(BASE, 'data', 'production.db')

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def rebuild_all():
    print('Starting comprehensive rebuild...')
    import openpyxl
    conn = get_connection()
    c = conn.cursor()
    for t in ['alerts','bom','production_cycles','shipping_plan','reports','schedules','process_equipment','processes','process_routes','work_orders','products','equipments','teams','users','daily_plans','production_requirements','publish_batches']:
        c.execute('DROP TABLE IF EXISTS ' + t)
    conn.commit()
    conn.close()
    from utils.db import init_database
    init_database()
    conn = get_connection()
    c = conn.cursor()
    teams = [(1,'前段','','白班','08:00','17:00'),(2,'焊接','','白班','08:00','17:00'),(3,'扣压','','白班','08:00','17:00'),(4,'装配包装','','白班','08:00','17:00')]
    for t in teams:
        c.execute('INSERT OR REPLACE INTO teams (id,name,leader,shift_type,shift_start,shift_end) VALUES (?,?,?,?,?,?)', t)
    equipments = [('QD-001','自动下料机1号',1,'自动','normal',100,'A区'),('QD-002','自动下料机2号',1,'自动','normal',100,'A区'),('QD-003','手动倒角机1号',1,'手动','normal',50,'A区'),('QD-004','手动倒角机2号',1,'手动','normal',50,'A区'),('QD-005','自动弯管机1号',1,'自动','normal',80,'B区'),('QD-006','自动弯管机2号',1,'自动','normal',80,'B区'),('QD-007','手动弯管机1号',1,'手动','normal',40,'B区'),('QD-008','冲孔机1号',1,'自动','normal',120,'C区'),('QD-009','去毛刺机1号',1,'自动','normal',150,'C区'),('QD-010','清洗机1号',1,'自动','normal',200,'C区'),('HJ-001','自动焊接机1号',2,'自动','normal',60,'D区'),('HJ-002','自动焊接机2号',2,'自动','normal',60,'D区'),('HJ-003','手动焊接机1号',2,'手动','normal',30,'D区'),('HJ-004','手动焊接机2号',2,'手动','normal',30,'D区'),('HJ-005','氩弧焊机1号',2,'手动','normal',40,'E区'),('KY-001','自动扣压机1号',3,'自动','normal',80,'F区'),('KY-002','自动扣压机2号',3,'自动','normal',80,'F区'),('KY-003','手动扣压机1号',3,'手动','normal',40,'F区'),('ZB-001','装配线1号',4,'手动','normal',50,'G区'),('ZB-002','装配线2号',4,'手动','normal',50,'G区'),('ZB-003','自动涂装线1号',4,'自动','normal',100,'H区'),('ZB-004','包装机1号',4,'自动','normal',200,'H区'),('ZB-005','成品检验台1号',4,'手动','normal',100,'H区')]
    for eq in equipments:
        c.execute('INSERT INTO equipments (equipment_code,equipment_name,team_id,equipment_type,status,capacity_per_hour,location) VALUES (?,?,?,?,?,?,?)', eq)
    print('Teams & Equipment inserted')
    # Import processes from Excel file
    proc_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), '工序_全部工序 (1).xlsx')
    if os.path.exists(proc_file):
        import openpyxl
        wb = openpyxl.load_workbook(proc_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            process_code = str(row[0]).strip()
            process_name = str(row[1]).strip() if row[1] else ''
            team_name = str(row[2]).strip() if row[2] else ''
            c.execute('INSERT OR REPLACE INTO processes (process_code,process_name,team_name) VALUES (?,?,?)', (process_code, process_name, team_name))
        print('Processes imported from Excel')
    else:
        print(f'Warning: Process file not found: {proc_file}')
    team_map = {'前段':1,'焊接':2,'扣压':3,'装配包装':4}
    c.execute('SELECT id,team_id FROM equipments')
    eq_by_team = {}
    for eq in c.fetchall():
        eq_by_team.setdefault(eq[1],[]).append(eq[0])
    c.execute('SELECT process_code,team_name FROM processes')
    mapped = 0
    for pc_code,tn in c.fetchall():
        tid = team_map.get(tn,1)
        if tid in eq_by_team:
            for eq_id in eq_by_team[tid]:
                c.execute('INSERT OR IGNORE INTO process_equipment (process_code,equipment_id,is_primary) VALUES (?,?,1)',(pc_code,eq_id))
                mapped += 1
    print(f'Process-equipment mappings: {mapped}')
    wb = openpyxl.load_workbook(os.path.join(BASE,'工艺路线.xlsx'))
    ws = wb.active
    c.execute('DELETE FROM process_routes')
    count = 0
    for row in ws.iter_rows(min_row=3,values_only=True):
        if not row or not row[0]:
            continue
        product_code = str(row[0]).strip()
        route_code = str(row[1]).strip() if row[1] else product_code
        route_name = str(row[2]).strip() if row[2] else ''
        process_list = str(row[3]).strip() if row[3] else ''
        c.execute('INSERT INTO process_routes (route_code,route_name,product_code,process_list) VALUES (?,?,?,?)',(route_code,route_name,product_code,process_list))
        count += 1
    print(f'Process routes imported: {count}')

    # Auto-fix: create missing processes from route process_list
    c.execute("SELECT process_list, route_code FROM process_routes")
    routes = c.fetchall()
    created = 0
    for route in routes:
        if route[0]:
            for proc_name in route[0].split(','):
                proc_name = proc_name.strip()
                if not proc_name:
                    continue
                c.execute("SELECT id FROM processes WHERE process_name=?", (proc_name,))
                if not c.fetchone():
                    team_name = ''
                    if '焊接' in (route[1] or ''):
                        team_name = '焊接'
                    elif '前段' in (route[1] or '') or '弯' in (route[1] or '') or '冲' in (route[1] or ''):
                        team_name = '前段'
                    elif '扣压' in (route[1] or ''):
                        team_name = '扣压'
                    elif '装配' in (route[1] or '') or '包装' in (route[1] or ''):
                        team_name = '装配包装'
                    proc_code = 'AUTO_' + str(created + 100)
                    c.execute("INSERT OR IGNORE INTO processes (process_code, process_name, team_name) VALUES (?, ?, ?)", (proc_code, proc_name, team_name))
                    created += 1
    conn.commit()
    print(f"Auto-created {created} missing processes")
    wb2 = openpyxl.load_workbook(os.path.join(BASE,'工单.xlsx'))
    ws2 = wb2.active
    count = 0
    for row in ws2.iter_rows(min_row=3,values_only=True):
        if not row or not row[0]:
            continue
        order_no = str(row[0]).strip()
        product_code = str(row[2]).strip() if row[2] else ''
        quantity = row[5] if row[5] else 0
        priority = str(row[6]).strip() if row[6] else 'P2'
        due_date = row[10].strftime('%Y-%m-%d') if hasattr(row[10],'strftime') else str(row[10]) if row[10] else None
        status = str(row[3]).strip() if len(row)>3 and row[3] else 'pending'
        completed_qty = row[13] if len(row)>13 and row[13] else 0
        process_progress = str(row[4]).strip() if len(row)>4 and row[4] else ''
        c.execute('INSERT OR REPLACE INTO work_orders (order_no,product_code,product_name,quantity,completed_qty,due_date,priority,status,process_progress) VALUES (?,?,?,?,?,?,?,?,?)',(order_no,product_code,product_code,quantity,completed_qty,due_date,priority,status,process_progress))
        count += 1
    print(f'Work orders imported: {count}')
    wb3 = openpyxl.load_workbook(os.path.join(BASE,'发货计划.xlsx'))
    ws3 = wb3.active
    headers = [cell.value for cell in ws3[1]]
    dates = [(i,h) for i,h in enumerate(headers) if hasattr(h,'strftime')]
    count = 0
    for row in ws3.iter_rows(min_row=2,values_only=True):
        if not row or not row[0]:
            continue
        pc = str(row[0]).strip()
        for ci,sd in dates:
            qty = row[ci] if ci<len(row) else None
            if qty and str(qty).isdigit() and int(qty)>0:
                c.execute('INSERT INTO shipping_plan (product_code,quantity,ship_date) VALUES (?,?,?)',(pc,int(qty),sd.strftime('%Y-%m-%d')))
                count += 1
    print(f'Shipping plans imported: {count}')
    wb4 = openpyxl.load_workbook(os.path.join(BASE,'产品生产周期与提前时间表.xlsx'))
    ws4 = wb4.active
    count = 0
    for row in ws4.iter_rows(min_row=2,values_only=True):
        if not row or not row[0]:
            continue
        c.execute('INSERT INTO production_cycles (product_code,production_days,lead_days) VALUES (?,?,?)',(str(row[0]).strip(),row[1] if row[1] else 1,row[2] if row[2] else 0))
        count += 1
    print(f'Production cycles imported: {count}')
    wb5 = openpyxl.load_workbook(os.path.join(BASE,'物料清单 (1).xlsx'))
    ws5 = wb5.active
    count = 0
    for row in ws5.iter_rows(min_row=2,values_only=True):
        if not row or not row[0]:
            continue
        parent_code = str(row[2]).strip() if row[2] else ''  # 父项产品编号
        parent_name = str(row[3]).strip() if row[3] else ''  # 父项产品名称
        child_code = str(row[6]).strip() if row[6] else ''   # 子项产品编号
        child_name = str(row[7]).strip() if row[7] else ''   # 子项产品名称
        # Skip purchased materials (starts with 01.)
        if child_code.startswith('01.'):
            continue
        qty = abs(float(row[10])) if row[10] else 1  # 单位用量
        unit = str(row[9]).strip() if row[9] else ''  # 子项单位
        process_team = str(row[11]).strip() if row[11] else None  # 备注
        c.execute('INSERT INTO bom (parent_product_code,parent_product_name,child_product_code,child_product_name,quantity,unit,process_team) VALUES (?,?,?,?,?,?,?)',
                  (parent_code, parent_name, child_code, child_name, qty, unit, process_team))
        count += 1
    print(f'BOM imported: {count}')
    users = [('admin','admin123','系统管理员','planner',None),('planner','planner123','计划员','planner',None),('qianduan','123456','前段班组','team',1),('hanjie','123456','焊接班组','team',2),('kouya','123456','扣压班组','team',3),('zhuangpei','123456','装配包装班组','team',4)]
    for u in users:
        c.execute('INSERT OR IGNORE INTO users (username,password,display_name,role,team_id) VALUES (?,?,?,?,?)', u)
    print('Default users created')
    conn.commit()
    conn.close()
    print('Rebuild complete!')

if __name__ == '__main__':
    rebuild_all()
