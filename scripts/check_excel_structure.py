# -*- coding: utf-8 -*-
import openpyxl
import os

file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "产品信息_全部.xlsx")
print(f"文件路径: {file_path}")
print(f"文件存在: {os.path.exists(file_path)}")

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"\n工作表: {ws.title}")
    print(f"总行数: {ws.max_row}")
    print(f"总列数: {ws.max_column}")
    
    # 读取表头（第一行）
    print("\n=== 表头信息 ===")
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value)
        print(f"列{col}: {cell_value}")
    
    # 读取前3行数据示例
    print("\n=== 前3行数据示例 ===")
    for row in range(2, min(5, ws.max_row + 1)):
        print(f"\n第{row}行:")
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            print(f"  {headers[col-1]}: {cell_value}")
    
    wb.close()
