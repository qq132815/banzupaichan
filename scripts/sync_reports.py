# -*- coding: utf-8 -*-
""""MES报工同步脚本 v3
流程: 登录MES -> 生产管理 -> 报工 -> 导出 -> 解析 -> 入库
Excel列结构 (21列, row1=title, row2=headers, row3+=data):
  col0:报工数 col1:报工良品数 col2:报工不良数 col3:报工单位 col4:报工良品率
  col5:生产人员 col6:报工开始时间 col7:报工结束时间 col8:审批状态 col9:审批人
  col10:审批时间 col11:创建人 col12:报工创建时间 col13:工序名称 col14:工单编号
  col15:产品编号 col16:产品名称 col17:关联单据号 col18:设备机台 col19:焊点数量
  col20:出勤人员备注
"""
import os, sys, time, sqlite3
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
SCREENSHOT_DIR = os.path.join(BASE_DIR, "screenshots")
DB_PATH = os.path.join(BASE_DIR, "data", "production.db")

MES_URL = "https://web.ycmes.cn/"
FACTORY_CODE = "606999"
USERNAME = "gs8888"
PASSWORD = "256448"

# Excel column indices (0-based, row 2 = headers)
COL_REPORT_QTY = 0       # 报工数
COL_GOOD_QTY = 1         # 报工良品数
COL_BAD_QTY = 2          # 报工不良数
COL_REPORT_UNIT = 3      # 报工单位
COL_GOOD_RATE = 4        # 报工良品率
COL_OPERATOR = 5         # 生产人员
COL_START_TIME = 6       # 报工开始时间
COL_END_TIME = 7         # 报工结束时间
COL_APPROVE_STATUS = 8   # 审批状态
COL_APPROVER = 9         # 审批人
COL_APPROVE_TIME = 10    # 审批时间
COL_CREATOR = 11         # 创建人
COL_CREATE_TIME = 12     # 报工创建时间
COL_PROCESS_NAME = 13    # 工序名称
COL_ORDER_NO = 14        # 工单编号
COL_PRODUCT_CODE = 15    # 产品编号
COL_PRODUCT_NAME = 16    # 产品名称
COL_RELATED_NO = 17      # 关联单据号
COL_EQUIPMENT = 18       # 设备机台
COL_WELD_COUNT = 19      # 焊点数量
COL_ATTENDANCE_NOTE = 20 # 出勤人员备注


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def screenshot(page, name):
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(SCREENSHOT_DIR, "%s_%s.png" % (name, ts))
    page.screenshot(path=filepath)
    print("  [screenshot] %s" % name)
    return filepath


def login(page):
    print("[1] 打开MES登录页...")
    for attempt in range(3):
        try:
            page.goto(MES_URL, wait_until="networkidle", timeout=60000)
            break
        except Exception:
            if attempt == 2:
                raise
            time.sleep(10)
    time.sleep(2)
    screenshot(page, "01_login_page")

    try:
        tab = page.locator("text=账号登录").first
        if tab.is_visible():
            tab.click()
            time.sleep(1)
    except Exception:
        pass

    fi = page.locator('input[placeholder*="工厂"]').first
    fi.click()
    fi.fill(FACTORY_CODE)
    time.sleep(0.3)

    ui = page.locator('input[placeholder*="用户名"]').first
    ui.click()
    ui.fill(USERNAME)
    time.sleep(0.3)

    pi = page.locator('input[type="password"]').first
    pi.click()
    pi.fill(PASSWORD)
    time.sleep(0.3)

    screenshot(page, "02_filled_form")

    page.locator("button").filter(has_text="登录").first.click()
    time.sleep(5)
    page.wait_for_load_state("networkidle", timeout=30000)
    screenshot(page, "03_after_login")

    if "login" in page.url.lower():
        print("  [ERROR] 登录失败")
        return False
    print("  [OK] 登录成功")
    return True


def navigate_and_export(page):
    print("[2] 导航到报工页面...")
    page.goto(MES_URL.rstrip("/") + "/production/output", wait_until="networkidle", timeout=60000)
    time.sleep(3)
    screenshot(page, "04_report_page")
    print("  [OK] 报工页面已加载")

    print("[3] 点击导出按钮...")
    export_btn = None
    for btn in page.locator("button").all():
        try:
            txt = (btn.text_content() or "").strip()
            if "导出" in txt and "模板" not in txt:
                export_btn = btn
                break
        except Exception:
            pass

    if not export_btn:
        print("  [ERROR] 未找到导出按钮!")
        screenshot(page, "05_no_export_btn")
        return None

    print("  等待下载...")
    with page.expect_download(timeout=300000) as dl_info:
        export_btn.click()
    download = dl_info.value
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(DOWNLOAD_DIR, "work_reports_%s.xlsx" % ts)
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    download.save_as(filepath)
    screenshot(page, "05_downloaded")
    print("  [OK] 下载完成: %s" % os.path.basename(filepath))
    return filepath


def safe_float(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_str(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def format_datetime(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v or "").strip()


def calc_hours(start, end):
    """计算工时(小时)"""
    try:
        fmt = "%Y-%m-%d %H:%M:%S"
        s = datetime.strptime(str(start)[:19], fmt)
        e = datetime.strptime(str(end)[:19], fmt)
        return round((e - s).total_seconds() / 3600, 1)
    except Exception:
        return 0


def parse_and_import(filepath):
    import openpyxl

    print("[4] 解析文件: %s" % os.path.basename(filepath))
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in ws[2]]
    print("  表头列数: %d" % len(headers))
    print("  表头: %s" % headers)

    conn = get_connection()
    c = conn.cursor()

    inserted = skipped = total_rows = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue
        total_rows += 1

        order_no = safe_str(row[COL_ORDER_NO])
        process_name = safe_str(row[COL_PROCESS_NAME])
        operator = safe_str(row[COL_OPERATOR])
        create_time = format_datetime(row[COL_CREATE_TIME])

        if not order_no or not process_name:
            skipped += 1
            continue

        # 去重: 工单编号+工序名称+生产人员+报工创建时间
        c.execute(
            "SELECT 1 FROM work_reports WHERE order_no=? AND process_name=? AND operator=? AND create_time=?",
            (order_no, process_name, operator, create_time))
        if c.fetchone():
            skipped += 1
            continue

        start_time = format_datetime(row[COL_START_TIME])
        end_time = format_datetime(row[COL_END_TIME])
        report_hours = calc_hours(start_time, end_time)

        c.execute(
            "INSERT INTO work_reports (order_no,product_code,product_name,process_name,"
            "report_qty,good_qty,bad_qty,report_unit,good_rate,operator,"
            "start_time,end_time,approve_status,approver,approve_time,"
            "creator,create_time,related_no,equipment,report_hours,weld_count,attendance_note)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (order_no,
             safe_str(row[COL_PRODUCT_CODE]),
             safe_str(row[COL_PRODUCT_NAME]),
             process_name,
             safe_float(row[COL_REPORT_QTY]),
             safe_float(row[COL_GOOD_QTY]),
             safe_float(row[COL_BAD_QTY]),
             safe_str(row[COL_REPORT_UNIT]),
             safe_str(row[COL_GOOD_RATE]),
             operator,
             start_time,
             end_time,
             safe_str(row[COL_APPROVE_STATUS]),
             safe_str(row[COL_APPROVER]),
             format_datetime(row[COL_APPROVE_TIME]),
             safe_str(row[COL_CREATOR]),
             create_time,
             safe_str(row[COL_RELATED_NO]),
             safe_str(row[COL_EQUIPMENT]),
             report_hours,
             safe_float(row[COL_WELD_COUNT]),
             safe_str(row[COL_ATTENDANCE_NOTE])))
        inserted += 1

    conn.commit()
    conn.close()
    wb.close()

    print("  总行数: %d, 新增: %d, 跳过(已存在): %d" % (total_rows, inserted, skipped))
    return inserted


def cleanup_old_files(prefix="work_reports_", keep=3):
    try:
        files = sorted([f for f in os.listdir(DOWNLOAD_DIR)
                        if f.startswith(prefix) and f.endswith(".xlsx")])
        for old in files[:-keep]:
            os.remove(os.path.join(DOWNLOAD_DIR, old))
    except Exception:
        pass


def main():
    from playwright.sync_api import sync_playwright

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    print("=" * 60)
    print("MES报工同步 v3")
    print("时间: %s" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 60)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()

        try:
            if not login(page):
                screenshot(page, "ERROR_login_failed")
                return False

            filepath = navigate_and_export(page)
            if not filepath:
                return False

            inserted = parse_and_import(filepath)
            print("\n[完成] 新增=%d" % inserted)
            screenshot(page, "06_sync_done")

            cleanup_old_files()
            print("IMPORTED:%d" % inserted)
            return True

        except Exception as e:
            print("\n[ERROR] %s" % e)
            import traceback
            traceback.print_exc()
            try:
                screenshot(page, "ERROR_sync_failed")
            except Exception:
                pass
            return False
        finally:
            browser.close()


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
