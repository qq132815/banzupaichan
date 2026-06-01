# -*- coding: utf-8 -*-
""""MES工单同步脚本 v4
流程: 登录MES -> 生产管理 -> 工单 -> 全部 -> 执行中 -> 导出 -> 解析 -> 入库
不合并子工单，每个工单编号作为独立记录
"""
import os, sys, time, sqlite3, re
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
SCREENSHOT_DIR = os.path.join(BASE_DIR, "screenshots")
DB_PATH = os.path.join(BASE_DIR, "data", "production.db")

MES_URL = "https://web.ycmes.cn/"
FACTORY_CODE = "606999"
USERNAME = "gs8888"
PASSWORD = "256448"

# Excel column indices (0-based, row 2 = headers)
COL_TASK_NO = 0        # 工单编号
COL_PRODUCT_CODE = 2   # 产品编号
COL_PRODUCT_NAME = 3   # 产品名称
COL_ORDER_STATUS = 4   # 工单状态 (执行中/已结束)
COL_PROGRESS_BAR = 5   # 工单进度条
COL_PRIORITY = 6       # 优先级
COL_PARENT_ORDER = 7   # 关联单据 (父工单号)
COL_PLAN_START = 9     # 计划开始时间
COL_PLAN_END = 10      # 计划结束时间
COL_PROGRESS_PCT = 12  # 工单进度(%)
COL_PLAN_QTY = 13      # 计划数
COL_DONE_QTY = 14      # 完工数
COL_BAD_QTY = 24       # 不良品数
COL_CREATE_TIME = 22   # 创建时间


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def screenshot(page, name):
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(SCREENSHOT_DIR, "%s_%s.png" % (name, ts))
    page.screenshot(path=filepath)
    print("  [screenshot] %s" % name)
    return filepath


def login(page):
    print("[1] 打开MES登录页...")
    for attempt in range(3):
        try:
            page.goto(MES_URL, wait_until="networkidle", timeout=60000)
            break
        except Exception:
            if attempt == 2:
                raise
            time.sleep(10)
    time.sleep(2)
    screenshot(page, "01_login_page")

    try:
        tab = page.locator("text=账号登录").first
        if tab.is_visible():
            tab.click()
            time.sleep(1)
    except Exception:
        pass

    fi = page.locator('input[placeholder*="工厂"]').first
    fi.click()
    fi.fill(FACTORY_CODE)
    time.sleep(0.3)

    ui = page.locator('input[placeholder*="用户名"]').first
    ui.click()
    ui.fill(USERNAME)
    time.sleep(0.3)

    pi = page.locator('input[type="password"]').first
    pi.click()
    pi.fill(PASSWORD)
    time.sleep(0.3)

    screenshot(page, "02_filled_form")

    page.locator("button").filter(has_text="登录").first.click()
    time.sleep(5)
    page.wait_for_load_state("networkidle", timeout=30000)
    screenshot(page, "03_after_login")

    if "login" in page.url.lower():
        print("  [ERROR] 登录失败")
        return False
    print("  [OK] 登录成功")
    return True


def navigate_and_export(page):
    print("[2] 导航到工单页面...")
    page.goto(MES_URL.rstrip("/") + "/production/work", wait_until="networkidle", timeout=60000)
    time.sleep(3)
    screenshot(page, "04_work_orders_page")
    print("  [OK] 工单页面已加载")

    print("[3] 筛选: 全部 -> 执行中...")
    try:
        all_tab = page.locator("text=全部").first
        all_tab.click()
        time.sleep(2)
        screenshot(page, "05_click_all")
        print("  [OK] 点击全部")
    except Exception as e:
        print("  [WARN] 点击全部失败: %s" % e)

    try:
        page.keyboard.press("Escape")
        time.sleep(0.5)
        exec_tab = page.locator("text=执行中").first
        exec_tab.scroll_into_view_if_needed()
        time.sleep(0.3)
        exec_tab.click(force=True)
        time.sleep(3)
        screenshot(page, "06_click_executing")
        print("  [OK] 点击执行中")
    except Exception as e:
        print("  [WARN] 点击执行中失败: %s" % e)

    print("[4] 点击导出按钮...")
    export_btn = None
    for btn in page.locator("button").all():
        try:
            txt = (btn.text_content() or "").strip()
            if "导出" in txt and "模板" not in txt:
                export_btn = btn
                break
        except Exception:
            pass

    if not export_btn:
        print("  [ERROR] 未找到导出按钮!")
        screenshot(page, "07_no_export_btn")
        return None

    print("  等待下载...")
    with page.expect_download(timeout=300000) as dl_info:
        export_btn.click()
    download = dl_info.value
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(DOWNLOAD_DIR, "work_orders_%s.xlsx" % ts)
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    download.save_as(filepath)
    screenshot(page, "07_downloaded")
    print("  [OK] 下载完成: %s" % os.path.basename(filepath))
    return filepath


def safe_float(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_str(v, default=""):
    if v is None:
        return default
    return str(v).strip()


def parse_progress_pct(s):
    if not s:
        return 0
    m = re.search(r'(\d+)', str(s))
    return float(m.group(1)) if m else 0


def format_datetime(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v or "").strip()


def format_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    if len(s) >= 10:
        return s[:10]
    return s


def parse_and_import(filepath):
    import openpyxl

    print("[5] 解析文件: %s" % os.path.basename(filepath))
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in ws[2]]
    print("  表头列数: %d" % len(headers))

    records = []
    total_rows = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue
        total_rows += 1

        task_no = safe_str(row[COL_TASK_NO])
        if not task_no:
            skipped += 1
            continue

        product_code = safe_str(row[COL_PRODUCT_CODE])
        product_name = safe_str(row[COL_PRODUCT_NAME])
        order_status = safe_str(row[COL_ORDER_STATUS])
        progress_bar = safe_str(row[COL_PROGRESS_BAR])
        priority = safe_str(row[COL_PRIORITY]) or "P2"
        plan_end = format_date(row[COL_PLAN_END])
        plan_qty = safe_float(row[COL_PLAN_QTY])
        done_qty = safe_float(row[COL_DONE_QTY])

        # 状态映射
        if order_status == "已结束":
            status = "completed"
        elif order_status == "执行中":
            status = "in_progress"
        else:
            status = "pending"

        create_time = format_datetime(row[COL_CREATE_TIME])
        
        records.append({
            "order_no": task_no,
            "product_code": product_code,
            "product_name": product_name,
            "quantity": plan_qty,
            "completed_qty": done_qty,
            "due_date": plan_end,
            "priority": priority,
            "status": status,
            "process_progress": progress_bar,
            "create_time": create_time,
        })

    print("  总行数: %d, 跳过: %d, 记录: %d" % (total_rows, skipped, len(records)))

    # 入库
    conn = get_connection()
    c = conn.cursor()
    inserted = updated = 0

    for r in records:
        c.execute("SELECT id FROM work_orders WHERE order_no=?", (r["order_no"],))
        existing = c.fetchone()
        if existing:
            c.execute(
                "UPDATE work_orders SET product_code=?, product_name=?, quantity=?, completed_qty=?,"
                " due_date=?, priority=?, status=?, process_progress=?, source='mes_sync', create_time=? WHERE order_no=?",
                (r["product_code"], r["product_name"], r["quantity"], r["completed_qty"],
                 r["due_date"], r["priority"], r["status"], r["process_progress"], r["create_time"], r["order_no"]))
            updated += 1
        else:
            c.execute(
                "INSERT INTO work_orders (order_no,product_code,product_name,quantity,completed_qty,"
                "due_date,priority,status,process_progress,source,create_time) VALUES (?,?,?,?,?,?,?,?,?,'mes_sync',?)",
                (r["order_no"], r["product_code"], r["product_name"], r["quantity"],
                 r["completed_qty"], r["due_date"], r["priority"], r["status"], r["process_progress"], r["create_time"]))
            inserted += 1

    conn.commit()
    conn.close()
    wb.close()

    print("  [DB] 插入: %d, 更新: %d" % (inserted, updated))
    return inserted, updated


def cleanup_old_files(prefix="work_orders_", keep=3):
    try:
        files = sorted([f for f in os.listdir(DOWNLOAD_DIR)
                        if f.startswith(prefix) and f.endswith(".xlsx")])
        for old in files[:-keep]:
            os.remove(os.path.join(DOWNLOAD_DIR, old))
    except Exception:
        pass


def main():
    from playwright.sync_api import sync_playwright

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    print("=" * 60)
    print("MES工单同步 v4")
    print("时间: %s" % datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 60)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()

        try:
            if not login(page):
                screenshot(page, "ERROR_login_failed")
                return False

            filepath = navigate_and_export(page)
            if not filepath:
                return False

            inserted, updated = parse_and_import(filepath)
            total = inserted + updated
            print("\n[完成] 插入=%d, 更新=%d, 总计=%d" % (inserted, updated, total))
            screenshot(page, "08_sync_done")

            cleanup_old_files()
            print("IMPORTED:%d" % total)
            return True

        except Exception as e:
            print("\n[ERROR] %s" % e)
            import traceback
            traceback.print_exc()
            try:
                screenshot(page, "ERROR_sync_failed")
            except Exception:
                pass
            return False
        finally:
            browser.close()


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
