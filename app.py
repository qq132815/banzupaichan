# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import json
import os
import re
import sys
import threading
import subprocess
from datetime import datetime, timedelta
from functools import wraps

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils.db import (init_database, get_all_teams, get_all_equipments, get_alerts,
                       get_dashboard_stats, get_connection, authenticate_user,
                       get_all_users, create_user, update_user, delete_user, reset_password,
                       get_or_create_daily_plan, get_daily_plan_by_id, get_all_daily_plans,
                       submit_daily_plan, approve_daily_plan, reject_daily_plan, return_daily_plan)
from utils.excel import (import_shipping_plan, import_production_cycles,
                          import_work_orders, import_processes, import_bom)
from utils.calc import (calculate_alerts, back_calculate_semi, calculate_quantity,
                       calculate_production_requirements, publish_requirements, get_published_requirements)

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.jinja_env.auto_reload = True

# ========== MES Sync Scheduler ==========
_sync_scheduler_running = False
_mes_sync_lock = threading.Lock()  # 防止多个Playwright实例并发登录MES冲突

def _run_attendance_sync(date_from, date_to):
    """Run attendance sync in background thread."""
    try:
        from utils.attendance_api import sync_attendance
        count = sync_attendance(date_from, date_to)
        print(f"[attendance] Auto-sync completed: {count} records for {date_from}~{date_to}")
    except Exception as e:
        print(f"[attendance] Auto-sync error: {e}")

def _run_sync_job(sync_type):
    """Run sync job in background thread."""
    log_id = None
    try:
        conn = get_connection()
        c = conn.cursor()
        c.execute("INSERT INTO sync_logs (sync_type, status, created_at, trigger_type) VALUES (?, 'running', datetime('now','localtime'), 'auto')", (sync_type,))
        log_id = c.lastrowid
        conn.commit()
        conn.close()
    except Exception:
        pass
    try:
        script = os.path.join(os.path.dirname(__file__), 'scripts', f'sync_{sync_type}.py')
        result = subprocess.run(
            [sys.executable, '-u', script],
            capture_output=True, text=True, timeout=300,
            cwd=os.path.dirname(__file__)
        )
        output = result.stdout.strip() if result.stdout else ''
        errors = result.stderr.strip() if result.stderr else ''
        if result.returncode == 0:
            count = 0
            for line in output.split('\n'):
                if '总计=' in line:
                    try:
                        count = int(line.split('总计=')[1].strip())
                    except ValueError:
                        pass
                    break
                if '新增:' in line:
                    try:
                        count = int(line.split('新增:')[1].split(',')[0].strip())
                    except ValueError:
                        pass
                    break
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='success', record_count=?, detail=? WHERE id=?", (count, output[-500:], log_id))
            conn.commit()
            conn.close()
            if sync_type == 'reports':
                _recalculate_work_report_efficiency()
                try:
                    _refresh_equipment_status()
                except Exception as eq_err:
                    print(f"[equipment_status] Error: {eq_err}")
            print(f"[{sync_type}] Sync completed: inserted/updated={count}")
        else:
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='error', detail=? WHERE id=?", (errors[-500:], log_id))
            conn.commit()
            conn.close()
            print(f"[{sync_type}] Sync error: {errors[-200:]}")
    except subprocess.TimeoutExpired:
        try:
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='error', detail='同步超时(>300秒)' WHERE id=?", (log_id,))
            conn.commit()
            conn.close()
        except Exception:
            pass
        print(f"[{sync_type}] Sync timeout")
    except Exception as e:
        try:
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='error', detail=? WHERE id=?", (str(e)[:500], log_id))
            conn.commit()
            conn.close()
        except Exception:
            pass
        print(f"[{sync_type}] Sync error: {e}")

def _sync_scheduler():
    """Background scheduler that checks and runs sync jobs."""
    global _sync_scheduler_running
    _sync_scheduler_running = True
    import time
    last_run = {'work_orders': 0, 'reports': 0, 'attendance': 0}

    def _run_synced(sync_type):
        with _mes_sync_lock:
            _run_sync_job(sync_type)
    
    while _sync_scheduler_running:
        try:
            conn = get_connection()
            c = conn.cursor()
            c.execute("SELECT key, value FROM system_settings WHERE key IN ('mes_work_order_interval','mes_report_interval')")
            settings = {r[0]: int(r[1]) for r in c.fetchall() if r[1].isdigit()}
            conn.close()
            
            wo_interval = settings.get('mes_work_order_interval', 0) * 60
            rpt_interval = settings.get('mes_report_interval', 0) * 60
            now = time.time()
            
            if wo_interval > 0 and now - last_run['work_orders'] >= wo_interval:
                last_run['work_orders'] = now
                t = threading.Thread(target=_run_synced, args=('work_orders',))
                t.daemon = True
                t.start()

            if rpt_interval > 0 and now - last_run['reports'] >= rpt_interval:
                last_run['reports'] = now
                t = threading.Thread(target=_run_synced, args=('reports',))
                t.daemon = True
                t.start()
            
            # Attendance auto-sync at 8:30 AM daily
            try:
                conn2 = get_connection()
                c2 = conn2.cursor()
                c2.execute("SELECT value FROM system_settings WHERE key='attendance_auto_sync'")
                auto_row = c2.fetchone()
                auto_sync = int(auto_row[0]) if auto_row and auto_row[0].isdigit() else 0
                conn2.close()
                
                if auto_sync:
                    now_dt = datetime.now()
                    today_830 = now_dt.replace(hour=8, minute=30, second=0, microsecond=0)
                    # Run if it's between 8:30 and 8:35 and hasn't run today at 8:30
                    last_att = last_run.get('attendance', 0)
                    if now_dt >= today_830 and now_dt < today_830.replace(minute=35):
                        if time.time() - last_att > 3600:  # at least 1 hour since last run
                            last_run['attendance'] = time.time()
                            yesterday = (now_dt - timedelta(days=1)).strftime('%Y-%m-%d')
                            t = threading.Thread(target=_run_attendance_sync, args=(yesterday, yesterday))
                            t.daemon = True
                            t.start()
                            print(f"[attendance] Auto-sync triggered for {yesterday}")
            except Exception as e:
                print(f"Attendance auto-sync error: {e}")
            
            # Equipment status auto-refresh at 08:15 and 20:30
            try:
                now_dt = datetime.now()
                refresh_times = [
                    now_dt.replace(hour=8, minute=10, second=0, microsecond=0),
                    now_dt.replace(hour=20, minute=15, second=0, microsecond=0),
                ]
                last_equip = last_run.get('equipment_status', 0)
                for rt in refresh_times:
                    if now_dt >= rt and now_dt < rt.replace(minute=rt.minute + 5):
                        if time.time() - last_equip > 3600:
                            last_run['equipment_status'] = time.time()
                            t = threading.Thread(target=_refresh_equipment_status)
                            t.daemon = True
                            t.start()
                            print("[equipment-status] Auto-refresh triggered at %s" % now_dt.strftime('%H:%M'))
                            break
            except Exception as e:
                print(f"Equipment status refresh error: {e}")

            time.sleep(60)
        except Exception as e:
            print(f"Scheduler error: {e}")
            time.sleep(60)

# Start scheduler in background
_scheduler_thread = threading.Thread(target=_sync_scheduler)
_scheduler_thread.daemon = True
_scheduler_thread.start()


def paginate_query(sql, params, page=1, page_size=50):
    """Execute a paginated query. Returns dict with data, total, page info."""
    conn = get_connection()
    c = conn.cursor()
    # Count total
    count_sql = "SELECT COUNT(*) FROM (" + sql + ")"
    c.execute(count_sql, params)
    total = c.fetchone()[0]
    # Apply pagination
    page = max(1, int(page or 1))
    page_size = min(200, max(1, int(page_size or 50)))
    offset = (page - 1) * page_size
    sql += " LIMIT ? OFFSET ?"
    params.extend([page_size, offset])
    c.execute(sql, params)
    data = [dict(row) for row in c.fetchall()]
    conn.close()
    total_pages = max(1, (total + page_size - 1) // page_size)
    return {"data": data, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}



def _natural_sort_key(code):
    """Generate a sort key for natural sorting (e.g. WG7-2 before WG7-10)."""
    import re
    parts = re.split(r'(\d+)', code or '')
    return [int(p) if p.isdigit() else p.lower() for p in parts]

app.secret_key = 'mes-scheduling-secret-2026'
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "imports")

# ========== Auth Decorators ==========
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'unauthorized'}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({'error': 'forbidden'}), 403
        return f(*args, **kwargs)
    return decorated

def planner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ('planner', 'admin'):
            return jsonify({'error': 'forbidden'}), 403
        return f(*args, **kwargs)
    return decorated

# ========== Login/Logout ==========
@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect('/')
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    user = authenticate_user(data.get('username',''), data.get('password',''))
    if not user:
        return jsonify({'error': '用户名或密码错误'}), 401
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['display_name'] = user['display_name']
    session['role'] = user['role']
    session['team_id'] = user['team_id']
    session['team_name'] = user.get('team_name', '')
    return jsonify({'ok': True, 'user': {
        'id': user['id'], 'username': user['username'],
        'display_name': user['display_name'], 'role': user['role'],
        'team_id': user['team_id'], 'team_name': user.get('team_name', '')
    }})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/me')
@login_required
def api_me():
    return jsonify({
        'id': session['user_id'], 'username': session['username'],
        'display_name': session['display_name'], 'role': session['role'],
        'team_id': session['team_id']
    })

# ========== Page Routes ==========
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/alerts-page')
@login_required
@planner_required
def alerts_page():
    return render_template('alerts.html')

@app.route('/schedule-page')
@login_required
def schedule_page():
    return render_template('schedule.html')

@app.route('/orders-page')
@login_required
@planner_required
def orders_page():
    return render_template('orders.html')

@app.route('/products/processes')
@login_required
def product_processes_page():
    return render_template('product_processes.html')

@app.route('/products/routes')
@login_required
def product_routes_page():
    return render_template('product_routes.html')

@app.route('/products/cycles')
@login_required
def product_cycles_page():
    return render_template('product_cycles.html')







@app.route('/test-select-order')
@login_required
def test_select_order_page():
    return render_template('test_select_order.html')

@app.route('/test-schedule-image')
@login_required
def test_schedule_image_page():
    return render_template('test_schedule_image.html')

@app.route('/test-product-image')
@login_required
def test_product_image_page():
    return render_template('test_product_image.html')

@app.route('/js-test')

@app.route('/test-image-debug')
def test_image_debug_page():
    return render_template('test_image_debug.html')
@login_required
def js_test_page():
    return render_template('js_test.html')

@app.route('/simple-test')
@login_required
def simple_test_page():
    return render_template('simple_test.html')

@app.route('/test-progress')
@login_required
def test_progress_page():
    return render_template('test_progress_modal.html')

@app.route('/products/definitions')
@login_required
def product_definitions_page():
    return render_template('product_definitions.html')

@app.route('/products/bom')
@login_required
def product_bom_page():
    return render_template('product_bom.html')

@app.route('/equipments-page')
@login_required
def equipments_page():
    return render_template('equipments.html')

@app.route('/personnel-page')
@login_required
def personnel_page():
    return render_template('personnel.html')

@app.route('/molds-page')
@login_required
def molds_page():
    return render_template('molds.html')

@app.route('/standard-hours-page')
@login_required
def standard_hours_page():
    return render_template('standard_hours.html')

@app.route('/statistics-page')
@login_required
@admin_required
def statistics_page():
    return render_template('statistics.html')

@app.route('/import-page')
@login_required
def import_page():
    return render_template('import.html')

@app.route('/admin/users')
@login_required
@admin_required
def admin_users_page():
    return render_template('admin_users.html')

@app.route('/admin/settings')
@login_required
@admin_required
def admin_settings_page():
    return render_template('admin_settings.html')

@app.route('/admin/mes-sync')
@login_required
@admin_required
def mes_sync_page():
    return render_template('mes_sync.html')

@app.route('/admin/permissions')
@login_required
@admin_required
def permissions_page():
    return render_template('admin_permissions.html')

@app.route('/work-reports-page')
@login_required
@admin_required
def work_reports_page():
    return render_template('work_reports.html')

@app.route('/reports/efficiency-daily')
@login_required
@admin_required
def report_efficiency_daily_page():
    return render_template('report_efficiency_daily.html')

@app.route('/reports/personal-efficiency')
@login_required
@admin_required
def report_personal_efficiency_page():
    return render_template('report_personal_efficiency.html')

@app.route('/reports/weekly')
@login_required
@admin_required
def report_weekly_page():
    return render_template('report_weekly.html')

@app.route('/reports/monthly')
@login_required
@admin_required
def report_monthly_page():
    return render_template('report_monthly.html')

@app.route('/planner/plans')
@login_required
def planner_plans_page():
    return render_template('planner_plans.html')

@app.route('/api/shipping-plan/template')
@login_required
def api_shipping_plan_template():
    import openpyxl
    from io import BytesIO
    from flask import send_file
    from datetime import date
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发货计划"

    # Row 1: Title (merged)
    ws.merge_cells('A1:AK1')
    ws['A1'] = 'X年X月发货计划（X月X日更新）'

    # Row 2: Headers
    headers = ['客户', '项目', '客户件号', '广升件号', '名称']
    for d in range(1, 32):
        headers.append(str(d) + '\u53f7')
    headers.append('\u5408\u8ba1')
    ws.append(headers)

    # Row 3: Weekday hints (dynamic based on current month)
    weekdays = ['\u661f\u671f\u4e00', '\u661f\u671f\u4e8c', '\u661f\u671f\u4e09', '\u661f\u671f\u56db', '\u661f\u671f\u4e94', '\u661f\u671f\u516d', '\u661f\u671f\u65e5']
    import calendar
    now = date.today()
    year, month = now.year, now.month
    days_in_month = calendar.monthrange(year, month)[1]
    row3 = [None, None, None, None, None]
    for d in range(1, 32):
        try:
            wd = date(year, month, d).weekday()
            row3.append(weekdays[wd])
        except:
            row3.append(None)
    row3.append(None)
    ws.append(row3)

    # Sample data rows
    ws.append(['\u5947\u745e', 'T1J PHEV', 'F26-8108010HV', '03.GS200-95A100A', '\u84b8\u53d1\u5668-\u538b\u7f29\u673a\u548c\u51b7\u51dd\u5668\u7ba1\u8def\u603b\u6210', None, 216, None, None, 216, None, None, 576])
    ws.append([None, None, 'F26-8108020HV', '03.GS200-95A200Z', '\u7535\u6c60\u51b7\u5374-\u538b\u7f29\u673a\u7ba1\u8def\u603b\u6210', None, None, None, None, 0, None, None, 800])

    # Style: make header row bold
    from openpyxl.styles import Font, Alignment
    for cell in ws[2]:
        if cell.value:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="\u53d1\u8d27\u8ba1\u5212\u5bfc\u5165\u6a21\u677f.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/planner/shipping-plan')
@login_required
@planner_required
def planner_shipping_plan_page():
    return render_template('planner_shipping_plan.html')

@app.route('/planner/semi-calc')
@login_required
@planner_required
def planner_semi_calc_page():
    return render_template('planner_semi_calc.html')

@app.route('/planner/publish-history')
@login_required
@planner_required
def planner_publish_history_page():
    return render_template('planner_publish_history.html')

@app.route('/schedule-records')
@login_required
def schedule_records_page():
    return render_template('schedule_records.html')

# ========== Dashboard API ==========
@app.route('/api/dashboard')
@login_required
def api_dashboard():
    stats = get_dashboard_stats()
    return jsonify(stats)

# ========== Team API ==========
@app.route('/api/teams')
@login_required
def api_teams():
    return jsonify(get_all_teams())

# ========== Equipment API ==========
@app.route('/api/equipments')
@login_required
def api_equipments():
    team_id = request.args.get('team_id', type=int)
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM equipments WHERE 1=1"
    params = []
    if team_id:
        sql += " AND team_id=?"
        params.append(team_id)
    if q:
        like = "%" + q + "%"
        sql += " AND (equipment_code LIKE ? OR equipment_name LIKE ? OR location LIKE ?)"
        params.extend([like, like, like])
    sql += " ORDER BY team_id"
    result = paginate_query(sql, params, page, page_size)
    result['data'].sort(key=lambda x: _natural_sort_key(x.get('equipment_code', '')))
    return jsonify(result)

@app.route('/api/equipments', methods=['POST'])
@login_required
def api_equipment_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO equipments (equipment_code, equipment_name, team_id, equipment_type, status, capacity_per_hour, location) VALUES (?, ?, ?, ?, ?, ?, ?)",
                  (data['equipment_code'], data['equipment_name'], data['team_id'],
                   data.get('equipment_type', 'auto'), data.get('status', 'normal'),
                   data.get('capacity_per_hour', 0), data.get('location', '')))
        conn.commit()
        return jsonify({'id': c.lastrowid, 'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/equipments/<int:eid>', methods=['PUT'])
@login_required
def api_equipment_update(eid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE equipments SET equipment_name=?, team_id=?, equipment_type=?, status=?, capacity_per_hour=?, location=? WHERE id=?",
              (data['equipment_name'], data['team_id'], data.get('equipment_type'),
               data.get('status', 'normal'), data.get('capacity_per_hour', 0),
               data.get('location', ''), eid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/equipments/<int:eid>', methods=['DELETE'])
@login_required
def api_equipment_delete(eid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM equipments WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ========== Product Definition API ==========
@app.route('/api/product-definitions')
@login_required
def api_product_definitions():
    q = request.args.get('q', '').strip()
    exact_product_code = request.args.get('exact_product_code', '').strip()
    category = request.args.get('category', '').strip()
    status = request.args.get('status', '').strip()
    product_type = request.args.get('product_type', '').strip()
    customer = request.args.get('customer', '').strip()
    field = request.args.get('field', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    
    # 如果请求特定字段的唯一值（用于筛选下拉框）
    if field:
        conn = get_connection()
        c = conn.cursor()
        if field == 'customer':
            c.execute("SELECT DISTINCT customer FROM products WHERE customer IS NOT NULL AND customer != '' ORDER BY customer")
            values = [row['customer'] for row in c.fetchall()]
        elif field == 'product_type':
            c.execute("SELECT DISTINCT product_type FROM products WHERE product_type IS NOT NULL AND product_type != '' ORDER BY product_type")
            values = [row['product_type'] for row in c.fetchall()]
        else:
            values = []
        conn.close()
        return jsonify(values)
    
    sql = "SELECT * FROM products WHERE 1=1"
    params = []
    
    # 优先使用精确匹配
    if exact_product_code:
        sql += " AND product_code=?"
        params.append(exact_product_code)
    elif q:
        like = "%" + q + "%"
        sql += " AND (product_code LIKE ? OR product_name LIKE ? OR specifications LIKE ? OR description LIKE ?)"
        params.extend([like, like, like, like])
    
    if category:
        sql += " AND category=?"
        params.append(category)
    if status:
        sql += " AND status=?"
        params.append(status)
    if product_type:
        sql += " AND product_type=?"
        params.append(product_type)
    if customer:
        sql += " AND customer=?"
        params.append(customer)
    
    sql += " ORDER BY product_code"
    result = paginate_query(sql, params, page, page_size)
    return jsonify(result)

@app.route('/api/product-definitions', methods=['POST'])
@login_required
def api_product_definition_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    try:
        # 下载图片（如果有URL）
        image_path = None
        if data.get('image_url'):
            image_path = download_product_image(data['image_url'], data['product_code'])
        
        c.execute("""INSERT INTO products 
            (product_code, product_name, product_type, specifications, unit, route_code,
             safety_stock, stock_qty, source, customer, basket_capacity, image_url, image_path, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (data['product_code'], data['product_name'], data.get('product_type', ''),
             data.get('specifications', ''), data.get('unit', ''), data.get('route_code', ''),
             data.get('safety_stock', 0), data.get('stock_qty', 0), data.get('source', ''),
             data.get('customer', ''), data.get('basket_capacity', 0), data.get('image_url', ''),
             image_path or '', data.get('description', '')))
        conn.commit()
        return jsonify({'id': c.lastrowid, 'ok': True, 'image_path': image_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/product-definitions/<int:pid>', methods=['PUT'])
@login_required
def api_product_definition_update(pid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    try:
        # 下载图片（如果有新URL且与原URL不同）
        image_path = None
        if data.get('image_url'):
            c.execute("SELECT image_url FROM products WHERE id=?", (pid,))
            old_url = c.fetchone()
            if old_url and old_url['image_url'] != data['image_url']:
                image_path = download_product_image(data['image_url'], data['product_code'])
            elif not old_url or not old_url['image_url']:
                image_path = download_product_image(data['image_url'], data['product_code'])
        
        if image_path:
            c.execute("""UPDATE products SET 
                product_name=?, product_type=?, specifications=?, unit=?, route_code=?,
                safety_stock=?, stock_qty=?, source=?, customer=?, basket_capacity=?,
                image_url=?, image_path=?, description=?, updated_at=datetime('now','localtime')
                WHERE id=?""",
                (data['product_name'], data.get('product_type', ''), data.get('specifications', ''),
                 data.get('unit', ''), data.get('route_code', ''), data.get('safety_stock', 0),
                 data.get('stock_qty', 0), data.get('source', ''), data.get('customer', ''),
                 data.get('basket_capacity', 0), data.get('image_url', ''), image_path,
                 data.get('description', ''), pid))
        else:
            c.execute("""UPDATE products SET 
                product_name=?, product_type=?, specifications=?, unit=?, route_code=?,
                safety_stock=?, stock_qty=?, source=?, customer=?, basket_capacity=?,
                image_url=?, description=?, updated_at=datetime('now','localtime')
                WHERE id=?""",
                (data['product_name'], data.get('product_type', ''), data.get('specifications', ''),
                 data.get('unit', ''), data.get('route_code', ''), data.get('safety_stock', 0),
                 data.get('stock_qty', 0), data.get('source', ''), data.get('customer', ''),
                 data.get('basket_capacity', 0), data.get('image_url', ''),
                 data.get('description', ''), pid))
        
        conn.commit()
        return jsonify({'ok': True, 'image_path': image_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/product-definitions/<int:pid>', methods=['DELETE'])
@login_required
def api_product_definition_delete(pid):
    conn = get_connection()
    c = conn.cursor()
    try:
        # 删除图片文件
        c.execute("SELECT image_path FROM products WHERE id=?", (pid,))
        result = c.fetchone()
        if result and result['image_path']:
            image_file = os.path.join(os.path.dirname(__file__), result['image_path'].lstrip('/'))
            if os.path.exists(image_file):
                os.remove(image_file)
        
        c.execute("DELETE FROM products WHERE id=?", (pid,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/product-definitions/categories')
@login_required
def api_product_categories():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT category FROM products WHERE category IS NOT NULL AND category != '' ORDER BY category")
    categories = [row['category'] for row in c.fetchall()]
    conn.close()
    return jsonify(categories)


@app.route('/api/product-definitions/download-all-images', methods=['POST'])
@login_required
def api_download_all_images():
    """批量下载所有产品图片"""
    try:
        import threading
        
        # 进度文件路径
        progress_file = os.path.join(os.path.dirname(__file__), 'data', 'download_progress.json')
        
        # 初始化进度文件
        with open(progress_file, 'w', encoding='utf-8') as f:
            json.dump({
                'status': 'running',
                'total': 0,
                'current': 0,
                'success': 0,
                'fail': 0,
                'skip': 0,
                'message': '正在准备下载...',
                'start_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }, f, ensure_ascii=False)
        
        def download_task():
            try:
                # 导入并执行下载脚本
                import sys
                script_path = os.path.join(os.path.dirname(__file__), 'scripts', 'download_all_images_progress.py')
                
                # 使用subprocess运行，并传入进度文件路径
                env = os.environ.copy()
                env['PROGRESS_FILE'] = progress_file
                
                result = subprocess.run(
                    [sys.executable, script_path], 
                    capture_output=True, text=True, timeout=1800,
                    env=env
                )
                
                # 更新进度文件为完成状态
                with open(progress_file, 'r', encoding='utf-8') as f:
                    progress = json.load(f)
                
                progress['status'] = 'completed'
                progress['message'] = f'下载完成: 成功{progress["success"]}个, 失败{progress["fail"]}个, 跳过{progress["skip"]}个'
                progress['end_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump(progress, f, ensure_ascii=False)
                
                print(f"图片下载完成: {result.stdout[-500:]}")
            except Exception as e:
                # 更新进度文件为错误状态
                with open(progress_file, 'r', encoding='utf-8') as f:
                    progress = json.load(f)
                
                progress['status'] = 'error'
                progress['message'] = f'下载失败: {str(e)}'
                
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump(progress, f, ensure_ascii=False)
                
                print(f"图片下载失败: {e}")
        
        # 在后台线程执行下载
        thread = threading.Thread(target=download_task)
        thread.daemon = True
        thread.start()
        
        return jsonify({'ok': True, 'message': '图片下载任务已启动'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/product-definitions/download-progress')
@login_required
def api_download_progress():
    """查询图片下载进度"""
    try:
        progress_file = os.path.join(os.path.dirname(__file__), 'data', 'download_progress.json')
        
        if os.path.exists(progress_file):
            with open(progress_file, 'r', encoding='utf-8') as f:
                progress = json.load(f)
            return jsonify(progress)
        else:
            return jsonify({
                'status': 'idle',
                'total': 0,
                'current': 0,
                'success': 0,
                'fail': 0,
                'skip': 0,
                'message': '未开始下载'
            })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/product-definitions/download-image', methods=['POST'])
@login_required
def api_download_product_image():
    """下载产品图片并保存到本地"""
    data = request.json
    image_url = data.get('image_url')
    product_code = data.get('product_code')
    
    if not image_url or not product_code:
        return jsonify({'error': '缺少参数'}), 400
    
    try:
        image_path = download_product_image(image_url, product_code)
        if image_path:
            return jsonify({'ok': True, 'image_path': image_path})
        else:
            return jsonify({'error': '图片下载失败'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def download_product_image(image_url, product_code):
    """下载产品图片并保存到本地"""
    import requests
    from urllib.parse import urlparse
    
    try:
        # 创建图片目录
        images_dir = os.path.join(os.path.dirname(__file__), 'static', 'images', 'products')
        os.makedirs(images_dir, exist_ok=True)
        
        # 生成文件名
        url_path = urlparse(image_url).path
        ext = os.path.splitext(url_path)[1] or '.jpg'
        filename = f"{product_code}{ext}"
        filepath = os.path.join(images_dir, filename)
        
        # 下载图片
        response = requests.get(image_url, timeout=30)
        response.raise_for_status()
        
        # 保存图片
        with open(filepath, 'wb') as f:
            f.write(response.content)
        
        # 返回相对路径
        return f"/static/images/products/{filename}"
    except Exception as e:
        print(f"下载图片失败: {e}")
        return None

# ========== Alert API ==========
# ========== Alert API ==========
@app.route('/api/alerts')
@login_required
def api_alerts():
    level = request.args.get('level')
    return jsonify(get_alerts(level))

@app.route('/api/alerts/calculate', methods=['POST'])
@login_required
def api_calc_alerts():
    count = calculate_alerts()
    return jsonify({'count': count})

@app.route('/api/back-calculate', methods=['POST'])
@login_required
@planner_required
def api_back_calc():
    count = back_calculate_semi()
    return jsonify({'new_orders': count})

# ========== Import API ==========
@app.route('/api/import/<data_type>', methods=['POST'])
@login_required
def api_import(data_type):
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    f = request.files['file']
    path = os.path.join(app.config["UPLOAD_FOLDER"], f.filename)
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    f.save(path)
    try:
        if data_type == 'shipping':
            plan_month = request.form.get('plan_month', '').strip()
            count = import_shipping_plan(path, plan_month=plan_month or None)
        elif data_type == 'cycles':
            count = import_production_cycles(path)
        elif data_type == 'orders':
            count = import_work_orders(path)
        elif data_type == 'processes':
            count = import_processes(path)
        elif data_type == 'bom':
            count = import_bom(path)
        elif data_type == 'process_routes':
            from utils.excel import import_process_routes
            count = import_process_routes(path)
        elif data_type == 'product_definitions':
            from utils.excel import import_product_definitions
            count = import_product_definitions(path)
        elif data_type == 'equipment':
            from utils.excel import import_equipment
            count = import_equipment(path)
        else:
            return jsonify({'error': 'unknown type'}), 400
        return jsonify({'count': count, 'success': True})
    except ValueError as e:
        return jsonify({'error': str(e), 'success': False})
    except Exception as e:
        return jsonify({'error': '导入失败: ' + str(e), 'success': False})

# ========== Daily Plan API ==========
@app.route('/api/daily-plan')
@login_required
def api_daily_plan():
    team_id = request.args.get('team_id', type=int)
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    if not team_id:
        team_id = session.get('team_id')
    if not team_id:
        return jsonify({'error': 'no team'}), 400
    plan = get_or_create_daily_plan(team_id, date)
    return jsonify(plan)

@app.route('/api/daily-plans')
@login_required
def api_daily_plans():
    date = request.args.get('date')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    team_id = request.args.get('team_id', type=int)
    # Team users can only see their own team's plans
    if session.get('role') == 'team':
        team_id = session.get('team_id')
    return jsonify(get_all_daily_plans(date, team_id, date_from, date_to))

@app.route('/api/daily-plan/<int:pid>/submit', methods=['POST'])
@login_required
def api_submit_plan(pid):
    plan = get_daily_plan_by_id(pid)
    if not plan:
        return jsonify({'error': 'not found'}), 404
    if plan['status'] not in ('draft', 'returned'):
        return jsonify({'error': '当前状态不允许提交'}), 400
    # Check team permission
    if session.get('role') == 'team' and plan['team_id'] != session.get('team_id'):
        return jsonify({'error': '只能提交自己班组的计划'}), 403
    submit_daily_plan(pid, session['user_id'])
    return jsonify({'ok': True})

@app.route('/api/daily-plan/<int:pid>/approve', methods=['POST'])
@login_required
@planner_required
def api_approve_plan(pid):
    approve_daily_plan(pid, session['user_id'])
    return jsonify({'ok': True})

@app.route('/api/daily-plan/<int:pid>/reject', methods=['POST'])
@login_required
@planner_required
def api_reject_plan(pid):
    data = request.json or {}
    reject_daily_plan(pid, session['user_id'], data.get('reason', ''))
    return jsonify({'ok': True})

@app.route('/api/daily-plan/<int:pid>', methods=['DELETE'])
@login_required
def api_delete_daily_plan(pid):
    plan = get_daily_plan_by_id(pid)
    if not plan:
        return jsonify({'error': 'not found'}), 404
    if plan['status'] not in ('draft', 'returned'):
        return jsonify({'error': '只有草稿状态可以删除'}), 400
    if session.get('role') == 'team' and plan['team_id'] != session.get('team_id'):
        return jsonify({'error': '只能删除自己班组的计划'}), 403
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM schedules WHERE daily_plan_id=?", (pid,))
    c.execute("DELETE FROM daily_plans WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/daily-plan/<int:pid>/return', methods=['POST'])
@login_required
def api_return_plan(pid):
    data = request.json or {}
    plan = get_daily_plan_by_id(pid)
    if not plan:
        return jsonify({'error': 'not found'}), 404
    # Planner can return any; team can request return on their own
    if session.get('role') == 'team' and plan['team_id'] != session.get('team_id'):
        return jsonify({'error': '无权操作'}), 403
    return_daily_plan(pid, session['user_id'], data.get('reason', ''))
    return jsonify({'ok': True})
# ========== Daily Plan Check API ==========
@app.route('/api/daily-plan/<int:pid>/check')
@login_required
def api_check_daily_plan(pid):
    """Auto-review: for each published requirement, check:
    1. Work order remaining >= required_qty? If yes, requirement is covered, skip.
    2. If not, check schedule qty on that date >= required_qty.
    3. Also check if schedule date > required_date (late delivery).
    """
    conn = get_connection()
    c = conn.cursor()

    # 1. Get the plan
    c.execute("SELECT dp.*, t.name as team_name FROM daily_plans dp LEFT JOIN teams t ON dp.team_id=t.id WHERE dp.id=?", (pid,))
    plan_row = c.fetchone()
    if not plan_row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    plan = dict(plan_row)
    plan_date = plan['plan_date']
    team_id = plan['team_id']
    team_name = plan.get('team_name', '')

    # 2. Get all schedules for this plan (for gantt)
    c.execute("""SELECT s.*, e.equipment_name, e.equipment_code
        FROM schedules s LEFT JOIN equipments e ON s.equipment_id=e.id
        WHERE s.daily_plan_id=? ORDER BY s.start_time""", (pid,))
    schedules = [dict(r) for r in c.fetchall()]

    # 3. Get published production requirements for this team
    c.execute("""SELECT product_code, product_name, required_date, required_quantity, team_name
        FROM production_requirements WHERE status='published' ORDER BY required_date""")
    all_reqs = c.fetchall()
    requirements = []
    for row in all_reqs:
        r = dict(row)
        tn = r.get('team_name', '') or ''
        if team_name and team_name not in tn:
            continue
        requirements.append(r)

    # 4. Get work order remaining
    c.execute("""SELECT product_code, product_name, SUM(COALESCE(quantity,0) - COALESCE(completed_qty,0)) as remaining
        FROM work_orders WHERE status != 'completed' GROUP BY product_code""")
    wo_remaining = {}
    for row in c.fetchall():
        pc = row[0] or ''
        pn = row[1] or ''
        rem = row[2] or 0
        wo_remaining[pc] = rem
        if pn:
            wo_remaining[pn] = rem

    conn.close()

    # 5. Build schedule qty index: product_code|schedule_date -> total_qty
    sched_qty = {}
    for s in schedules:
        pc = s.get('product_code', '')
        sd = s.get('schedule_date', '')
        if not pc:
            continue
        key = pc + '|' + sd
        sched_qty[key] = sched_qty.get(key, 0) + (s.get('quantity', 0) or 0)

    # 6. Run checks per requirement
    issues = []
    checked_products = set()
    for req in requirements:
        pc = req['product_code']
        pn = req.get('product_name', '') or pc
        req_date = req['required_date']
        req_qty = req['required_quantity'] or 0
        if req_qty <= 0:
            continue
        checked_products.add(pc)

        # Step A: Check work order remaining
        # wo_rem = remaining qty (quantity - completed_qty)
        # wo_rem == 0 => fully done, requirement satisfied -> skip
        # wo_rem > 0 but < req_qty => mostly done, actual need = wo_rem
        # wo_rem >= req_qty => nothing done, actual need = req_qty
        wo_rem = wo_remaining.get(pc, 0)
        if wo_rem <= 0:
            continue

        actual_need = min(wo_rem, req_qty)

        # Step B: Check schedule qty on required date covers actual need
        sched_key = pc + '|' + req_date
        scheduled = sched_qty.get(sched_key, 0)
        gap = actual_need - scheduled

        if gap > 0:
            issues.append({
                'type': 'qty_gap', 'level': 'error',
                'product_code': pc, 'product_name': pn,
                'required_date': req_date, 'required_qty': req_qty,
                'wo_remaining': wo_rem, 'actual_need': actual_need,
                'scheduled_qty': scheduled, 'gap': gap,
                'message': pn + ' ' + req_date + ': 需求' + str(int(req_qty)) + '，工单未完成' + str(int(wo_rem)) + '，排班' + str(int(scheduled)) + '，缺口' + str(int(gap))
            })

        # Step C: Check late schedules (after required date)
        for key, qty in sched_qty.items():
            if key.startswith(pc + '|'):
                sched_date = key.split('|')[1]
                if sched_date > req_date and qty > 0:
                    issues.append({
                        'type': 'late_schedule', 'level': 'warning',
                        'product_code': pc, 'product_name': pn,
                        'required_date': req_date, 'schedule_date': sched_date,
                        'scheduled_qty': qty,
                        'message': pn + ': ' + sched_date + '排班(' + str(int(qty)) + ')晚于需求' + req_date
                    })

    # Info: products in schedule but no requirement
    for s in schedules:
        pc = s.get('product_code', '')
        if pc and pc not in checked_products and pc not in [r['product_code'] for r in requirements]:
            if not any(i['product_code'] == pc and i['type'] == 'no_requirement' for i in issues):
                issues.append({
                    'type': 'no_requirement', 'level': 'info',
                    'product_code': pc, 'product_name': s.get('product_name', '') or pc,
                    'message': pc + ': 无对应生产需求'
                })

    # Verdict
    err = sum(1 for i in issues if i['level'] == 'error')
    warn = sum(1 for i in issues if i['level'] == 'warning')
    if err > 0:
        verdict, verdict_text = 'error', str(err) + '项缺口问题，建议关注'
    elif warn > 0:
        verdict, verdict_text = 'warning', str(warn) + '项日期问题'
    else:
        verdict, verdict_text = 'pass', '审核通过'

    return jsonify({
        'plan': plan, 'schedules': schedules,
        'requirements': requirements, 'wo_remaining': wo_remaining,
        'issues': issues, 'verdict': verdict, 'verdict_text': verdict_text
    })




# ========== Requirement Check API (for schedule page) ==========
@app.route('/api/requirement-check')
@login_required
def api_requirement_check():
    """Check published requirements vs work orders + schedules for a team and date."""
    team_id = request.args.get('team_id', type=int)
    date = request.args.get('date', '').strip()
    if not team_id or not date:
        return jsonify({'error': 'team_id and date required'}), 400

    conn = get_connection()
    c = conn.cursor()

    c.execute("SELECT name FROM teams WHERE id=?", (team_id,))
    team_row = c.fetchone()
    team_name = team_row[0] if team_row else ''

    c.execute("""SELECT product_code, product_name, required_date, required_quantity, team_name
        FROM production_requirements WHERE status='published' ORDER BY required_date""")
    requirements = []
    for row in c.fetchall():
        r = dict(row)
        tn = r.get('team_name', '') or ''
        if team_name and team_name not in tn:
            continue
        requirements.append(r)

    c.execute("""SELECT product_code, product_name, SUM(COALESCE(quantity,0) - COALESCE(completed_qty,0)) as remaining
        FROM work_orders WHERE status != 'completed' GROUP BY product_code""")
    wo_remaining = {}
    for row in c.fetchall():
        pc, pn, rem = row[0] or '', row[1] or '', row[2] or 0
        wo_remaining[pc] = rem
        if pn:
            wo_remaining[pn] = rem

    c.execute("""SELECT s.product_code, SUM(s.quantity) as total_qty
        FROM schedules s WHERE s.team_id=? AND s.schedule_date=?
        GROUP BY s.product_code""", (team_id, date))
    sched_qty = {}
    for row in c.fetchall():
        sched_qty[row[0]] = row[1] or 0

    conn.close()

    issues = []
    for req in requirements:
        pc = req['product_code']
        pn = req.get('product_name', '') or pc
        req_date = req['required_date']
        req_qty = req['required_quantity'] or 0
        if req_qty <= 0:
            continue
        wo_rem = wo_remaining.get(pc, 0)
        if wo_rem <= 0:
            continue
        actual_need = min(wo_rem, req_qty)
        scheduled = sched_qty.get(pc, 0)
        gap = actual_need - scheduled
        if gap > 0:
            issues.append({
                'type': 'qty_gap', 'level': 'error',
                'product_code': pc, 'product_name': pn,
                'required_date': req_date, 'required_qty': req_qty,
                'wo_remaining': wo_rem, 'actual_need': actual_need,
                'scheduled_qty': scheduled, 'gap': gap,
                'message': pn + ' ' + req_date + ': 需求' + str(int(req_qty)) + '，工单未完成' + str(int(wo_rem)) + '，排班' + str(int(scheduled)) + '，缺口' + str(int(gap))
            })

    for pc, qty in sched_qty.items():
        for req in requirements:
            if req['product_code'] == pc and date > req['required_date'] and qty > 0:
                issues.append({
                    'type': 'late_schedule', 'level': 'warning',
                    'product_code': pc, 'product_name': req.get('product_name', '') or pc,
                    'required_date': req['required_date'], 'schedule_date': date,
                    'scheduled_qty': qty,
                    'message': (req.get('product_name','') or pc) + ': ' + date + '排班(' + str(int(qty)) + ')晚于需求' + req['required_date']
                })

    return jsonify({
        'requirements': requirements, 'wo_remaining': wo_remaining,
        'sched_qty': sched_qty, 'issues': issues,
        'team_name': team_name, 'date': date
    })

# ========== Schedule API ==========
@app.route('/api/schedule', methods=['GET'])
@login_required
def api_schedule_list():
    conn = get_connection()
    c = conn.cursor()
    team_id = request.args.get('team_id', type=int)
    date = request.args.get('date')
    q = "SELECT s.*, e.equipment_name, e.equipment_code, w.product_name FROM schedules s LEFT JOIN equipments e ON s.equipment_id=e.id LEFT JOIN work_orders w ON s.work_order_no=w.order_no WHERE 1=1"
    params = []
    if team_id:
        q += " AND s.team_id=?"
        params.append(team_id)
    if date:
        q += " AND s.schedule_date=?"
        params.append(date)
    q += " ORDER BY s.start_time"
    c.execute(q, params)
    r = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(r)

@app.route('/api/schedule', methods=['POST'])
@login_required
def api_schedule_create():
    data = request.json
    # Check daily plan status
    plan_date = data.get('schedule_date', datetime.now().strftime('%Y-%m-%d'))
    team_id = data.get('team_id', session.get('team_id'))
    plan = get_or_create_daily_plan(team_id, plan_date)
    if plan['status'] == 'approved':
        return jsonify({'error': '该日计划已审批通过，不能修改。请联系计划员退回。'}), 403
    # Team can only edit their own
    if session.get('role') == 'team' and team_id != session.get('team_id'):
        return jsonify({'error': '只能排自己班组的计划'}), 403

    conn = get_connection()
    c = conn.cursor()
    eq_id = data.get('equipment_id')
    start = data.get('start_time')
    end = data.get('end_time')
    if eq_id and plan_date and start and end:
        c.execute("SELECT id, work_order_no, start_time, end_time FROM schedules WHERE equipment_id=? AND schedule_date=? AND task_status != 'cancelled'",
                  (eq_id, plan_date))
        for ex in c.fetchall():
            if start < (ex["end_time"] or "") and end > (ex["start_time"] or ""):
                conn.close()
                return jsonify({'error': 'conflict', 'conflict_with': ex["work_order_no"],
                                'conflict_time': str(ex["start_time"]) + "-" + str(ex["end_time"])}), 409
    c.execute("INSERT INTO schedules (daily_plan_id, equipment_id, work_order_no, product_code, process_code, process_name, schedule_date, start_time, end_time, quantity, hours, capacity_per_hour, is_overtime, team_id, task_status, priority, operator) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
              (plan['id'], eq_id, data.get('work_order_no'), data.get('product_code'),
               data.get('process_code'), data.get('process_name'), plan_date, start, end,
               data.get('quantity', 0), data.get('hours', 0),
               data.get('capacity_per_hour', 0), data.get('is_overtime', 0),
               team_id, 'planned', data.get('priority', 'P2'),
               data.get('operator', '')))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/schedule/<int:sid>', methods=['PUT'])
@login_required
def api_schedule_update(sid):
    data = request.json
    plan_date = data.get('schedule_date')
    team_id = data.get('team_id', session.get('team_id'))
    if plan_date:
        plan = get_or_create_daily_plan(team_id, plan_date)
        if plan['status'] == 'approved':
            return jsonify({'error': '该日计划已审批通过，不能修改'}), 403

    conn = get_connection()
    c = conn.cursor()
    eq_id = data.get('equipment_id')
    start = data.get('start_time')
    end = data.get('end_time')
    if eq_id and plan_date and start and end:
        c.execute("SELECT id, work_order_no, start_time, end_time FROM schedules WHERE equipment_id=? AND schedule_date=? AND id != ? AND task_status != 'cancelled'",
                  (eq_id, plan_date, sid))
        for ex in c.fetchall():
            if start < (ex["end_time"] or "") and end > (ex["start_time"] or ""):
                conn.close()
                return jsonify({'error': 'conflict', 'conflict_with': ex["work_order_no"]}), 409
    c.execute("UPDATE schedules SET equipment_id=?, start_time=?, end_time=?, quantity=?, hours=?, capacity_per_hour=?, is_overtime=?, operator=?, process_code=?, process_name=?, priority=?, work_order_no=?, schedule_date=?, team_id=?, product_code=? WHERE id=?",
              (eq_id, start, end, data.get('quantity', 0), data.get('hours', 0),
               data.get('capacity_per_hour', 0), data.get('is_overtime', 0),
               data.get('operator', ''), data.get('process_code'),
               data.get('process_name'), data.get('priority', 'P2'),
               data.get('work_order_no'), plan_date, team_id, data.get('product_code'), sid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/schedule/<int:sid>', methods=['DELETE'])
@login_required
def api_schedule_delete(sid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM schedules WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Order API ==========
@app.route('/api/orders')
@login_required
def api_orders():
    status = request.args.get('status')
    keyword = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM work_orders WHERE 1=1"
    params = []
    if status:
        sql += " AND status=?"
        params.append(status)
    if keyword:
        like = "%" + keyword + "%"
        sql += " AND (order_no LIKE ? OR product_code LIKE ? OR product_name LIKE ?)"
        params.extend([like, like, like])
    sql += " ORDER BY create_time DESC"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/orders', methods=['POST'])
@login_required
def api_order_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO work_orders (order_no, product_code, product_name, quantity, completed_qty, due_date, priority, status, source) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
              (data['order_no'], data['product_code'], data.get('product_name', ''),
               data.get('quantity', 0), data.get('completed_qty', 0),
               data.get('due_date'), data.get('priority', 'P2'),
               data.get('status', 'pending'), data.get('source', '')))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/orders/<int:oid>', methods=['PUT'])
@login_required
def api_order_update(oid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE work_orders SET order_no=?, product_code=?, product_name=?, quantity=?, completed_qty=?, due_date=?, priority=?, status=?, source=? WHERE id=?",
              (data['order_no'], data['product_code'], data.get('product_name', ''),
               data.get('quantity', 0), data.get('completed_qty', 0),
               data.get('due_date'), data.get('priority', 'P2'),
               data.get('status', 'pending'), data.get('source', ''), oid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/orders/<int:oid>', methods=['DELETE'])
@login_required
def api_order_delete(oid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM work_orders WHERE id=?", (oid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/orders/clear', methods=['POST'])
@login_required
@planner_required
def api_orders_clear():
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM work_orders")
    deleted = c.rowcount
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'deleted': deleted})

@app.route('/api/search-orders')
@login_required
def api_search_orders():
    keyword = request.args.get('q', '')
    team_name = session.get('team_name', '')
    conn = get_connection()
    c = conn.cursor()
    if team_name:
        # Get process names belonging to this team
        c.execute("SELECT DISTINCT process_name FROM processes WHERE team_name LIKE ?", ('%' + team_name + '%',))
        team_procs = [r[0] for r in c.fetchall()]
        # Find product_codes whose process_routes contain any of these processes
        matching_products = set()
        c.execute("SELECT product_code, process_list FROM process_routes")
        for r in c.fetchall():
            pc = r[0]
            pl = r[1] or ''
            for tp in team_procs:
                if tp in pl:
                    matching_products.add(pc)
                    break
        if not matching_products:
            conn.close()
            return jsonify([])
        placeholders = ','.join(['?' for _ in matching_products])
        params = ['%' + keyword + '%', '%' + keyword + '%', '%' + keyword + '%'] + list(matching_products)
        c.execute(f"SELECT * FROM work_orders WHERE (order_no LIKE ? OR product_code LIKE ? OR product_name LIKE ?) AND product_code IN ({placeholders}) LIMIT 20", params)
    else:
        c.execute("SELECT * FROM work_orders WHERE order_no LIKE ? OR product_code LIKE ? OR product_name LIKE ? LIMIT 20",
                  ('%' + keyword + '%', '%' + keyword + '%', '%' + keyword + '%'))
    orders = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(orders)

# ========== Product Process API ==========
@app.route('/api/product-processes')
@login_required
def api_product_processes():
    product_code = request.args.get('product_code', '')
    team_name = session.get('team_name', '')
    conn = get_connection()
    c = conn.cursor()
    # Match on product_code column for precise lookup
    c.execute("SELECT route_code, route_name, process_list FROM process_routes WHERE product_code = ?", (product_code,))
    routes = c.fetchall()
    if not routes:
        # Fallback: try route_code exact match
        c.execute("SELECT route_code, route_name, process_list FROM process_routes WHERE route_code = ?", (product_code,))
        routes = c.fetchall()
    result = []
    seen = set()
    for r in routes:
        route_code = r[0]
        route_name = r[1]
        process_list = r[2]
        if process_list:
            for proc_name in process_list.split(','):
                proc_name = proc_name.strip()
                if not proc_name:
                    continue
                key = route_code + '|' + proc_name
                if key in seen:
                    continue
                seen.add(key)
                c.execute("SELECT process_code, team_name FROM processes WHERE process_name=?", (proc_name,))
                proc = c.fetchone()
                if proc:
                    # Filter by team if user is a team member
                    if team_name and proc[1] and team_name not in proc[1]:
                        continue
                    result.append({
                        'route_code': route_code, 'route_name': route_name,
                        'process_code': proc[0], 'process_name': proc_name,
                        'team_name': proc[1]
                    })
    conn.close()
    return jsonify(result)

@app.route('/api/process-equipment')
@login_required
def api_process_equipment():
    process_code = request.args.get('process_code', '')
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT e.id, e.equipment_code, e.equipment_name, e.equipment_type, e.capacity_per_hour, e.location FROM equipments e JOIN process_equipment pe ON e.id = pe.equipment_id WHERE pe.process_code = ?", (process_code,))
    equips = [dict(zip(['id', 'equipment_code', 'equipment_name', 'equipment_type', 'capacity_per_hour', 'location'], row)) for row in c.fetchall()]
    conn.close()
    return jsonify(equips)

# ========== Processes CRUD ==========
@app.route('/api/processes')
@login_required
def api_processes():
    q = request.args.get('q', '')
    team = request.args.get('team', '')
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM processes WHERE 1=1"
    params = []
    if q:
        sql += " AND (process_name LIKE ? OR process_code LIKE ?)"
        params.extend(['%'+q+'%', '%'+q+'%'])
    if team:
        sql += " AND team_name LIKE ?"
        params.append('%' + team + '%')
    sql += " ORDER BY team_name, process_code"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/processes', methods=['POST'])
@login_required
def api_process_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO processes (process_code, process_name, team_name) VALUES (?, ?, ?)",
                  (data['process_code'], data['process_name'], data.get('team_name', '')))
        conn.commit()
        return jsonify({'id': c.lastrowid, 'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/processes/<int:pid>', methods=['PUT'])
@login_required
def api_process_update(pid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE processes SET process_code=?, process_name=?, team_name=? WHERE id=?",
              (data['process_code'], data['process_name'], data.get('team_name', ''), pid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/processes/<int:pid>', methods=['DELETE'])
@login_required
def api_process_delete(pid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM processes WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Production Requirements API ==========
@app.route('/api/production-requirements')
@login_required
def api_production_requirements():
    status = request.args.get('status', '')
    conn = get_connection()
    c = conn.cursor()
    if status:
        c.execute("SELECT * FROM production_requirements WHERE status=? ORDER BY required_date, product_code", (status,))
    else:
        c.execute("SELECT * FROM production_requirements ORDER BY required_date, product_code")
    results = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(results)

@app.route('/api/production-requirements/calculate', methods=['POST'])
@login_required
@planner_required
def api_calculate_requirements():
    count = calculate_production_requirements()
    return jsonify({'count': count})

@app.route('/api/production-requirements/publish', methods=['POST'])
@login_required
@planner_required
def api_publish_requirements():
    data = request.json or {}
    ids = data.get('ids', [])
    count = publish_requirements(ids if ids else None)
    # Create publish batch record
    if count > 0:
        conn = get_connection()
        c = conn.cursor()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        batch_no = 'PB-' + datetime.now().strftime('%Y%m%d%H%M%S')
        c.execute("SELECT COUNT(*), COALESCE(SUM(required_quantity),0) FROM production_requirements WHERE status='published' AND (batch_id IS NULL OR batch_id=0)", )
        row = c.fetchone()
        total_count = row[0] if row else 0
        total_qty = row[1] if row else 0
        c.execute("INSERT INTO publish_batches (batch_no, published_at, total_count, total_quantity, published_by) VALUES (?,?,?,?,?)",
                  (batch_no, now, total_count, total_qty, session.get('username','')))
        batch_id = c.lastrowid
        c.execute("UPDATE production_requirements SET batch_id=?, published_at=? WHERE status='published' AND (batch_id IS NULL OR batch_id=0)",
                  (batch_id, now))
        conn.commit()
        conn.close()
    return jsonify({'published': count})

@app.route('/api/publish-batches')
@login_required
@planner_required
def api_publish_batches():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM publish_batches ORDER BY published_at DESC")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/publish-batches/<int:batch_id>/details')
@login_required
@planner_required
def api_publish_batch_details(batch_id):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM production_requirements WHERE batch_id=? ORDER BY parent_chain, required_date", (batch_id,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route('/api/schedule-warnings')
@login_required
def api_schedule_warnings():
    team_name = session.get('team_name', '')
    if not team_name:
        team_name = request.args.get('team', '')
    conn = get_connection()
    c = conn.cursor()
    # Get published requirements for this team
    c.execute("""SELECT product_code, product_name, required_date, required_quantity, team_name
        FROM production_requirements WHERE status='published' ORDER BY required_date""")
    rows = c.fetchall()
    # Filter by team (team_name may be comma-separated in requirements)
    filtered = []
    for r in rows:
        tn = r[4] or ''
        if team_name and team_name not in tn:
            continue
        filtered.append({'product_code': r[0], 'product_name': r[1], 'required_date': r[2], 'required_quantity': r[3]})
    # Get incomplete quantities from work_orders
    c.execute("""SELECT product_code, product_name, SUM(COALESCE(quantity,0) - COALESCE(completed_qty,0)) as remaining
        FROM work_orders WHERE status != 'completed' GROUP BY product_code""")
    wo_remaining = {}
    for r in c.fetchall():
        key = r[1] or r[0]  # use product_name first, fallback to code
        wo_remaining[key] = r[2] if r[2] else 0
    conn.close()
    return jsonify({'requirements': filtered, 'work_order_remaining': wo_remaining})

@app.route('/api/production-requirements/for-team')
@login_required
def api_requirements_for_team():
    """Get published requirements for current team"""
    team_name = request.args.get('team', '')
    if not team_name:
        team_name = session.get('team_name', '')
    
    # Date range is optional
    start_date = request.args.get('start', '')
    end_date = request.args.get('end', '')
    
    requirements = get_published_requirements(team_name, start_date if start_date else None, end_date if end_date else None)
    return jsonify(requirements)

@app.route('/api/production-requirements/<int:rid>', methods=['DELETE'])
@login_required
@planner_required
def api_delete_requirement(rid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM production_requirements WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Reference Panel APIs ==========
@app.route('/api/reference/semi-finished-output')
@login_required
def api_semi_finished_output():
    """Return work reports from last 2 days for semi-finished product processes.
    Semi-finished = second-to-last process in route (or the only process if just one).
    """
    from datetime import datetime, timedelta
    conn = get_connection()
    c = conn.cursor()

    # 1. Build product_code -> semi-finished process name mapping
    c.execute("SELECT product_code, process_list FROM process_routes WHERE process_list IS NOT NULL AND process_list != ''")
    routes = c.fetchall()
    semi_process_map = {}  # product_code -> set of semi-finished process names
    for pc, pl in routes:
        procs = [p.strip() for p in pl.split(',') if p.strip()]
        if not procs:
            continue
        if len(procs) == 1:
            target = procs[0]
        else:
            target = procs[-2]
        if pc not in semi_process_map:
            semi_process_map[pc] = set()
        semi_process_map[pc].add(target)

    if not semi_process_map:
        conn.close()
        return jsonify([])

    # 2. Get reports from last 2 days (join personnel for team name)
    cutoff = (datetime.now() - timedelta(days=2)).strftime('%Y-%m-%d')
    c.execute("""SELECT r.product_code, r.product_name, r.process_name, r.report_qty, r.good_qty,
                 r.create_time, r.operator, t.name as team_name
                 FROM work_reports r
                 LEFT JOIN personnel p ON r.operator = p.name
                 LEFT JOIN teams t ON p.team_id = t.id
                 WHERE r.create_time >= ?
                 ORDER BY r.create_time DESC""", (cutoff,))
    all_reports = c.fetchall()

    # 3. Filter: keep only reports where process matches semi-finished process for that product
    team_data = {}
    for r in all_reports:
        pc, pn, proc, qty, gqty, ctime, op, team = r
        allowed = semi_process_map.get(pc)
        if allowed and proc in allowed:
            tn = team or '未知班组'
            if tn not in team_data:
                team_data[tn] = []
            team_data[tn].append({
                'product_code': pc or '',
                'product_name': pn or '',
                'process_name': proc or '',
                'report_qty': qty or 0,
                'good_qty': gqty or 0,
                'create_time': ctime or '',
                'operator': op or ''
            })

    conn.close()

    # 4. Format response
    result = []
    for tn in sorted(team_data.keys()):
        result.append({'team': tn, 'records': team_data[tn]})

    return jsonify(result)

# ========== Process Routes CRUD ==========
@app.route('/api/process-routes')
@login_required
def api_process_routes():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM process_routes WHERE 1=1"
    params = []
    if q:
        like = "%" + q + "%"
        sql += " AND (route_code LIKE ? OR route_name LIKE ? OR product_code LIKE ? OR process_list LIKE ?)"
        params.extend([like, like, like, like])
    sql += " ORDER BY route_code"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/process-routes', methods=['POST'])
@login_required
def api_process_route_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO process_routes (route_code, route_name, product_code, process_list, remark) VALUES (?,?,?,?,?)",
              (data['route_code'], data.get('route_name',''), data.get('product_code',''), data.get('process_list',''), data.get('remark','')))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/process-routes/<int:rid>', methods=['PUT'])
@login_required
def api_process_route_update(rid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE process_routes SET route_code=?, route_name=?, product_code=?, process_list=?, remark=? WHERE id=?",
              (data['route_code'], data.get('route_name',''), data.get('product_code',''), data.get('process_list',''), data.get('remark',''), rid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/process-routes/<int:rid>', methods=['DELETE'])
@login_required
def api_process_route_delete(rid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM process_routes WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Production Cycles CRUD ==========
@app.route('/api/production-cycles')
@login_required
def api_production_cycles():
    conn = get_connection()
    c = conn.cursor()
    keyword = request.args.get('q', '')
    if keyword:
        c.execute("SELECT * FROM production_cycles WHERE product_code LIKE ? LIMIT 50", ('%' + keyword + '%',))
    else:
        c.execute("SELECT * FROM production_cycles LIMIT 100")
    r = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(r)

@app.route('/api/production-cycles', methods=['POST'])
@login_required
def api_production_cycle_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO production_cycles (product_code, production_days, lead_days) VALUES (?,?,?)",
              (data['product_code'], data.get('production_days', 1), data.get('lead_days', 0)))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/production-cycles/<int:cid>', methods=['PUT'])
@login_required
def api_production_cycle_update(cid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE production_cycles SET product_code=?, production_days=?, lead_days=? WHERE id=?",
              (data['product_code'], data.get('production_days', 1), data.get('lead_days', 0), cid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/production-cycles/<int:cid>', methods=['DELETE'])
@login_required
def api_production_cycle_delete(cid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM production_cycles WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== BOM CRUD ==========
@app.route('/api/bom')
@login_required
def api_bom():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM bom WHERE 1=1"
    params = []
    if q:
        like = "%" + q + "%"
        sql += " AND (parent_product_code LIKE ? OR parent_product_name LIKE ? OR child_product_code LIKE ? OR child_product_name LIKE ?)"
        params.extend([like, like, like, like])
    sql += " ORDER BY parent_product_code"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/bom', methods=['POST'])
@login_required
def api_bom_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO bom (parent_product_code, parent_product_name, child_product_code, child_product_name, quantity, unit, process_team) VALUES (?,?,?,?,?,?,?)",
              (data['parent_product_code'], data.get('parent_product_name',''), data.get('child_product_code',''),
               data.get('child_product_name',''), data.get('quantity', 1), data.get('unit',''), data.get('process_team','')))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'ok': True})

@app.route('/api/bom/<int:bid>', methods=['PUT'])
@login_required
def api_bom_update(bid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE bom SET parent_product_code=?, parent_product_name=?, child_product_code=?, child_product_name=?, quantity=?, unit=?, process_team=? WHERE id=?",
              (data['parent_product_code'], data.get('parent_product_name',''), data.get('child_product_code',''),
               data.get('child_product_name',''), data.get('quantity', 1), data.get('unit',''), data.get('process_team',''), bid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/bom/<int:bid>', methods=['DELETE'])
@login_required
def api_bom_delete(bid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM bom WHERE id=?", (bid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Personnel API ==========
@app.route('/api/personnel')
@login_required
def api_personnel():
    team_id = request.args.get('team_id', type=int)
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM personnel WHERE 1=1"
    params = []
    if team_id:
        sql += " AND team_id=?"
        params.append(team_id)
    if q:
        like = "%" + q + "%"
        sql += " AND (name LIKE ? OR user_id LIKE ? OR department LIKE ? OR position LIKE ?)"
        params.extend([like, like, like, like])
    sql += " ORDER BY team_id, name"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/personnel', methods=['POST'])
@login_required
def api_personnel_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO personnel (user_id, name, department, position, team_id) VALUES (?, ?, ?, ?, ?)",
              (data.get('user_id',''), data['name'], data.get('department',''), data.get('position',''), data.get('team_id')))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/personnel/<int:pid>', methods=['PUT'])
@login_required
def api_personnel_update(pid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE personnel SET user_id=?, name=?, department=?, position=?, team_id=?, is_active=? WHERE id=?",
              (data.get('user_id',''), data['name'], data.get('department',''), data.get('position',''), data.get('team_id'), data.get('is_active',1), pid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/personnel/<int:pid>', methods=['DELETE'])
@login_required
def api_personnel_delete(pid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM personnel WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Molds API ==========
@app.route('/api/molds')
@login_required
def api_molds():
    team_id = request.args.get('team_id', type=int)
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM molds WHERE 1=1"
    params = []
    if team_id:
        sql += " AND team_id=?"
        params.append(team_id)
    if q:
        like = "%" + q + "%"
        sql += " AND (mold_code LIKE ? OR mold_name LIKE ? OR product_code LIKE ? OR location LIKE ?)"
        params.extend([like, like, like, like])
    sql += " ORDER BY mold_code"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/molds', methods=['POST'])
@login_required
def api_molds_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO molds (mold_code, mold_name, mold_type, product_code, status, location, team_id, remark) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
              (data['mold_code'], data.get('mold_name',''), data.get('mold_type',''), data.get('product_code',''), data.get('status','normal'), data.get('location',''), data.get('team_id'), data.get('remark','')))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/molds/<int:mid>', methods=['PUT'])
@login_required
def api_molds_update(mid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE molds SET mold_code=?, mold_name=?, mold_type=?, product_code=?, status=?, location=?, team_id=?, remark=? WHERE id=?",
              (data['mold_code'], data.get('mold_name',''), data.get('mold_type',''), data.get('product_code',''), data.get('status','normal'), data.get('location',''), data.get('team_id'), data.get('remark',''), mid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/molds/<int:mid>', methods=['DELETE'])
@login_required
def api_molds_delete(mid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM molds WHERE id=?", (mid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Personnel & Mold Import ==========
@app.route('/api/import/personnel', methods=['POST'])
@login_required
def api_import_personnel():
    from utils.excel import import_personnel
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '请上传文件'}), 400
    path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    f.save(path)
    try:
        count = import_personnel(path)
        return jsonify({'ok': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/import/molds', methods=['POST'])
@login_required
def api_import_molds():
    from utils.excel import import_molds
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '请上传文件'}), 400
    path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    f.save(path)
    try:
        count = import_molds(path)
        return jsonify({'ok': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# ========== Standard Hours API ==========
@app.route('/api/standard-hours')
@login_required
def api_standard_hours():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = "SELECT * FROM standard_hours WHERE 1=1"
    params = []
    if q:
        like = "%" + q + "%"
        sql += " AND (product_code LIKE ? OR product_name LIKE ? OR process_name LIKE ?)"
        params.extend([like, like, like])
    sql += " ORDER BY product_code, process_name"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/standard-hours', methods=['POST'])
@login_required
def api_standard_hours_create():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO standard_hours (product_code, product_name, process_name, team_name, standard_hours, setup_time, remark) VALUES (?, ?, ?, ?, ?, ?, ?)",
                  (data['product_code'], data.get('product_name',''), data['process_name'], data.get('team_name',''), data.get('standard_hours',0), data.get('setup_time',0), data.get('remark','')))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 400
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/standard-hours/<int:sid>', methods=['PUT'])
@login_required
def api_standard_hours_update(sid):
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE standard_hours SET product_code=?, product_name=?, process_name=?, team_name=?, standard_hours=?, setup_time=?, remark=? WHERE id=?",
              (data['product_code'], data.get('product_name',''), data['process_name'], data.get('team_name',''), data.get('standard_hours',0), data.get('setup_time',0), data.get('remark',''), sid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/standard-hours/<int:sid>', methods=['DELETE'])
@login_required
def api_standard_hours_delete(sid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM standard_hours WHERE id=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/standard-hours/init', methods=['POST'])
@login_required
def api_standard_hours_init():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT product_code, route_name, process_list, remark FROM process_routes WHERE process_list IS NOT NULL AND process_list != ''")
    routes = c.fetchall()
    count = 0
    for route in routes:
        product_code = route[0] or ''
        product_name = route[1] or ''
        process_list = route[2] or ''
        team_name = route[3] or ''
        processes = [p.strip() for p in process_list.split(',') if p.strip()]
        for proc in processes:
            try:
                c.execute("INSERT OR IGNORE INTO standard_hours (product_code, product_name, process_name, team_name) VALUES (?, ?, ?, ?)",
                          (product_code, product_name, proc, team_name))
                if c.rowcount > 0:
                    count += 1
            except:
                pass
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'count': count})

# ========== Work Reports API ==========
@app.route('/api/work-reports')
@login_required
def api_work_reports():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    sql = """SELECT r.*, t.name as team_name FROM work_reports r
              LEFT JOIN personnel p ON r.operator = p.name
              LEFT JOIN teams t ON p.team_id = t.id
              WHERE 1=1"""
    params = []
    date = request.args.get('date', '').strip()
    if date:
        sql += " AND r.create_time LIKE ?"
        params.append(date + '%')
    team_filter = request.args.get('team', '').strip()
    if team_filter:
        sql += " AND t.name = ?"
        params.append(team_filter)
    if q:
        like = "%" + q + "%"
        sql += " AND (r.order_no LIKE ? OR r.product_code LIKE ? OR r.product_name LIKE ? OR r.process_name LIKE ? OR r.operator LIKE ?)"
        params.extend([like, like, like, like, like])
    sql += " ORDER BY r.create_time DESC"
    return jsonify(paginate_query(sql, params, page, page_size))

@app.route('/api/work-reports', methods=['DELETE'])
@login_required
@planner_required
def api_work_reports_clear():
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM work_reports")
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/import/work-reports', methods=['POST'])
@login_required
@planner_required
def api_import_work_reports():
    from utils.excel import import_work_reports
    f = request.files.get('file')
    if not f:
        return jsonify({'error': '请上传文件'}), 400
    path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    f.save(path)
    try:
        count = import_work_reports(path)
        _recalculate_work_report_efficiency()
        try:
            _refresh_equipment_status()
        except Exception:
            pass
        return jsonify({'ok': True, 'success': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# ========== System Settings API ==========
@app.route('/api/settings', methods=['GET'])
@login_required
def api_get_settings():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT key, value FROM system_settings")
    r = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return jsonify(r)

@app.route('/api/settings', methods=['POST'])
@login_required
@planner_required
def api_save_settings():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    for key, value in data.items():
        c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES (?, ?, datetime('now','localtime'))", (key, str(value)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Permissions API ==========
@app.route('/api/permissions', methods=['GET'])
@login_required
def api_get_permissions():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT key, value FROM system_settings WHERE key LIKE 'perm_%'")
    rows = c.fetchall()
    conn.close()
    
    permissions = {}
    for row in rows:
        key = row[0].replace('perm_', '')
        try:
            permissions[key] = eval(row[1])
        except:
            permissions[key] = {'admin': True, 'planner': True, 'team': False}
    
    # Default permissions if not set
    defaults = {
        'workbench': {'admin': True, 'planner': True, 'team': True},
        'schedule': {'admin': False, 'planner': False, 'team': True},
        'shipping_plan': {'admin': True, 'planner': True, 'team': False},
        'plan_approval': {'admin': True, 'planner': True, 'team': False},
        'alerts': {'admin': True, 'planner': True, 'team': False},
        'orders': {'admin': True, 'planner': True, 'team': False},
        'basic_data': {'admin': True, 'planner': True, 'team': True},
        'basic_data_edit': {'admin': True, 'planner': True, 'team': False},
        'production_reports': {'admin': True, 'planner': False, 'team': False},
        'system_management': {'admin': True, 'planner': False, 'team': False}
    }
    
    for key, default in defaults.items():
        if key not in permissions:
            permissions[key] = default
    
    return jsonify(permissions)

@app.route('/api/permissions', methods=['POST'])
@login_required
@admin_required
def api_save_permissions():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    for key, value in data.items():
        c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES (?, ?, datetime('now','localtime'))", ('perm_' + key, str(value)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== MES Sync API ==========
@app.route('/api/mes-sync/settings', methods=['GET'])
@login_required
@planner_required
def api_get_sync_settings():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT key, value FROM system_settings WHERE key IN ('mes_work_order_interval','mes_report_interval')")
    r = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return jsonify({
        'work_order_interval': int(r.get('mes_work_order_interval', 0)),
        'report_interval': int(r.get('mes_report_interval', 0))
    })

@app.route('/api/mes-sync/settings', methods=['POST'])
@login_required
@planner_required
def api_save_sync_settings():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    wo_interval = data.get('work_order_interval', 0)
    rpt_interval = data.get('report_interval', 0)
    c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES ('mes_work_order_interval', ?, datetime('now','localtime'))", (str(wo_interval),))
    c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES ('mes_report_interval', ?, datetime('now','localtime'))", (str(rpt_interval),))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/mes-sync/work-orders', methods=['POST'])
@login_required
@planner_required
def api_sync_work_orders():
    import subprocess
    log_id = None
    try:
        conn = get_connection()
        c = conn.cursor()
        c.execute("INSERT INTO sync_logs (sync_type, status, created_at, trigger_type) VALUES ('work_orders', 'running', datetime('now','localtime'), 'manual')")
        log_id = c.lastrowid
        conn.commit()
        conn.close()
    except Exception:
        pass
    try:
        script = os.path.join(os.path.dirname(__file__), 'scripts', 'sync_work_orders.py')
        with _mes_sync_lock:
            result = subprocess.run(
                [sys.executable, '-u', script],
                capture_output=True, text=True, timeout=300,
                cwd=os.path.dirname(__file__)
            )
        output = result.stdout.strip() if result.stdout else ''
        errors = result.stderr.strip() if result.stderr else ''
        if result.returncode == 0:
            count = 0
            for line in output.split('\n'):
                if '总计=' in line:
                    try:
                        count = int(line.split('总计=')[1].strip())
                    except ValueError:
                        pass
                    break
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='success', record_count=?, detail=? WHERE id=?", (count, output[-500:], log_id))
            conn.commit()
            conn.close()
            _recalculate_work_report_efficiency()
            try:
                _check_material_alerts()
            except Exception:
                pass
            try:
                _refresh_equipment_status()
            except Exception:
                pass
            return jsonify({'ok': True, 'count': count, 'output': output[-1000:]})
        else:
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='error', detail=? WHERE id=?", (errors[-500:], log_id))
            conn.commit()
            conn.close()
            return jsonify({'ok': False, 'error': errors[-500:]})
    except subprocess.TimeoutExpired:
        conn = get_connection()
        c = conn.cursor()
        c.execute("UPDATE sync_logs SET status='error', detail='同步超时(>300秒)' WHERE id=?", (log_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': False, 'error': '同步超时(>300秒)'})
    except Exception as e:
        conn = get_connection()
        c = conn.cursor()
        c.execute("UPDATE sync_logs SET status='error', detail=? WHERE id=?", (str(e)[:500], log_id))
        conn.commit()
        conn.close()
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/mes-sync/reports', methods=['POST'])
@login_required
@planner_required
def api_sync_reports():
    import subprocess
    log_id = None
    try:
        conn = get_connection()
        c = conn.cursor()
        c.execute("INSERT INTO sync_logs (sync_type, status, created_at, trigger_type) VALUES ('reports', 'running', datetime('now','localtime'), 'manual')")
        log_id = c.lastrowid
        conn.commit()
        conn.close()
    except Exception:
        pass
    try:
        script = os.path.join(os.path.dirname(__file__), 'scripts', 'sync_reports.py')
        with _mes_sync_lock:
            result = subprocess.run(
                [sys.executable, '-u', script],
                capture_output=True, text=True, timeout=300,
                cwd=os.path.dirname(__file__)
            )
        output = result.stdout.strip() if result.stdout else ''
        errors = result.stderr.strip() if result.stderr else ''
        if result.returncode == 0:
            count = 0
            for line in output.split('\n'):
                if '总计=' in line:
                    try:
                        count = int(line.split('总计=')[1].strip())
                    except ValueError:
                        pass
                    break
                if '插入=' in line:
                    try:
                        count += int(line.split('插入=')[1].split(',')[0].strip())
                    except ValueError:
                        pass
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='success', record_count=?, detail=? WHERE id=?", (count, output[-500:], log_id))
            conn.commit()
            conn.close()
            return jsonify({'ok': True, 'count': count, 'output': output[-1000:]})
        else:
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE sync_logs SET status='error', detail=? WHERE id=?", (errors[-500:], log_id))
            conn.commit()
            conn.close()
            return jsonify({'ok': False, 'error': errors[-500:]})
    except subprocess.TimeoutExpired:
        conn = get_connection()
        c = conn.cursor()
        c.execute("UPDATE sync_logs SET status='error', detail='同步超时(>300秒)' WHERE id=?", (log_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': False, 'error': '同步超时(>300秒)'})
    except Exception as e:
        conn = get_connection()
        c = conn.cursor()
        c.execute("UPDATE sync_logs SET status='error', detail=? WHERE id=?", (str(e)[:500], log_id))
        conn.commit()
        conn.close()
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/mes-sync/logs', methods=['GET'])
@login_required
@planner_required
def api_get_sync_logs():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, sync_type, status, record_count, detail, created_at, trigger_type FROM sync_logs ORDER BY id DESC LIMIT 50")
    rows = c.fetchall()
    conn.close()
    return jsonify([{'id':r[0],'sync_type':r[1],'status':r[2],'record_count':r[3],'detail':r[4],'created_at':r[5],'trigger_type':r[6]} for r in rows])

# ========== Schedules Lookup API ==========
@app.route('/api/schedules')
@login_required
def api_schedules():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 200, type=int)
    sql = "SELECT s.*, e.equipment_name FROM schedules s LEFT JOIN equipments e ON s.equipment_id=e.id WHERE 1=1"
    params = []
    if q:
        like = "%" + q + "%"
        sql += " AND (s.product_code LIKE ? OR s.process_name LIKE ? OR s.work_order_no LIKE ?)"
        params.extend([like, like, like])
    sql += " ORDER BY s.schedule_date DESC, s.start_time"
    return jsonify(paginate_query(sql, params, page, page_size))

# ========== Standard Hours Capacity API ==========
@app.route('/api/standard-hours-capacity')
@login_required
def api_standard_hours_capacity():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    page_size = request.args.get('page_size', 50, type=int)
    conn = get_connection()
    c = conn.cursor()
    
    # 过滤条件: 排除半成品入库和含清洗的工序
    exclude_filter = " AND sh.process_name != '\u534a\u6210\u54c1\u5165\u5e93' AND sh.process_name NOT LIKE '%\u6e05\u6d17%'"
    
    # Base query
    sql = "SELECT sh.* FROM standard_hours sh WHERE 1=1" + exclude_filter
    params = []
    if q:
        like = "%" + q + "%"
        sql += " AND (sh.product_code LIKE ? OR sh.product_name LIKE ? OR sh.process_name LIKE ?)"
        params.extend([like, like, like])
    sql += " ORDER BY sh.product_code, sh.process_name"
    
    # Count
    count_sql = "SELECT COUNT(*) FROM (" + sql + ")"
    c.execute(count_sql, params)
    total = c.fetchone()[0]
    
    # Paginate
    page = max(1, int(page or 1))
    page_size = min(200, max(1, int(page_size or 50)))
    offset = (page - 1) * page_size
    c.execute(sql + " LIMIT ? OFFSET ?", params + [page_size, offset])
    rows = [dict(row) for row in c.fetchall()]
    
    # Enrich each row with capacity data
    for row in rows:
        pc = row['product_code']
        pn = row['process_name']
        
        # 排班产能: from schedules
        c.execute("SELECT DISTINCT capacity_per_hour FROM schedules WHERE product_code=? AND process_name=? AND capacity_per_hour > 0", (pc, pn))
        sched_caps = [r[0] for r in c.fetchall()]
        row['schedule_capacities'] = sched_caps
        if len(sched_caps) == 1:
            row['schedule_capacity'] = sched_caps[0]
        elif len(sched_caps) > 1:
            row['schedule_capacity'] = -1
        else:
            row['schedule_capacity'] = 0
        
        # 报工产能: from work_reports
        c.execute("SELECT report_qty, report_hours, order_no, operator, equipment, start_time, end_time FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0", (pc, pn))
        reports = c.fetchall()
        caps = []
        report_details = []
        for r in reports:
            qty, hrs, ono, op, eq, st, et = r
            if hrs > 0 and qty > 0:
                cap = round(qty / hrs, 1)
                caps.append(cap)
                report_details.append({'qty': qty, 'hours': hrs, 'capacity': cap, 'order_no': ono, 'operator': op, 'equipment': eq, 'start': st, 'end': et})
        row['report_avg'] = round(sum(caps) / len(caps), 1) if caps else 0
        row['report_max'] = max(caps) if caps else 0
        row['report_min'] = min(caps) if caps else 0
        row['report_count'] = len(caps)
    
    # 优化的进度计算: 用单条SQL统计
    progress_sql = """
        SELECT 
            COUNT(DISTINCT sh.product_code || '|' || sh.process_name) as total,
            COUNT(DISTINCT CASE WHEN s.product_code IS NOT NULL OR wr.product_code IS NOT NULL THEN sh.product_code || '|' || sh.process_name END) as completed
        FROM standard_hours sh
        LEFT JOIN (
            SELECT DISTINCT product_code, process_name 
            FROM schedules 
            WHERE capacity_per_hour > 0
        ) s ON sh.product_code = s.product_code AND sh.process_name = s.process_name
        LEFT JOIN (
            SELECT DISTINCT product_code, process_name 
            FROM work_reports 
            WHERE report_hours > 0 AND report_qty > 0
        ) wr ON sh.product_code = wr.product_code AND sh.process_name = wr.process_name
        WHERE 1=1""" + exclude_filter.replace("sh.", "sh.")
    c.execute(progress_sql)
    prog = c.fetchone()
    total_combos = prog[0] or 0
    has_cap_count = prog[1] or 0
    
    conn.close()
    total_pages = max(1, (total + page_size - 1) // page_size)
    return jsonify({
        'data': rows, 
        'total': total, 
        'page': page, 
        'page_size': page_size, 
        'total_pages': total_pages,
        'progress': {
            'total': total_combos,
            'completed': has_cap_count,
            'percent': round(has_cap_count / total_combos * 100, 1) if total_combos > 0 else 0
        }
    })

# ========== System Config API ==========
@app.route('/api/system-config')
@login_required
def api_get_config():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT key, value FROM system_config")
    config = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return jsonify(config)

@app.route('/api/system-config', methods=['POST'])
@login_required
@planner_required
def api_save_config():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    for key, value in data.items():
        c.execute("INSERT OR REPLACE INTO system_config (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


def _refresh_equipment_status():
    """Refresh equipment status from today's reports only.
    Matches report equipment to equipments table with team validation."""
    try:
        conn = get_connection()
        c = conn.cursor()
        c.execute("UPDATE equipments SET status='normal'")
        c.execute("SELECT id, equipment_name, equipment_code, team_id FROM equipments")
        eq_list = []
        for row in c.fetchall():
            eq_list.append({'id': row[0], 'name': row[1], 'code': row[2], 'team_id': row[3]})
        c.execute("SELECT id, name FROM teams")
        team_name_map = {r[0]: r[1] for r in c.fetchall()}
        c.execute("SELECT process_name, team_name FROM processes WHERE team_name IS NOT NULL AND team_name != ''")
        process_team_map = {}
        for r in c.fetchall():
            process_team_map[r[0]] = [t.strip() for t in r[1].split(',') if t.strip()]
        c.execute("""
            SELECT wr.equipment, wr.order_no, wr.process_name
            FROM work_reports wr
            INNER JOIN (
                SELECT equipment, MAX(create_time) as max_time
                FROM work_reports
                WHERE equipment IS NOT NULL AND equipment != ''
                  AND date(create_time) = date('now','localtime')
                GROUP BY equipment
            ) latest ON wr.equipment = latest.equipment AND wr.create_time = latest.max_time
            WHERE wr.equipment IS NOT NULL AND wr.equipment != ''
              AND wr.order_no IS NOT NULL AND wr.order_no != ''
        """)
        latest_reports = c.fetchall()
        import unicodedata as _ud
        def _norm(s):
            if not s: return ''
            s = _ud.normalize('NFKC', s).strip().lower()
            s = re.split(r'[\uff0c\u3002\uff1b\uff1a\uff08(]', s)[0]
            s = s.replace('/', '-').replace('.', '-').replace('_', '-')
            s = s.replace('\u53f7', '').replace('\u53f0', '').replace('\u7ebf', '')
            while '--' in s: s = s.replace('--', '-')
            return s.replace(' ', '')
        def _strip(s):
            return re.sub(r'[^a-z0-9]', '', _norm(s)) if s else ''
        updated = 0
        for equip_raw, order_no, process_name in latest_reports:
            if not order_no or not process_name:
                continue
            matched_eq = None
            for eq in eq_list:
                if eq['name'] == equip_raw or eq['code'] == equip_raw:
                    matched_eq = eq
                    break
            if not matched_eq:
                nk = _norm(equip_raw)
                for eq in eq_list:
                    if _norm(eq['name']) == nk or _norm(eq['code']) == nk:
                        matched_eq = eq
                        break
            if not matched_eq:
                sk = _strip(equip_raw)
                if sk:
                    for eq in eq_list:
                        if _strip(eq['name']) == sk or _strip(eq['code']) == sk:
                            matched_eq = eq
                            break
            if not matched_eq:
                continue
            eq_team_name = team_name_map.get(matched_eq['team_id'], '')
            process_teams = process_team_map.get(process_name, [])
            if process_teams and eq_team_name and eq_team_name not in process_teams:
                continue
            c.execute("SELECT process_progress FROM work_orders WHERE order_no=?", (order_no,))
            row = c.fetchone()
            if not row or not row[0]:
                continue
            progress_str = row[0]
            process_completed = False
            if process_name:
                pattern = re.escape(process_name) + r'.*?[\u3010(\d+)/(\d+)\u3011]'
                match = re.search(pattern, progress_str)
                if match:
                    done = int(match.group(1))
                    total = int(match.group(2))
                    if total > 0 and done >= total:
                        process_completed = True
                else:
                    process_completed = False
            if not process_completed:
                c.execute("UPDATE equipments SET status='running' WHERE id=?", (matched_eq['id'],))
                updated += 1
        conn.commit()
        conn.close()
        print(f"[equipment-status] Refreshed: {updated} equipment set to running")
        return updated
    except Exception as e:
        print(f"[equipment-status] Refresh error: {e}")
        return 0

def _recalculate_work_report_efficiency():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT value FROM system_config WHERE key='capacity_baseline'")
    row = c.fetchone()
    baseline = int(row[0]) if row else 1
    capacity_map = {}
    c.execute("SELECT DISTINCT product_code, process_name FROM standard_hours WHERE process_name != '半成品入库' AND process_name NOT LIKE '%清洗%'")
    all_combos = c.fetchall()
    for pc, pn in all_combos:
        cap = 0
        if baseline == 1:
            c.execute("SELECT DISTINCT capacity_per_hour FROM schedules WHERE product_code=? AND process_name=? AND capacity_per_hour > 0", (pc, pn))
            caps = [r[0] for r in c.fetchall()]
            cap = sum(caps) / len(caps) if caps else 0
        elif baseline == 2:
            c.execute("SELECT report_qty, report_hours FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0 AND report_qty > 0", (pc, pn))
            reports = c.fetchall()
            if reports:
                caps = [r[0]/r[1] for r in reports if r[1] > 0]
                cap = sum(caps) / len(caps) if caps else 0
        elif baseline == 3:
            c.execute("SELECT report_qty, report_hours FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0 AND report_qty > 0", (pc, pn))
            reports = c.fetchall()
            if reports:
                caps = [r[0]/r[1] for r in reports if r[1] > 0]
                cap = max(caps) if caps else 0
        elif baseline == 4:
            c.execute("SELECT report_qty, report_hours FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0 AND report_qty > 0", (pc, pn))
            reports = c.fetchall()
            if reports:
                caps = [r[0]/r[1] for r in reports if r[1] > 0]
                cap = min(caps) if caps else 0
        if cap > 0:
            capacity_map[(pc, pn)] = cap
    c.execute("SELECT id, product_code, process_name, report_qty, report_hours FROM work_reports WHERE report_hours > 0 AND report_qty > 0")
    reports = c.fetchall()
    updated = 0
    for rid, pc, pn, qty, hrs in reports:
        key = (pc, pn)
        if key in capacity_map and capacity_map[key] > 0:
            efficiency = round((qty / hrs) / capacity_map[key] * 100, 1)
            c.execute("UPDATE work_reports SET efficiency=? WHERE id=?", (efficiency, rid))
            updated += 1
        else:
            c.execute("UPDATE work_reports SET efficiency=0 WHERE id=?", (rid,))
    conn.commit()
    conn.close()

@app.route('/api/recalculate-efficiency', methods=['POST'])
@login_required
@planner_required
def api_recalculate_efficiency():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT value FROM system_config WHERE key='capacity_baseline'")
    row = c.fetchone()
    baseline = int(row[0]) if row else 1
    capacity_map = {}
    c.execute("SELECT DISTINCT product_code, process_name FROM standard_hours WHERE process_name != '半成品入库' AND process_name NOT LIKE '%清洗%'")
    all_combos = c.fetchall()
    for pc, pn in all_combos:
        cap = 0
        if baseline == 1:
            c.execute("SELECT DISTINCT capacity_per_hour FROM schedules WHERE product_code=? AND process_name=? AND capacity_per_hour > 0", (pc, pn))
            caps = [r[0] for r in c.fetchall()]
            cap = sum(caps) / len(caps) if caps else 0
        elif baseline == 2:
            c.execute("SELECT report_qty, report_hours FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0 AND report_qty > 0", (pc, pn))
            reports = c.fetchall()
            if reports:
                caps = [r[0]/r[1] for r in reports if r[1] > 0]
                cap = sum(caps) / len(caps) if caps else 0
        elif baseline == 3:
            c.execute("SELECT report_qty, report_hours FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0 AND report_qty > 0", (pc, pn))
            reports = c.fetchall()
            if reports:
                caps = [r[0]/r[1] for r in reports if r[1] > 0]
                cap = max(caps) if caps else 0
        elif baseline == 4:
            c.execute("SELECT report_qty, report_hours FROM work_reports WHERE product_code=? AND process_name=? AND report_hours > 0 AND report_qty > 0", (pc, pn))
            reports = c.fetchall()
            if reports:
                caps = [r[0]/r[1] for r in reports if r[1] > 0]
                cap = min(caps) if caps else 0
        if cap > 0:
            capacity_map[(pc, pn)] = cap
    c.execute("SELECT id, product_code, process_name, report_qty, report_hours FROM work_reports WHERE report_hours > 0 AND report_qty > 0")
    reports = c.fetchall()
    updated = 0
    for rid, pc, pn, qty, hrs in reports:
        key = (pc, pn)
        if key in capacity_map and capacity_map[key] > 0:
            efficiency = round((qty / hrs) / capacity_map[key] * 100, 1)
            c.execute("UPDATE work_reports SET efficiency=? WHERE id=?", (efficiency, rid))
            updated += 1
        else:
            c.execute("UPDATE work_reports SET efficiency=0 WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'updated': updated, 'baseline': baseline})

# ========== Personnel Reports API ==========

def _get_personnel_map():
    """Build name -> {team_id, department, position} mapping."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT name, team_id, department, position FROM personnel")
    mapping = {row[0]: {'team_id': row[1], 'department': row[2], 'position': row[3]} for row in c.fetchall()}
    conn.close()
    return mapping

def _get_standard_capacity_map():
    """Build (product_code, process_name) -> standard_hours mapping."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT product_code, process_name, standard_hours FROM standard_hours WHERE standard_hours > 0")
    mapping = {}
    for row in c.fetchall():
        key = (row[0], row[1])
        if key not in mapping:
            mapping[key] = []
        mapping[key].append(row[2])
    conn.close()
    # Average for each combo
    return {k: sum(v)/len(v) for k, v in mapping.items()}

def _parse_attendance_names(note):
    """Parse attendance_note into a list of person names, filtering out non-name annotations."""
    if not note or not note.strip():
        return []
    note = note.strip()
    parts = re.split(r'[,\uff0c\u3001\u3002.\s]+', note)
    names = []
    skip_keywords = ['H)', '\u62a5\u5e9f', '\u8c03\u673a', '\u5f85\u6599', '\u5f02\u5e38', '\u81ea\u52a8\u710a', '\u64cd\u4f5c', '\u8bbe\u5907', '\u5355\u6a21', '\u8fb9\u5f2f', '\u5b66\u4e60', '\u4e2d\u9014', '\u5206\u949f', '\u5c0f\u65f6']
    for p in parts:
        p = p.strip()
        if not p:
            continue
        if any(kw in p for kw in skip_keywords):
            continue
        if re.search(r'\d+[Hh]', p):
            continue
        p = re.sub(r'[()\uff08\uff09\d]+$', '', p).strip()
        if not p:
            continue
        if len(p) >= 2 and len(p) <= 10:
            names.append(p)
    return names

def _expand_reports_with_attendance_note(date_from, date_to):
    """Load work reports in date range, expand via attendance_note.
    Returns dict: {person_name: {'qty': N, 'hours': N, 'good_qty': N, 'efficiency': N, 'report_count': N}}
    """
    conn = get_connection()
    c = conn.cursor()
    c.execute("""SELECT operator, attendance_note, report_qty, report_hours, good_qty, efficiency, create_time
        FROM work_reports WHERE create_time BETWEEN ? AND ? AND report_hours > 0""",
        (date_from + ' 00:00:00', date_to + ' 23:59:59'))
    rows = c.fetchall()
    conn.close()
    per_person = {}
    for operator, note, qty, hours, good_qty, eff, create_time in rows:
        qty = qty or 0
        hours = hours or 0
        good_qty = good_qty or 0
        eff = eff or 0
        names = set()
        if operator:
            names.add(operator)
        if note:
            extra = _parse_attendance_names(note)
            names.update(extra)
        for nm in names:
            if nm not in per_person:
                per_person[nm] = {'qty': [], 'hours': [], 'good_qty': [], 'efficiency': []}
            per_person[nm]['qty'].append(qty)
            per_person[nm]['hours'].append(hours)
            per_person[nm]['good_qty'].append(good_qty)
            if eff > 0:
                per_person[nm]['efficiency'].append(eff)
    result = {}
    for nm, data in per_person.items():
        total_qty = sum(data['qty'])
        total_hours = sum(data['hours'])
        total_good = sum(data['good_qty'])
        effs = data['efficiency']
        avg_eff = round(sum(effs) / len(effs), 1) if effs else 0
        result[nm] = {
            'qty': total_qty, 'hours': total_hours, 'good_qty': total_good,
            'efficiency': avg_eff, 'report_count': len(data['qty'])
        }
    return result

def _expand_personal_report_by_date(name, date_from, date_to):
    """For a specific person, get daily report data including attendance_note expansion."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("""SELECT operator, attendance_note, report_qty, report_hours, good_qty, efficiency, create_time
        FROM work_reports WHERE create_time BETWEEN ? AND ? AND report_hours > 0""",
        (date_from + ' 00:00:00', date_to + ' 23:59:59'))
    rows = c.fetchall()
    conn.close()
    daily = {}
    for operator, note, qty, hours, good_qty, eff, create_time in rows:
        qty = qty or 0
        hours = hours or 0
        good_qty = good_qty or 0
        eff = eff or 0
        is_match = False
        if operator == name:
            is_match = True
        elif note:
            extra = _parse_attendance_names(note)
            if name in extra:
                is_match = True
        if not is_match:
            continue
        dt = create_time[:10]
        if dt not in daily:
            daily[dt] = {'qty': [], 'hours': [], 'good_qty': [], 'efficiency': []}
        daily[dt]['qty'].append(qty)
        daily[dt]['hours'].append(hours)
        daily[dt]['good_qty'].append(good_qty)
        if eff > 0:
            daily[dt]['efficiency'].append(eff)
    result = {}
    for dt, data in daily.items():
        result[dt] = {
            'qty': sum(data['qty']),
            'hours': round(sum(data['hours']), 1),
            'good_qty': sum(data['good_qty']),
            'efficiency': round(sum(data['efficiency']) / len(data['efficiency']), 1) if data['efficiency'] else 0
        }
    return result

@app.route('/api/reports/daily')
@login_required
def api_report_daily():
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    team_id = request.args.get('team_id', type=int)
    
    personnel_map = _get_personnel_map()
    
    # Get work reports expanded via attendance_note
    report_data = _expand_reports_with_attendance_note(date, date)
    
    conn = get_connection()
    c = conn.cursor()
    # Get attendance for the date (join by user_id via personnel)
    c.execute("""SELECT p.name, a.work_hours, a.is_overtime, a.leave_type, a.normal_hours, a.overtime_hours 
        FROM attendance a INNER JOIN personnel p ON a.user_id = p.user_id 
        WHERE a.work_date=?""", (date,))
    attend_data = {r[0]: {'hours': r[1] or 0, 'overtime': r[2], 'leave': r[3] or '', 'normal': r[4] or 0, 'ot': r[5] or 0} for r in c.fetchall()}
    conn.close()
    
    rows = []
    for name, info in personnel_map.items():
        if team_id and info['team_id'] != team_id:
            continue
        rd = report_data.get(name, {'qty': 0, 'hours': 0, 'good_qty': 0, 'efficiency': 0})
        ad = attend_data.get(name, {'hours': 0, 'overtime': 0, 'leave': '', 'normal': 0, 'ot': 0})
        # Include if has any data: work report OR attendance record
        has_attendance = name in attend_data
        has_report = name in report_data
        if not has_report and not has_attendance:
            continue
        
        good_rate = round(rd['good_qty'] / rd['qty'] * 100, 1) if rd['qty'] > 0 else 0
        eff = rd['efficiency']
        att_hrs = ad['hours']
        prod_hrs = round(rd['hours'], 1)
        util_rate = round(prod_hrs / att_hrs * 100, 1) if att_hrs > 0 else 0

        rows.append({
            'name': name,
            'department': info['department'],
            'attendance_hours': att_hrs,
            'normal_hours': ad['normal'],
            'overtime_hours': ad['ot'],
            'production_hours': prod_hrs,
            'qty': rd['qty'],
            'good_rate': good_rate,
            'efficiency': eff,
            'utilization_rate': util_rate,
            'is_overtime': ad['overtime'],
            'leave_type': ad['leave'],
        })
    
    return jsonify({'date': date, 'data': rows})

@app.route('/api/reports/personal')
@login_required
def api_report_personal():
    name = request.args.get('name', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    if not name or not date_from or not date_to:
        return jsonify({'error': '参数不完整'}), 400
    
    conn = get_connection()
    c = conn.cursor()
    std_cap = _get_standard_capacity_map()
    
    # Get work reports with efficiency
    c.execute("""SELECT create_time, SUM(report_qty), SUM(report_hours), SUM(good_qty),
        AVG(CASE WHEN efficiency > 0 THEN efficiency END) as avg_eff
        FROM work_reports WHERE operator=? AND create_time BETWEEN ? AND ? AND report_hours > 0
        GROUP BY SUBSTR(create_time, 1, 10)""",
        (name, date_from + ' 00:00:00', date_to + ' 23:59:59'))
    report_data = {}
    for r in c.fetchall():
        dt = r[0][:10]
        report_data[dt] = {'qty': r[1] or 0, 'hours': round(r[2] or 0, 1), 'good_qty': r[3] or 0, 'efficiency': round(r[4] or 0, 1)}

    # Get attendance (match by user_id via personnel)
    c.execute("""SELECT a.work_date, a.work_hours, a.is_overtime, a.leave_type, a.normal_hours, a.overtime_hours, a.check_in
        FROM attendance a INNER JOIN personnel p ON a.user_id = p.user_id 
        WHERE p.name=? AND a.work_date BETWEEN ? AND ?""", 
              (name, date_from, date_to))
    attend_data = {}
    for r in c.fetchall():
        attend_data[r[0]] = {'hours': r[1] or 0, 'overtime': r[2], 'leave': r[3] or '', 'normal': r[4] or 0, 'ot': round(r[5] or 0, 1), 'check_in': r[6] or ''}

    conn.close()

    # Build date columns
    days = []
    current = datetime.strptime(date_from, '%Y-%m-%d').date()
    end = datetime.strptime(date_to, '%Y-%m-%d').date()
    while current <= end:
        ds = current.strftime('%Y-%m-%d')
        rd = report_data.get(ds, {'qty': 0, 'hours': 0, 'good_qty': 0, 'efficiency': 0})
        ad = attend_data.get(ds, {'hours': 0, 'overtime': 0, 'leave': '', 'ot': 0, 'normal': 0, 'check_in': ''})
        good_rate = round(rd['good_qty'] / rd['qty'] * 100, 1) if rd['qty'] > 0 else 0
        util_rate = round(rd['hours'] / ad['hours'] * 100, 1) if ad['hours'] > 0 else 0
        days.append({
            'date': ds,
            'qty': rd['qty'],
            'hours': rd['hours'],
            'good_rate': good_rate,
            'efficiency': rd['efficiency'],
            'attendance_hours': ad['hours'],
            'overtime_hours': ad['ot'],
            'normal_hours': round(ad['normal'], 1),
            'utilization_rate': util_rate,
            'leave_type': ad['leave'],
            'status': '请假' if ad['leave'] else ('迟到' if ad['check_in'] and ad['check_in'][11:16]>'08:00' else ('加班' if ad['ot']>0 else ('正常' if ad['hours']>0 else ('旷工' if ad['check_in'] or rd['qty']>0 or rd['hours']>0 else ('休息' if datetime.strptime(ds, '%Y-%m-%d').weekday()>=5 else '无数据'))))),
        })
        current += timedelta(days=1)
    
    return jsonify({'name': name, 'days': days})

@app.route('/api/reports/weekly')
@login_required
def api_report_weekly():
    week_start = request.args.get('week_start', '')
    team_id = request.args.get('team_id', type=int)
    if not week_start:
        today = datetime.now().date()
    else:
        today = datetime.strptime(week_start, '%Y-%m-%d').date()
    monday = today - timedelta(days=today.weekday())
    week_start = monday.strftime('%Y-%m-%d')
    week_end = (monday + timedelta(days=6)).strftime('%Y-%m-%d')
    
    personnel_map = _get_personnel_map()
    std_cap = _get_standard_capacity_map()
    
    # Work reports for the week expanded via attendance_note
    expanded = _expand_reports_with_attendance_note(week_start, week_end)
    report_data = {}
    for nm, rd in expanded.items():
        report_data[nm] = {'qty': rd['qty'], 'hours': rd['hours'], 'good_qty': rd['good_qty'], 'days': rd['report_count'], 'efficiency': rd['efficiency']}
    
    conn = get_connection()
    c = conn.cursor()
    # Attendance for the week
    c.execute("""SELECT p.name, SUM(a.work_hours), SUM(a.is_overtime), COUNT(*), SUM(a.normal_hours), SUM(a.overtime_hours) FROM attendance a INNER JOIN personnel p ON a.user_id = p.user_id WHERE a.work_date BETWEEN ? AND ? GROUP BY p.name""",
              (week_start, week_end))
    attend_data = {}
    for r in c.fetchall():
        attend_data[r[0]] = {'total_hours': r[1] or 0, 'overtime_days': r[2] or 0, 'work_days': r[3] or 0, 'normal': r[4] or 0, 'ot': r[5] or 0}
    conn.close()
    
    rows = []
    for name, info in personnel_map.items():
        if team_id and info['team_id'] != team_id:
            continue
        rd = report_data.get(name, {'qty': 0, 'hours': 0, 'good_qty': 0, 'days': 0})
        ad = attend_data.get(name, {'total_hours': 0, 'overtime_days': 0, 'work_days': 0, 'normal': 0, 'ot': 0})
        if rd['qty'] == 0 and rd['hours'] == 0 and ad['total_hours'] == 0:
            continue
        
        good_rate = round(rd['good_qty'] / rd['qty'] * 100, 1) if rd['qty'] > 0 else 0
        eff = rd.get('efficiency', 0)
        util_rate = round(rd['hours'] / ad['total_hours'] * 100, 1) if ad['total_hours'] > 0 else 0
        daily_avg = round(rd['qty'] / rd['days'], 1) if rd['days'] > 0 else 0
        
        rows.append({
            'name': name, 'department': info['department'],
            'work_days': ad['work_days'],
            'total_attendance_hours': ad['total_hours'],
            'total_production_hours': round(rd['hours'], 1),
            'total_qty': rd['qty'],
            'good_rate': good_rate,
            'avg_efficiency': eff,
            'normal_hours': round(ad['normal'], 1),
            'overtime_hours': round(ad['ot'], 1),
            'normal_hours': round(ad['normal'], 1),
            'overtime_hours': round(ad['ot'], 1),
            'overtime_days': ad['overtime_days'],
            'utilization_rate': util_rate,
            'daily_avg_qty': daily_avg,
        })
    
    return jsonify({'week_start': week_start, 'week_end': week_end, 'data': rows})

@app.route('/api/reports/monthly')
@login_required
def api_report_monthly():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    team_id = request.args.get('team_id', type=int)
    date_from = month + '-01'
    import calendar
    y, m = map(int, month.split('-'))
    last_day = calendar.monthrange(y, m)[1]
    date_to = f"{month}-{last_day:02d}"
    
    personnel_map = _get_personnel_map()
    std_cap = _get_standard_capacity_map()
    
    # Work reports expanded via attendance_note
    expanded = _expand_reports_with_attendance_note(date_from, date_to)
    report_data = {}
    for nm, rd in expanded.items():
        report_data[nm] = {'qty': rd['qty'], 'hours': rd['hours'], 'good_qty': rd['good_qty'], 'days': rd['report_count'], 'efficiency': rd['efficiency']}
    
    conn = get_connection()
    c = conn.cursor()
    # Attendance
    c.execute("""SELECT p.name, SUM(a.work_hours), SUM(a.is_overtime), COUNT(*), SUM(CASE WHEN a.leave_type != '' AND a.leave_type IS NOT NULL THEN 1 ELSE 0 END), SUM(a.normal_hours), SUM(a.overtime_hours) FROM attendance a INNER JOIN personnel p ON a.user_id = p.user_id WHERE a.work_date BETWEEN ? AND ? GROUP BY p.name""",
              (date_from, date_to))
    attend_data = {}
    for r in c.fetchall():
        attend_data[r[0]] = {'total_hours': r[1] or 0, 'overtime_days': r[2] or 0, 'work_days': r[3] or 0, 'leave_days': r[4] or 0, 'normal': r[5] or 0, 'ot': r[6] or 0}
    
    # Planned qty from schedules
    c.execute("SELECT team_id, SUM(quantity) FROM schedules WHERE schedule_date BETWEEN ? AND ? GROUP BY team_id",
              (date_from, date_to))
    planned_by_team = {r[0]: r[1] for r in c.fetchall()}
    
    conn.close()
    
    rows = []
    for name, info in personnel_map.items():
        if team_id and info['team_id'] != team_id:
            continue
        rd = report_data.get(name, {'qty': 0, 'hours': 0, 'good_qty': 0, 'days': 0})
        ad = attend_data.get(name, {'total_hours': 0, 'overtime_days': 0, 'work_days': 0, 'leave_days': 0, 'normal': 0, 'ot': 0})
        if rd['qty'] == 0 and rd['hours'] == 0 and ad['total_hours'] == 0:
            continue
        
        good_rate = round(rd['good_qty'] / rd['qty'] * 100, 1) if rd['qty'] > 0 else 0
        eff = rd.get('efficiency', 0)
        util_rate = round(rd['hours'] / ad['total_hours'] * 100, 1) if ad['total_hours'] > 0 else 0
        daily_avg = round(rd['qty'] / rd['days'], 1) if rd['days'] > 0 else 0
        planned = planned_by_team.get(info['team_id'], 0)
        achieve_rate = round(rd['qty'] / planned * 100, 1) if planned > 0 else 0
        
        rows.append({
            'name': name, 'department': info['department'],
            'work_days': ad['work_days'],
            'total_attendance_hours': ad['total_hours'],
            'total_production_hours': round(rd['hours'], 1),
            'total_qty': rd['qty'],
            'good_rate': good_rate,
            'avg_efficiency': eff,
            'normal_hours': round(ad['normal'], 1),
            'overtime_hours': round(ad['ot'], 1),
            'normal_hours': round(ad['normal'], 1),
            'overtime_hours': round(ad['ot'], 1),
            'overtime_days': ad['overtime_days'],
            'utilization_rate': util_rate,
            'daily_avg_qty': daily_avg,
            'achieve_rate': achieve_rate,
            'leave_days': ad['leave_days'],
        })
    
    return jsonify({'month': month, 'data': rows})

@app.route('/api/attendance/sync', methods=['POST'])
@login_required
@planner_required
def api_attendance_sync():
    data = request.json
    date_from = data.get('date_from', '')
    date_to = data.get('date_to', '')
    if not date_from or not date_to:
        return jsonify({'error': '请选择日期范围'}), 400
    try:
        from utils.attendance_api import sync_attendance
        count = sync_attendance(date_from, date_to)
        return jsonify({'ok': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/attendance/test', methods=['POST'])
@login_required
@planner_required
def api_attendance_test():
    try:
        from utils.attendance_api import fetch_attendance
        from datetime import datetime, timedelta
        test_date = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
        records = fetch_attendance(test_date)
        if records is not None:
            return jsonify({'ok': True, 'count': len(records), 'date': test_date})
        else:
            return jsonify({'ok': False, 'error': '无法获取考勤数据，请检查API地址和Key'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/attendance', methods=['GET'])
@login_required
def api_attendance():
    name = request.args.get('name', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    conn = get_connection()
    c = conn.cursor()
    sql = "SELECT * FROM attendance WHERE 1=1"
    params = []
    if name:
        sql += " AND name=?"
        params.append(name)
    if date_from:
        sql += " AND work_date>=?"
        params.append(date_from)
    if date_to:
        sql += " AND work_date<=?"
        params.append(date_to)
    sql += " ORDER BY work_date DESC LIMIT 100"
    c.execute(sql, params)
    r = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(r)


# ========== Attendance Settings API ==========
@app.route('/api/attendance/settings', methods=['GET'])
@login_required
def api_attendance_settings_get():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT key, value FROM system_settings WHERE key LIKE 'attendance_%'")
    config = {row[0]: row[1] for row in c.fetchall()}
    conn.close()
    return jsonify({
        'api_url': config.get('attendance_api_url') or 'http://10.6.201.10:7777/ddkq/api/third-party/attendance',
        'api_key': config.get('attendance_api_key') or 'tk_cs_20260601',
        'auto_sync': config.get('attendance_auto_sync') or '0',
    })

@app.route('/api/attendance/settings', methods=['POST'])
@login_required
@planner_required
def api_attendance_settings_save():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES ('attendance_api_url', ?, datetime('now'))", (data.get('api_url', ''),))
    c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES ('attendance_api_key', ?, datetime('now'))", (data.get('api_key', ''),))
    c.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES ('attendance_auto_sync', ?, datetime('now'))", (str(data.get('auto_sync', '0')),))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ========== Export API ==========
@app.route('/api/export/<data_type>')
@login_required
def api_export(data_type):
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    if data_type == 'process_routes':
        ws.title = '工艺路线'
        c.execute("SELECT route_code, route_name, product_code, process_list, remark FROM process_routes ORDER BY route_code")
        ws.append(['工艺路线编号', '路线名称', '产品编码', '工序列表', '备注'])
    elif data_type == 'cycles':
        ws.title = '生产周期'
        c.execute("SELECT product_code, production_days, lead_days FROM production_cycles ORDER BY product_code")
        ws.append(['产品编码', '生产周期(天)', '提前时间(天)'])
    elif data_type == 'bom':
        ws.title = '物料清单'
        c.execute("SELECT parent_product_code, parent_product_name, child_product_code, child_product_name, quantity, unit, process_team FROM bom ORDER BY parent_product_code")
        ws.append(['父件编码', '父件名称', '子件编码', '子件名称', '用量', '单位', '生产班组'])
    elif data_type == 'product_definitions':
        ws.title = '产品定义'
        c.execute("""SELECT product_type, product_code, product_name, specifications, unit, route_code,
                     safety_stock, stock_qty, source, image_url, customer, basket_capacity 
                     FROM products ORDER BY product_code""")
        ws.append(['产品类型', '产品编号', '产品名称', '产品规格', '单位', '工艺路线', 
                   '最小安全库存', '库存数量', '产品来源', '产品图', '客户', '每筐容量'])
    elif data_type == 'work_orders':
        ws.title = '工单数据'
        c.execute("SELECT order_no, product_code, product_name, quantity, completed_qty, due_date, priority, status, process_progress, source FROM work_orders ORDER BY order_no")
        ws.append(['工单编号', '产品编码', '产品名称', '计划数量', '完成数量', '交期', '优先级', '状态', '工序进度', '来源'])
    else:
        return jsonify({'error': 'unsupported'}), 400
    for row in c.fetchall():
        ws.append(list(row))
    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, as_attachment=True, download_name=data_type + '_export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ========== User Management API ==========
@app.route('/api/users')
@login_required
@planner_required
def api_users():
    return jsonify(get_all_users())

@app.route('/api/users', methods=['POST'])
@login_required
@planner_required
def api_user_create():
    data = request.json
    uid = create_user(data['username'], data['password'], data['display_name'],
                      data.get('role', 'team'), data.get('team_id'))
    if uid:
        return jsonify({'id': uid, 'ok': True})
    return jsonify({'error': '用户名已存在'}), 400

@app.route('/api/users/<int:uid>', methods=['PUT'])
@login_required
@planner_required
def api_user_update(uid):
    data = request.json
    update_user(uid, data['display_name'], data.get('role', 'team'),
                data.get('team_id'), data.get('is_active', 1))
    return jsonify({'ok': True})

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@login_required
@planner_required
def api_user_delete(uid):
    delete_user(uid)
    return jsonify({'ok': True})

@app.route('/api/users/<int:uid>/reset-password', methods=['POST'])
@login_required
@planner_required
def api_user_reset_pwd(uid):
    data = request.json
    reset_password(uid, data.get('new_password', '123456'))
    return jsonify({'ok': True})


# ========== Supply Chain Management API ==========
@app.route('/supply-chain')
@login_required
@planner_required
def supply_chain_page():
    return render_template('supply_chain.html')

@app.route('/api/supply-chain/overview')
@login_required
@planner_required
def api_supply_overview():
    """Get supply chain overview: all items with their data."""
    conn = get_connection()
    c = conn.cursor()

    # Get all supply items
    c.execute("SELECT * FROM supply_items ORDER BY customer, product_code")
    items = [dict(row) for row in c.fetchall()]

    # Get latest snapshot data per type
    data_types = ['local_stock', '3pl_stock', 'in_transit', 'forecast', 'shipped']
    for item in items:
        pc = item['product_code']
        for dt in data_types:
            c.execute("SELECT value FROM supply_data WHERE product_code=? AND data_type=? ORDER BY imported_at DESC LIMIT 1", (pc, dt))
            row = c.fetchone()
            item[dt] = row[0] if row else 0

        # Get in-production from work_orders
        c.execute("SELECT SUM(COALESCE(quantity,0) - COALESCE(completed_qty,0)) FROM work_orders WHERE product_code=? AND status != 'completed'", (pc,))
        row = c.fetchone()
        item['in_production'] = row[0] if row and row[0] else 0

        # Get daily shipping plan (next 60 days)
        c.execute("SELECT ship_date, quantity FROM shipping_plan WHERE product_code=? ORDER BY ship_date", (pc,))
        item['daily_plan'] = [{'date': r[0], 'qty': r[1]} for r in c.fetchall()]

        # Calculate stock total
        item['stock_total'] = (item.get('local_stock', 0) or 0) + (item.get('3pl_stock', 0) or 0) + (item.get('in_transit', 0) or 0)

        # Calculate shortage indicators
        available = item['stock_total'] + item['in_production']
        cumulative_demand = 0
        item['stock_runout_date'] = None
        item['production_runout_date'] = None
        from datetime import date as _date, datetime as _dt
        today_str = _date.today().strftime('%Y-%m-%d')
        for dp in item['daily_plan']:
            if dp['date'] < today_str:
                continue
            cumulative_demand += dp['qty'] or 0
            if item['stock_runout_date'] is None and item['stock_total'] < cumulative_demand:
                item['stock_runout_date'] = dp['date']
            if item['production_runout_date'] is None and available < cumulative_demand:
                item['production_runout_date'] = dp['date']

    conn.close()
    return jsonify(items)

@app.route('/api/supply-chain/templates', methods=['GET'])
@login_required
@planner_required
def api_get_templates():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM import_templates ORDER BY name")
    r = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(r)

@app.route('/api/supply-chain/templates', methods=['POST'])
@login_required
@planner_required
def api_create_template():
    data = request.json
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO import_templates (name, template_type, header_keywords, column_mapping) VALUES (?, ?, ?, ?)",
                  (data['name'], data['template_type'],
                   json.dumps(data.get('header_keywords', {}), ensure_ascii=False),
                   json.dumps(data.get('column_mapping', {}), ensure_ascii=False)))
        conn.commit()
        return jsonify({'ok': True, 'id': c.lastrowid})
    except Exception as e:
        return jsonify({'error': str(e)}), 400
    finally:
        conn.close()

@app.route('/api/supply-chain/templates/<int:tid>', methods=['DELETE'])
@login_required
@planner_required
def api_delete_template(tid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM import_templates WHERE id=?", (tid,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/supply-chain/detect-template', methods=['POST'])
@login_required
@planner_required
def api_detect_template():
    """Upload Excel and detect which template it matches."""
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    f = request.files['file']
    import os, openpyxl
    path = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    f.save(path)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets_info = []
    for ws in wb.worksheets:
        headers = []
        for r in range(1, min(4, ws.max_row + 1)):
            row_vals = []
            for c_idx in range(1, min(ws.max_column + 1, 30)):
                v = ws.cell(row=r, column=c_idx).value
                if v is not None:
                    row_vals.append(str(v).strip())
            headers.append(row_vals)
        sheets_info.append({
            'name': ws.title,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'headers': headers
        })

    # Try to match templates
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM import_templates")
    templates = [dict(row) for row in c.fetchall()]
    conn.close()

    matches = []
    for sheet in sheets_info:
        all_text = ' '.join(sum(sheet['headers'], [])).lower()
        best_match = None
        best_score = 0
        for tmpl in templates:
            kw = json.loads(tmpl['header_keywords']) if tmpl['header_keywords'] else {}
            score = 0
            for k in kw.get('keywords', []):
                if k.lower() in all_text:
                    score += 1
            if score > best_score:
                best_score = score
                best_match = tmpl
        matches.append({
            'sheet_name': sheet['name'],
            'headers': sheet['headers'],
            'max_row': sheet['max_row'],
            'detected_template': best_match['name'] if best_match and best_score >= 2 else None,
            'template_id': best_match['id'] if best_match and best_score >= 2 else None,
            'confidence': best_score
        })

    return jsonify({'file': f.filename, 'sheets': matches, 'file_path': path})

@app.route('/api/supply-chain/import', methods=['POST'])
@login_required
@planner_required
def api_supply_import():
    """Import data using a confirmed template."""
    data = request.json
    file_path = data.get('file_path')
    template_id = data.get('template_id')
    sheet_name = data.get('sheet_name')

    if not file_path or not template_id:
        return jsonify({'error': 'file_path and template_id required'}), 400

    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM import_templates WHERE id=?", (template_id,))
    tmpl = c.fetchone()
    if not tmpl:
        conn.close()
        return jsonify({'error': 'template not found'}), 404
    tmpl = dict(tmpl)
    mapping = json.loads(tmpl['column_mapping']) if tmpl['column_mapping'] else {}

    import openpyxl, uuid
    from datetime import date as _date
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Find the sheet
    ws = None
    for s in wb.worksheets:
        if s.title == sheet_name:
            ws = s
            break
    if not ws:
        ws = wb.active

    batch_id = str(uuid.uuid4())[:8]
    count = 0
    header_row = mapping.get('header_row', 2)
    data_start = mapping.get('data_start', 4)

    code_col = mapping.get('code_col')  # column letter or index
    customer_col = mapping.get('customer_col')
    project_col = mapping.get('project_col')
    basket_col = mapping.get('basket_col')

    # Snapshot columns (single value per row)
    snapshot_cols = mapping.get('snapshot_cols', {})

    # Daily plan columns
    plan_date_row = mapping.get('plan_date_row', 1)
    plan_start_col = mapping.get('plan_start_col')
    plan_end_col = mapping.get('plan_end_col')

    # Helper to get column index from letter
    def col_idx(col):
        if col is None:
            return None
        if isinstance(col, int):
            return col
        col = col.upper()
        result = 0
        for ch in col:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result

    ci_code = col_idx(code_col)
    ci_customer = col_idx(customer_col)
    ci_project = col_idx(project_col)
    ci_basket = col_idx(basket_col)

    for r in range(data_start, ws.max_row + 1):
        code_val = ws.cell(row=r, column=ci_code).value if ci_code else None
        if not code_val:
            continue
        pc = str(code_val).strip()
        if not pc:
            continue

        customer = str(ws.cell(row=r, column=ci_customer).value).strip() if ci_customer else ''
        project = str(ws.cell(row=r, column=ci_project).value).strip() if ci_project else ''
        basket = 0
        if ci_basket:
            bv = ws.cell(row=r, column=ci_basket).value
            basket = float(bv) if bv else 0

        # Upsert supply_items
        c.execute("SELECT id FROM supply_items WHERE product_code=?", (pc,))
        if c.fetchone():
            c.execute("UPDATE supply_items SET customer=?, project=?, basket_capacity=?, updated_at=datetime('now','localtime') WHERE product_code=?",
                      (customer, project, basket, pc))
        else:
            c.execute("INSERT INTO supply_items (product_code, customer, project, basket_capacity) VALUES (?, ?, ?, ?)",
                      (pc, customer, project, basket))

        # Import snapshot data (e.g. local_stock, 3pl_stock, in_transit)
        for dtype, col_letter in snapshot_cols.items():
            ci = col_idx(col_letter)
            if ci:
                val = ws.cell(row=r, column=ci).value
                if val is not None:
                    try:
                        c.execute("INSERT INTO supply_data (product_code, data_type, value, batch_id) VALUES (?, ?, ?, ?)",
                                  (pc, dtype, float(val), batch_id))
                    except (ValueError, TypeError):
                        pass

        # Import daily plan data
        if plan_start_col and plan_end_col:
            ci_start = col_idx(plan_start_col)
            ci_end = col_idx(plan_end_col)
            for ci in range(ci_start, ci_end + 1):
                # Get date from header row
                date_val = ws.cell(row=plan_date_row, column=ci).value
                if date_val is None:
                    continue
                if hasattr(date_val, 'strftime'):
                    d_str = date_val.strftime('%Y-%m-%d')
                else:
                    d_str = str(date_val).strip()[:10]
                qty_val = ws.cell(row=r, column=ci).value
                if qty_val is not None:
                    try:
                        c.execute("INSERT INTO supply_data (product_code, data_type, data_date, value, batch_id) VALUES (?, ?, ?, ?, ?)",
                                  (pc, 'daily_plan', d_str, float(qty_val), batch_id))
                    except (ValueError, TypeError):
                        pass

        count += 1

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'count': count, 'batch_id': batch_id})

@app.route('/api/supply-chain/items', methods=['DELETE'])
@login_required
@planner_required
def api_clear_supply_data():
    """Clear all supply data (for re-import)."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM supply_data")
    c.execute("DELETE FROM supply_items")
    conn.commit()
    deleted = c.rowcount
    conn.close()
    return jsonify({'ok': True, 'deleted': deleted})

# ========== Shipping Plan API ==========
@app.route('/api/shipping-plan')
@login_required
def api_shipping_plan():
    month = request.args.get('month', '').strip()
    conn = get_connection()
    c = conn.cursor()
    if month:
        c.execute("SELECT * FROM shipping_plan WHERE plan_month=? ORDER BY customer, product_code, ship_date", (month,))
    else:
        c.execute("SELECT DISTINCT plan_month FROM shipping_plan WHERE plan_month IS NOT NULL ORDER BY plan_month DESC LIMIT 1")
        row = c.fetchone()
        if row:
            c.execute("SELECT * FROM shipping_plan WHERE plan_month=? ORDER BY customer, product_code, ship_date", (row[0],))
        else:
            c.execute("SELECT * FROM shipping_plan ORDER BY ship_date LIMIT 500")
    r = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify(r)

@app.route('/api/shipping-plan/months')
@login_required
def api_shipping_plan_months():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT DISTINCT plan_month FROM shipping_plan WHERE plan_month IS NOT NULL ORDER BY plan_month DESC")
    months = [row[0] for row in c.fetchall()]
    conn.close()
    return jsonify(months)

@app.route('/api/shipping-plan', methods=['DELETE'])
@login_required
@planner_required
def api_shipping_plan_delete():
    month = request.args.get('month', '').strip()
    conn = get_connection()
    c = conn.cursor()
    if month:
        c.execute("DELETE FROM shipping_plan WHERE plan_month=?", (month,))
    else:
        c.execute("DELETE FROM shipping_plan")
    deleted = c.rowcount
    conn.commit()
    conn.close()
    return jsonify({'deleted': deleted})

# ========== Statistics API ==========
@app.route('/api/statistics')
@login_required
def api_statistics():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    if not date_from:
        date_from = (datetime.now() - timedelta(days=6)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = datetime.now().strftime('%Y-%m-%d')
    conn = get_connection()
    c = conn.cursor()

    # 1. Load all teams
    c.execute("SELECT id, name FROM teams ORDER BY id")
    teams = [dict(row) for row in c.fetchall()]

    # 2. Load all process->team mapping once
    c.execute("SELECT process_name, team_name FROM processes WHERE team_name IS NOT NULL AND team_name != '' AND team_name != '报工权限'")
    proc_team_map = {}
    for r in c.fetchall():
        proc_team_map[r[0]] = r[1]

    # 3. Bulk load schedules for date range
    c.execute("SELECT team_id, schedule_date, product_code, process_name, quantity, capacity_per_hour FROM schedules WHERE schedule_date BETWEEN ? AND ?", (date_from, date_to))
    all_schedules = c.fetchall()

    # 4. Bulk load work_reports for date range
    c.execute("SELECT product_code, process_name, report_qty, report_hours, create_time FROM work_reports WHERE create_time BETWEEN ? AND ?",
              (date_from + ' 00:00:00', date_to + ' 23:59:59'))
    all_reports = c.fetchall()
    conn.close()

    # 5. Build schedule index: (team_id, date) -> total_qty
    sched_by_team_date = {}
    # Build schedule capacity index: (product_code, process_name, date) -> capacity
    sched_capacity = {}
    for s in all_schedules:
        tid, dt, pc, pn, qty, cap = s[0], s[1], s[2], s[3], s[4] or 0, s[5] or 0
        key = (tid, dt)
        sched_by_team_date[key] = sched_by_team_date.get(key, 0) + qty
        if cap > 0:
            sched_capacity[(pc, pn, dt)] = cap

    # 6. Build report index: (product_code, process_name, date) -> {qty, hours}
    report_by_key = {}
    for r in all_reports:
        pc, pn, qty, hrs, ct = r[0], r[1], r[2] or 0, r[3] or 0, r[4] or ''
        dt = ct[:10] if ct else ''
        key = (pc, pn, dt)
        if key not in report_by_key:
            report_by_key[key] = {'qty': 0, 'hours': 0}
        report_by_key[key]['qty'] += qty
        report_by_key[key]['hours'] += hrs

    # 7. Build team->process set
    team_procs = {}
    for pn, tn in proc_team_map.items():
        for team in teams:
            if team['name'] in tn:
                team_procs.setdefault(team['id'], set()).add(pn)

    # 8. Generate result
    result = {}
    for team in teams:
        tid = team['id']
        tname = team['name']
        procs = team_procs.get(tid, set())
        days = []
        current = datetime.strptime(date_from, '%Y-%m-%d').date()
        end = datetime.strptime(date_to, '%Y-%m-%d').date()
        while current <= end:
            ds = current.strftime('%Y-%m-%d')
            planned = sched_by_team_date.get((tid, ds), 0)
            # Sum completed qty for this team's processes on this date
            completed = 0
            efficiencies = []
            for (pc, pn, dt), data in report_by_key.items():
                if dt == ds and pn in procs:
                    completed += data['qty']
                    if data['hours'] > 0:
                        cap = sched_capacity.get((pc, pn, ds), 0)
                        if cap > 0:
                            eff = round((data['qty'] / data['hours']) / cap * 100, 1)
                            efficiencies.append(eff)
            rate = round(completed / planned * 100, 1) if planned > 0 else 0
            eff_avg = round(sum(efficiencies) / len(efficiencies), 1) if efficiencies else 0
            days.append({'date': ds, 'planned_qty': planned, 'completed_qty': completed, 'completion_rate': rate, 'efficiency': eff_avg})
            current += timedelta(days=1)
        result[tid] = {'team_name': tname, 'days': days}

    return jsonify(result)



# ========== Plan Record APIs ==========

@app.route('/api/plan/today-progress')
@login_required
def api_plan_today_progress():
    """Real-time daily plan progress for all teams."""
    today = request.args.get('date', '') or datetime.now().strftime('%Y-%m-%d')
    conn = get_connection()
    c = conn.cursor()

    # Get today's approved/draft plans
    c.execute("""SELECT dp.id, dp.team_id, t.name as team_name, dp.status, dp.plan_date
        FROM daily_plans dp LEFT JOIN teams t ON dp.team_id=t.id
        WHERE dp.plan_date=? AND dp.status IN ('approved','draft')""", (today,))
    plans = [dict(row) for row in c.fetchall()]

    # Get scheduled tasks for today
    c.execute("""SELECT s.team_id, t.name as team_name, SUM(s.quantity) as planned_qty,
        COUNT(*) as task_count
        FROM schedules s LEFT JOIN teams t ON s.team_id=t.id
        WHERE s.schedule_date=?
        GROUP BY s.team_id""", (today,))
    sched_by_team = {r[0]: {'planned_qty': r[2] or 0, 'task_count': r[3]} for r in c.fetchall()}

    # Get actual production from work_reports today
    c.execute("""SELECT p.id as team_id, SUM(wr.report_qty) as actual_qty,
        SUM(wr.report_hours) as actual_hours
        FROM work_reports wr
        INNER JOIN processes pr ON wr.process_name=pr.process_name
        INNER JOIN teams p ON pr.team_name LIKE '%' || p.name || '%'
        WHERE wr.create_time LIKE ? AND wr.report_hours > 0
        GROUP BY p.id""", (today + '%',))
    actual_by_team = {r[0]: {'actual_qty': r[1] or 0, 'actual_hours': r[2] or 0} for r in c.fetchall()}

    # Get all teams
    c.execute("SELECT id, name FROM teams ORDER BY id")
    teams = c.fetchall()
    conn.close()

    result = []
    for tid, tname in teams:
        sched = sched_by_team.get(tid, {'planned_qty': 0, 'task_count': 0})
        actual = actual_by_team.get(tid, {'actual_qty': 0, 'actual_hours': 0})
        planned = sched['planned_qty']
        completed = actual['actual_qty']
        rate = round(completed / planned * 100, 1) if planned > 0 else 0
        status = 'completed' if rate >= 100 and planned > 0 else ('in_progress' if completed > 0 else 'not_started')
        result.append({
            'team_id': tid,
            'team_name': tname,
            'planned_qty': planned,
            'actual_qty': completed,
            'completion_rate': rate,
            'task_count': sched['task_count'],
            'actual_hours': round(actual['actual_hours'], 1),
            'status': status,
        })
    return jsonify({'date': today, 'teams': result})


@app.route('/api/plan/<int:plan_id>/gantt')
@login_required
def api_plan_gantt(plan_id):
    """Get Gantt data for a specific plan: planned schedules + actual work reports."""
    conn = get_connection()
    c = conn.cursor()

    # Get plan info
    c.execute("""SELECT dp.*, t.name as team_name FROM daily_plans dp
        LEFT JOIN teams t ON dp.team_id=t.id WHERE dp.id=?""", (plan_id,))
    plan = dict(c.fetchone()) if c.rowcount else None
    if not plan:
        conn.close()
        return jsonify({'error': 'Plan not found'}), 404

    plan_date = plan['plan_date']
    team_id = plan['team_id']

    # Get planned schedules
    c.execute("""SELECT s.*, e.equipment_name, e.equipment_code
        FROM schedules s
        LEFT JOIN equipments e ON s.equipment_id=e.id
        WHERE s.team_id=? AND s.schedule_date=?
        ORDER BY s.start_time""", (team_id, plan_date))
    schedules = [dict(row) for row in c.fetchall()]

    # Get actual work reports for this team on this date
    c.execute("""SELECT wr.process_name, wr.equipment, wr.order_no,
        SUM(wr.report_qty) as qty, SUM(wr.report_hours) as hours,
        MIN(wr.start_time) as first_start, MAX(wr.end_time) as last_end
        FROM work_reports wr
        INNER JOIN processes pr ON wr.process_name=pr.process_name
        INNER JOIN teams t ON pr.team_name LIKE '%' || t.name || '%'
        WHERE t.id=? AND wr.create_time LIKE ? AND wr.report_hours > 0
        GROUP BY wr.process_name, wr.equipment, wr.order_no""",
              (team_id, plan_date + '%'))
    reports = [dict(row) for row in c.fetchall()]

    # Summary
    total_planned = sum(s.get('quantity', 0) or 0 for s in schedules)
    total_completed = sum(r.get('qty', 0) or 0 for r in reports)
    rate = round(total_completed / total_planned * 100, 1) if total_planned > 0 else 0

    conn.close()
    return jsonify({
        'plan': plan,
        'schedules': schedules,
        'reports': reports,
        'summary': {
            'total_planned': total_planned,
            'total_completed': total_completed,
            'completion_rate': rate,
        }
    })


@app.route('/api/plan/today-gantt/<int:team_id>')
@login_required
def api_plan_today_gantt(team_id):
    """Get Gantt data for a team on a given date: planned schedules + actual reports."""
    today = request.args.get('date', '') or datetime.now().strftime('%Y-%m-%d')
    conn = get_connection()
    c = conn.cursor()

    # Get planned schedules
    c.execute("""SELECT s.*, e.equipment_name, e.equipment_code
        FROM schedules s
        LEFT JOIN equipments e ON s.equipment_id=e.id
        WHERE s.team_id=? AND s.schedule_date=?
        ORDER BY s.start_time""", (team_id, today))
    schedules = [dict(row) for row in c.fetchall()]

    # Get actual reports today (individual rows for time-based overlay)
    c.execute("""SELECT wr.process_name, wr.equipment, wr.order_no, wr.product_name,
        wr.report_qty as qty, wr.start_time, wr.end_time
        FROM work_reports wr
        INNER JOIN processes pr ON wr.process_name=pr.process_name
        INNER JOIN teams t ON pr.team_name LIKE '%' || t.name || '%'
        WHERE t.id=? AND wr.create_time LIKE ?
        ORDER BY wr.start_time""",
              (team_id, today + '%'))
    reports = [dict(row) for row in c.fetchall()]

    # Team name
    c.execute("SELECT name FROM teams WHERE id=?", (team_id,))
    row = c.fetchone()
    team_name = row[0] if row else ''

    total_planned = sum(s.get('quantity', 0) or 0 for s in schedules)
    total_completed = sum(r.get('qty', 0) or 0 for r in reports)
    rate = round(total_completed / total_planned * 100, 1) if total_planned > 0 else 0

    conn.close()
    return jsonify({
        'date': today,
        'team_id': team_id,
        'team_name': team_name,
        'schedules': schedules,
        'reports': reports,
        'summary': {
            'total_planned': total_planned,
            'total_completed': total_completed,
            'completion_rate': rate,
        }
    })


@app.route('/api/plan/completion-list')
@login_required
def api_plan_completion_list():
    """List plans with completion data."""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    team_id = request.args.get('team_id', type=int)
    status = request.args.get('status', '')

    if not date_from:
        date_from = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = datetime.now().strftime('%Y-%m-%d')

    conn = get_connection()
    c = conn.cursor()

    q = """SELECT dp.id, dp.plan_date, dp.status, t.name as team_name, dp.team_id,
        (SELECT SUM(s.quantity) FROM schedules s WHERE s.daily_plan_id=dp.id) as planned_qty,
        (SELECT COUNT(*) FROM schedules s WHERE s.daily_plan_id=dp.id) as task_count
        FROM daily_plans dp LEFT JOIN teams t ON dp.team_id=t.id
        WHERE dp.plan_date BETWEEN ? AND ?"""
    params = [date_from, date_to]

    if team_id:
        q += " AND dp.team_id=?"
        params.append(team_id)
    if status:
        q += " AND dp.status=?"
        params.append(status)

    q += " ORDER BY dp.plan_date DESC, dp.team_id"
    c.execute(q, params)
    plans = []
    for row in c.fetchall():
        p = dict(row)
        # Get actual completed qty from work_reports
        if p['plan_date']:
            c.execute("""SELECT SUM(wr.report_qty)
                FROM work_reports wr
                INNER JOIN processes pr ON wr.process_name=pr.process_name
                INNER JOIN teams t ON pr.team_name LIKE '%' || t.name || '%'
                WHERE t.id=? AND wr.create_time LIKE ? AND wr.report_hours > 0""",
                      (p['team_id'], p['plan_date'] + '%'))
            r = c.fetchone()
            p['completed_qty'] = r[0] or 0 if r else 0
            p['completion_rate'] = round(p['completed_qty'] / p['planned_qty'] * 100, 1) if p['planned_qty'] and p['planned_qty'] > 0 else 0
        else:
            p['completed_qty'] = 0
            p['completion_rate'] = 0
        plans.append(p)

    conn.close()
    return jsonify(plans)

# ========== Team Statistics APIs ==========

def _get_team_process_ids(team_id):
    """Get list of process_names that belong to a team via LIKE matching."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT name FROM teams WHERE id=?", (team_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return set()
    team_name = row[0]
    c.execute("SELECT process_name FROM processes WHERE team_name LIKE ?", ('%' + team_name + '%',))
    result = set(r[0] for r in c.fetchall())
    conn.close()
    return result

def _get_all_team_process_map():
    """Return {team_id: set of process_names}."""
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, name FROM teams ORDER BY id")
    teams = c.fetchall()
    result = {}
    for tid, tname in teams:
        c.execute("SELECT process_name FROM processes WHERE team_name LIKE ?", ('%' + tname + '%',))
        result[tid] = set(r[0] for r in c.fetchall())
    conn.close()
    return result

def _build_attendance_by_team_date(date_from, date_to):
    """Return {(team_id, date): attendance info}"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("""SELECT p.team_id, a.work_date,
        COUNT(DISTINCT a.user_id) as total_users,
        SUM(CASE WHEN a.is_overtime = 1 THEN 1 ELSE 0 END) as ot_users,
        SUM(a.work_hours) as total_hours,
        SUM(a.normal_hours) as normal_hours,
        SUM(a.overtime_hours) as ot_hours
        FROM attendance a
        INNER JOIN personnel p ON a.user_id = p.user_id
        WHERE a.work_date BETWEEN ? AND ?
        GROUP BY p.team_id, a.work_date""", (date_from, date_to))
    result = {}
    for r in c.fetchall():
        result[(r[0], r[1])] = {
            'total_users': r[2],
            'ot_users': r[3],
            'total_hours': r[4] or 0,
            'normal_hours': r[5] or 0,
            'ot_hours': r[6] or 0,
        }
    conn.close()
    return result

def _build_schedules_by_team_date(date_from, date_to):
    """Return {(team_id, date): planned_qty} from production_requirements."""
    conn = get_connection()
    c = conn.cursor()
    # Build team name -> id map
    c.execute("SELECT id, name FROM teams")
    team_map = {r[1]: r[0] for r in c.fetchall()}
    # Use production_requirements as planned qty source
    c.execute("""SELECT team_name, required_date, SUM(required_quantity)
        FROM production_requirements
        WHERE required_date BETWEEN ? AND ?
        GROUP BY team_name, required_date""",
              (date_from, date_to))
    result = {}
    for r in c.fetchall():
        tn, dt, qty = r[0], r[1], r[2] or 0
        if not tn or not dt:
            continue
        # Match team by exact name or inclusion
        matched_tid = None
        for tname, tid in team_map.items():
            if tn == tname or tname in tn:
                matched_tid = tid
                break
        if matched_tid is not None:
            key = (matched_tid, dt)
            result[key] = result.get(key, 0) + qty
    conn.close()
    return result

def _build_reports_by_proc_date(date_from, date_to):
    """Return {(process_name, date): {qty, good_qty, hours, efficiency}}"""
    conn = get_connection()
    c = conn.cursor()
    c.execute("""SELECT process_name, SUBSTR(create_time,1,10) as dt,
        SUM(report_qty), SUM(good_qty), SUM(report_hours),
        AVG(CASE WHEN efficiency > 0 THEN efficiency END)
        FROM work_reports
        WHERE create_time BETWEEN ? AND ?
        GROUP BY process_name, dt""",
              (date_from + ' 00:00:00', date_to + ' 23:59:59'))
    result = {}
    for r in c.fetchall():
        result[(r[0], r[1])] = {
            'qty': r[2] or 0,
            'good_qty': r[3] or 0,
            'hours': r[4] or 0,
            'efficiency': r[5] or 0,
        }
    conn.close()
    return result

def _compute_team_day_stats(team_proc_set, sched_data, report_data, attend_data, team_id, date_str):
    """Compute all metrics for one team on one day."""
    planned = sched_data.get((team_id, date_str), 0)
    completed = 0
    good_qty = 0
    total_hours = 0
    efficiencies = []
    for (pn, dt), rd in report_data.items():
        if dt == date_str and pn in team_proc_set:
            completed += rd['qty']
            good_qty += rd['good_qty']
            total_hours += rd['hours']
            if rd['efficiency'] > 0:
                efficiencies.append(rd['efficiency'])

    completion_rate = round(completed / planned * 100, 1) if planned > 0 else 0
    prod_eff = round(sum(efficiencies) / len(efficiencies), 1) if efficiencies else 0
    good_rate = round(good_qty / completed * 100, 1) if completed > 0 else 0

    att = attend_data.get((team_id, date_str), {})
    total_users = att.get('total_users', 0)
    ot_users = att.get('ot_users', 0)
    ot_ratio = round(ot_users / total_users * 100, 1) if total_users > 0 else 0
    att_hours = att.get('total_hours', 0)
    util_rate = round(total_hours / att_hours * 100, 1) if att_hours > 0 else 0

    return {
        'planned_qty': planned,
        'completed_qty': completed,
        'completion_rate': completion_rate,
        'production_efficiency': prod_eff,
        'good_rate': good_rate,
        'overtime_count': ot_users,
        'overtime_ratio': ot_ratio,
        'total_attendance_hours': round(att_hours, 1),
        'total_production_hours': round(total_hours, 1),
        'utilization_rate': util_rate,
    }

def _compute_team_summary(days_data):
    """Compute weighted summary from a list of day stats."""
    total_planned = sum(d['planned_qty'] for d in days_data)
    total_completed = sum(d['completed_qty'] for d in days_data)
    total_good = 0
    total_prod_hours = 0
    total_att_hours = 0
    eff_vals = []
    att_day_count = 0
    ot_day_count = 0

    for d in days_data:
        total_prod_hours += d['total_production_hours']
        total_att_hours += d['total_attendance_hours']
        if d['total_attendance_hours'] > 0:
            att_day_count += 1
        if d['overtime_count'] > 0:
            ot_day_count += 1
        if d['completed_qty'] > 0 and d['good_rate'] > 0:
            total_good += int(d['good_rate'] / 100 * d['completed_qty'])
        if d['production_efficiency'] > 0:
            eff_vals.append(d['production_efficiency'])

    return {
        'completion_rate': round(total_completed / total_planned * 100, 1) if total_planned > 0 else 0,
        'production_efficiency': round(sum(eff_vals) / len(eff_vals), 1) if eff_vals else 0,
        'good_rate': round(total_good / total_completed * 100, 1) if total_completed > 0 else 0,
        'overtime_ratio': round(ot_day_count / att_day_count * 100, 1) if att_day_count > 0 else 0,
        'utilization_rate': round(total_prod_hours / total_att_hours * 100, 1) if total_att_hours > 0 else 0,
        'planned_qty': total_planned,
        'completed_qty': total_completed,
    }


@app.route('/api/stats/daily')
@login_required
def api_stats_daily():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    if not date_from or not date_to:
        return jsonify({'error': 'date_from and date_to required'}), 400

    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, name FROM teams ORDER BY id")
    teams_list = [{'id': r[0], 'name': r[1]} for r in c.fetchall()]
    member_counts = {}
    c.execute("SELECT team_id, COUNT(*) FROM personnel WHERE is_active=1 GROUP BY team_id")
    for r in c.fetchall():
        member_counts[r[0]] = r[1]
    conn.close()

    team_proc_map = _get_all_team_process_map()
    sched_data = _build_schedules_by_team_date(date_from, date_to)
    report_data = _build_reports_by_proc_date(date_from, date_to)
    attend_data = _build_attendance_by_team_date(date_from, date_to)

    result_teams = []
    for team in teams_list:
        tid = team['id']
        team_procs = team_proc_map.get(tid, set())
        days = []
        current = datetime.strptime(date_from, '%Y-%m-%d').date()
        end = datetime.strptime(date_to, '%Y-%m-%d').date()
        while current <= end:
            ds = current.strftime('%Y-%m-%d')
            day_stats = _compute_team_day_stats(team_procs, sched_data, report_data, attend_data, tid, ds)
            day_stats['date'] = ds
            days.append(day_stats)
            current += timedelta(days=1)
        summary = _compute_team_summary(days)
        result_teams.append({
            'id': tid, 'name': team['name'],
            'member_count': member_counts.get(tid, 0),
            'summary': summary, 'days': days,
        })
    return jsonify({'mode': 'daily', 'date_from': date_from, 'date_to': date_to, 'teams': result_teams})


@app.route('/api/stats/weekly')
@login_required
def api_stats_weekly():
    week_start = request.args.get('week_start', '')
    if not week_start:
        return jsonify({'error': 'week_start required'}), 400
    monday = datetime.strptime(week_start, '%Y-%m-%d').date()
    weeks = []
    for i in range(3, -1, -1):
        ws = monday - timedelta(weeks=i)
        we = ws + timedelta(days=6)
        weeks.append((ws.strftime('%Y-%m-%d'), we.strftime('%Y-%m-%d')))
    date_from = weeks[0][0]
    date_to = weeks[-1][1]

    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, name FROM teams ORDER BY id")
    teams_list = [{'id': r[0], 'name': r[1]} for r in c.fetchall()]
    member_counts = {}
    c.execute("SELECT team_id, COUNT(*) FROM personnel WHERE is_active=1 GROUP BY team_id")
    for r in c.fetchall():
        member_counts[r[0]] = r[1]
    conn.close()

    team_proc_map = _get_all_team_process_map()
    sched_data = _build_schedules_by_team_date(date_from, date_to)
    report_data = _build_reports_by_proc_date(date_from, date_to)
    attend_data = _build_attendance_by_team_date(date_from, date_to)

    result_teams = []
    for team in teams_list:
        tid = team['id']
        team_procs = team_proc_map.get(tid, set())
        weeks_data = []
        for ws, we in weeks:
            week_days = []
            current = datetime.strptime(ws, '%Y-%m-%d').date()
            end = datetime.strptime(we, '%Y-%m-%d').date()
            while current <= end:
                ds = current.strftime('%Y-%m-%d')
                day_stats = _compute_team_day_stats(team_procs, sched_data, report_data, attend_data, tid, ds)
                week_days.append(day_stats)
                current += timedelta(days=1)
            ws_summary = _compute_team_summary(week_days)
            ws_summary['week_start'] = ws
            ws_summary['week_end'] = we
            weeks_data.append(ws_summary)
        overall = _compute_team_summary([{'planned_qty': w['planned_qty'], 'completed_qty': w['completed_qty'],
            'completion_rate': w['completion_rate'], 'production_efficiency': w['production_efficiency'],
            'good_rate': w['good_rate'], 'overtime_ratio': w['overtime_ratio'],
            'utilization_rate': w['utilization_rate'], 'total_production_hours': 0,
            'total_attendance_hours': 0, 'overtime_count': 0} for w in weeks_data])
        overall['planned_qty'] = sum(w['planned_qty'] for w in weeks_data)
        overall['completed_qty'] = sum(w['completed_qty'] for w in weeks_data)
        if overall['planned_qty'] > 0:
            overall['completion_rate'] = round(overall['completed_qty'] / overall['planned_qty'] * 100, 1)
        result_teams.append({
            'id': tid, 'name': team['name'],
            'member_count': member_counts.get(tid, 0),
            'summary': overall, 'weeks': weeks_data,
        })
    return jsonify({'mode': 'weekly', 'week_start': week_start, 'teams': result_teams})


@app.route('/api/stats/monthly')
@login_required
def api_stats_monthly():
    month = request.args.get('month', '')
    if not month:
        return jsonify({'error': 'month required (YYYY-MM)'}), 400
    y, m_val = map(int, month.split('-'))
    import calendar
    last_day = calendar.monthrange(y, m_val)[1]
    date_from = month + '-01'
    date_to = '%s-%02d' % (month, last_day)

    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, name FROM teams ORDER BY id")
    teams_list = [{'id': r[0], 'name': r[1]} for r in c.fetchall()]
    member_counts = {}
    c.execute("SELECT team_id, COUNT(*) FROM personnel WHERE is_active=1 GROUP BY team_id")
    for r in c.fetchall():
        member_counts[r[0]] = r[1]
    conn.close()

    team_proc_map = _get_all_team_process_map()
    sched_data = _build_schedules_by_team_date(date_from, date_to)
    report_data = _build_reports_by_proc_date(date_from, date_to)
    attend_data = _build_attendance_by_team_date(date_from, date_to)

    result_teams = []
    for team in teams_list:
        tid = team['id']
        team_procs = team_proc_map.get(tid, set())
        days = []
        current = datetime.strptime(date_from, '%Y-%m-%d').date()
        end = datetime.strptime(date_to, '%Y-%m-%d').date()
        while current <= end:
            ds = current.strftime('%Y-%m-%d')
            day_stats = _compute_team_day_stats(team_procs, sched_data, report_data, attend_data, tid, ds)
            day_stats['date'] = ds
            days.append(day_stats)
            current += timedelta(days=1)
        summary = _compute_team_summary(days)
        result_teams.append({
            'id': tid, 'name': team['name'],
            'member_count': member_counts.get(tid, 0),
            'summary': summary,
            'months': [{'month': month, 'planned_qty': summary['planned_qty'],
                'completed_qty': summary['completed_qty'],
                'completion_rate': summary['completion_rate'],
                'production_efficiency': summary['production_efficiency'],
                'good_rate': summary['good_rate'],
                'overtime_ratio': summary['overtime_ratio'],
                'utilization_rate': summary['utilization_rate']}],
        })
    return jsonify({'mode': 'monthly', 'month': month, 'teams': result_teams})

@app.route('/api/workshop/3d-status')
@login_required
def api_workshop_3d_status():
    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, equipment_code, equipment_name, team_id, status, 
                   pos_x, pos_y, pos_z, rotation_z, model_type
            FROM equipments
        """)
        equipments = []
        for row in cursor.fetchall():
            eq = {
                'id': row[0],
                'code': row[1],
                'name': row[2],
                'team_id': row[3],
                'team_name': '',  # will be filled later
                'status': row[4] or 'normal',
                'position': {'x': row[5], 'y': row[6], 'z': row[7]},
                'rotation': row[8],
                'model': row[9]
            }
            equipments.append(eq)

        # Get teams for team name lookup
        cursor.execute("SELECT id, name FROM teams")
        teams_map = {row[0]: row[1] for row in cursor.fetchall()}

        # Get latest work report for each equipment (qty + efficiency)
        cursor.execute("""
            SELECT wr.equipment, wr.report_qty, wr.efficiency, wr.order_no, wr.process_name, wr.create_time, wr.product_name
            FROM work_reports wr
            INNER JOIN (
                SELECT equipment, MAX(create_time) as max_time
                FROM work_reports
                WHERE equipment IS NOT NULL AND equipment != ''
                GROUP BY equipment
            ) latest ON wr.equipment = latest.equipment AND wr.create_time = latest.max_time
            WHERE wr.equipment IS NOT NULL AND wr.equipment != ''
        """)
        report_map = {}
        for row in cursor.fetchall():
            report_map[row[0]] = {
                'report_qty': row[1],
                'efficiency': row[2],
                'order_no': row[3],
                'process_name': row[4],
                'report_time': row[5],
                'product_name': row[6]
            }
        conn.close()

        # Build normalized index for fuzzy matching (handles Chinese full-width, mixed separators, etc.)
        import unicodedata as _ud

        def _norm_eq(s):
            """Normalize equipment name: NFKC, lowercase, strip annotations, unify separators"""
            if not s: return ''
            s = _ud.normalize('NFKC', s).strip().lower()
            # Strip Chinese annotations like "???8:00-9:10"
            s = re.split(r'[?,??;?:?(]', s)[0]
            s = s.replace('/', '-').replace('.', '-').replace('_', '-')
            s = s.replace('?', '').replace('?', '').replace('?', '')
            while '--' in s: s = s.replace('--', '-')
            return s.replace(' ', '')

        def _strip_non_alnum(s):
            """Keep only a-z and 0-9 for maximum fuzzy matching"""
            return re.sub(r'[^a-z0-9]', '', _norm_eq(s)) if s else ''

        # Build multi-level indexes from report_map
        norm_report_map = {}     # normalized -> report data
        stripped_report_map = {} # alphanumeric-only -> report data
        for key, val in report_map.items():
            nk = _norm_eq(key)
            if nk and nk not in norm_report_map:
                norm_report_map[nk] = val
            sk = _strip_non_alnum(key)
            if sk and sk not in stripped_report_map:
                stripped_report_map[sk] = val

        # Attach report data to equipments with 3-level fuzzy matching
        for eq in equipments:
            rpt = None
            # Level 1: Exact match by name or code
            rpt = report_map.get(eq['name']) or report_map.get(eq['code'])
            if not rpt:
                # Level 2: Normalized match (handles ?/?, full-width, separators)
                n_name = _norm_eq(eq['name'])
                n_code = _norm_eq(eq['code'])
                rpt = norm_report_map.get(n_name) or norm_report_map.get(n_code)
            if not rpt:
                # Level 3: Alphanumeric-only match (handles DJ?Dw8?2 -> djxdw82)
                s_name = _strip_non_alnum(eq['name'])
                s_code = _strip_non_alnum(eq['code'])
                rpt = stripped_report_map.get(s_name) or stripped_report_map.get(s_code)
            if rpt:
                eq['report_qty'] = rpt['report_qty']
                eq['efficiency'] = rpt['efficiency']
                eq['order_no'] = rpt['order_no']
                eq['process_name'] = rpt['process_name']
                eq['report_time'] = rpt['report_time']

                eq['team_name'] = teams_map.get(eq['team_id'], '')
                eq['product_name'] = rpt.get('product_name', '')

        return jsonify({'success': True, 'data': equipments})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workshop/refresh-equipment-status', methods=['POST'])
@login_required
@planner_required
def api_refresh_equipment_status():
    """手动触发设备状态刷新（根据报工数据）"""
    try:
        updated = _refresh_equipment_status()
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Workshop Layout API routes

@app.route('/workshop-layout-page')
@login_required
def workshop_layout_page():
    return render_template('workshop_layout.html')

@app.route('/api/workshop-layouts', methods=['GET'])
@login_required
def api_get_workshop_layouts():
    try:
        team_id = request.args.get('team_id')
        conn = get_connection()
        c = conn.cursor()
        if team_id:
            c.execute("SELECT * FROM workshop_layouts WHERE team_id=? ORDER BY updated_at DESC", (team_id,))
        else:
            c.execute("SELECT * FROM workshop_layouts ORDER BY updated_at DESC")
        rows = c.fetchall()
        layouts = []
        for r in rows:
            layouts.append({
                'id': r[0], 'team_id': r[1], 'name': r[2],
                'layout_data': r[3], 'created_at': r[4], 'updated_at': r[5]
            })
        conn.close()
        return jsonify({'success': True, 'data': layouts})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workshop-layouts', methods=['POST'])
@login_required
def api_save_workshop_layout():
    try:
        data = request.json
        name = data.get('name', '默认布局').strip()
        if not name:
            return jsonify({'success': False, 'error': '布局名称不能为空'}), 400
        conn = get_connection()
        c = conn.cursor()
        # Check duplicate name
        if data.get('id'):
            c.execute("SELECT id FROM workshop_layouts WHERE name=? AND id!=?", (name, data['id']))
        else:
            c.execute("SELECT id FROM workshop_layouts WHERE name=?", (name,))
        if c.fetchone():
            conn.close()
            return jsonify({'success': False, 'error': '布局名称已存在，请使用其他名称'}), 400
        if data.get('id'):
            c.execute("UPDATE workshop_layouts SET name=?, layout_data=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                      (name, data.get('layout_data', '{}'), data['id']))
        else:
            c.execute("INSERT INTO workshop_layouts (team_id, name, layout_data) VALUES (?,?,?)",
                      (data.get('team_id'), name, data.get('layout_data', '{}')))
        conn.commit()
        layout_id = data.get('id') or c.lastrowid
        conn.close()
        return jsonify({'success': True, 'id': layout_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/workshop-layouts/<int:lid>', methods=['DELETE'])
@login_required
def api_delete_workshop_layout(lid):
    try:
        conn = get_connection()
        c = conn.cursor()
        c.execute("SELECT id FROM workshop_layouts WHERE id=?", (lid,))
        if not c.fetchone():
            conn.close()
            return jsonify({'success': False, 'error': '布局不存在'}), 404
        c.execute("DELETE FROM workshop_layouts WHERE id=?", (lid,))
        conn.commit()
        deleted = c.rowcount
        conn.close()
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



# ========== Report Export APIs ==========
@app.route('/api/reports/daily/export')
@login_required
def api_report_daily_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    import io
    
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    team_id = request.args.get('team_id', type=int)
    
    personnel_map = _get_personnel_map()
    report_data = _expand_reports_with_attendance_note(date, date)
    
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT p.name, a.work_hours, a.is_overtime, a.leave_type, a.normal_hours, a.overtime_hours FROM attendance a INNER JOIN personnel p ON a.user_id = p.user_id WHERE a.work_date=?", (date,))
    attend_data = {r[0]: {'hours': r[1] or 0, 'overtime': r[2], 'leave': r[3] or '', 'normal': r[4] or 0, 'ot': r[5] or 0} for r in c.fetchall()}
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = u'\u65e5\u62a5\u8868'
    
    headers = [u'\u59d3\u540d', u'\u73ed\u7ec4', u'\u6253\u5361\u5de5\u65f6', u'\u5e38\u89c4\u5de5\u65f6', u'\u52a0\u73ed\u5de5\u65f6', u'\u751f\u4ea7\u5de5\u65f6', u'\u603b\u4ea7\u91cf', u'\u826f\u54c1\u7387', u'\u751f\u4ea7\u6548\u7387', u'\u5de5\u65f6\u5229\u7528\u7387', u'\u52a0\u73ed']
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True, size=11)
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center'); cell.border = bdr
    
    row_num = 2
    for name, info in personnel_map.items():
        if team_id and info['team_id'] != team_id: continue
        rd = report_data.get(name, {'qty': 0, 'hours': 0, 'good_qty': 0, 'efficiency': 0})
        ad = attend_data.get(name, {'hours': 0, 'overtime': 0, 'leave': '', 'normal': 0, 'ot': 0})
        if name not in report_data and name not in attend_data: continue
        good_rate = round(rd['good_qty'] / rd['qty'] * 100, 1) if rd['qty'] > 0 else 0
        att_hrs = ad['hours']; prod_hrs = round(rd['hours'], 1)
        util_rate = round(prod_hrs / att_hrs * 100, 1) if att_hrs > 0 else 0
        vals = [name, info['department'], att_hrs, ad['normal'], ad['ot'], prod_hrs, rd['qty'],
                str(good_rate)+'%' if good_rate > 0 else '-',
                str(rd['efficiency'])+'%' if rd['efficiency'] > 0 else '-',
                str(util_rate)+'%' if util_rate > 0 else '-',
                u'\u52a0\u73ed' if ad['overtime'] else '-']
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = bdr; cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    for col in range(1, len(headers)+1): ws.column_dimensions[chr(64+col)].width = 14
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=u'\u65e5\u62a5\u8868_'+date+'.xlsx')


@app.route('/api/reports/personal/export')
@login_required
def api_report_personal_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    import io
    from datetime import timedelta
    
    name = request.args.get('name', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    if not name: return jsonify({'error': '请输入员工姓名'}), 400
    
    report_by_date = _expand_personal_report_by_date(name, date_from, date_to)

    conn = get_connection(); c = conn.cursor()
    c.execute("SELECT a.work_date, a.work_hours, a.is_overtime, a.leave_type, a.normal_hours, a.overtime_hours FROM attendance a INNER JOIN personnel p ON a.user_id=p.user_id WHERE p.name=? AND a.work_date BETWEEN ? AND ? ORDER BY a.work_date",
              (name, date_from, date_to))
    attend_by_date = {r[0]: {'hours': r[1] or 0, 'overtime': r[2], 'leave': r[3] or '', 'normal': r[4] or 0, 'ot': r[5] or 0} for r in c.fetchall()}
    conn.close()
    
    days = []
    d = datetime.strptime(date_from, '%Y-%m-%d'); end = datetime.strptime(date_to, '%Y-%m-%d')
    while d <= end:
        ds = d.strftime('%Y-%m-%d')
        rd = report_by_date.get(ds, {'qty': 0, 'hours': 0, 'good_qty': 0, 'efficiency': 0})
        ad = attend_by_date.get(ds, {'hours': 0, 'overtime': 0, 'leave': '', 'normal': 0, 'ot': 0})
        att_hrs = ad['hours']; prod_hrs = rd['hours']
        good_rate = round(rd['good_qty']/rd['qty']*100, 1) if rd['qty'] > 0 else 0
        util_rate = round(prod_hrs/att_hrs*100, 1) if att_hrs > 0 else 0
        status = ad['leave'] if ad['leave'] else ('加班' if ad['overtime'] else ('正常' if att_hrs > 0 else ('休息' if ds in attend_by_date else '-')))
        days.append({'date': ds, 'attendance_hours': att_hrs, 'normal_hours': ad['normal'], 'overtime_hours': ad['ot'],
                     'hours': prod_hrs, 'qty': rd['qty'], 'good_rate': good_rate, 'efficiency': rd['efficiency'],
                     'utilization_rate': util_rate, 'status': status})
        d += timedelta(days=1)
    
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = name + '效率查询'
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True, size=11)
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['指标'] + [day['date'][5:] for day in days]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center'); cell.border = bdr
    
    fields = [('attendance_hours','打卡工时'),('normal_hours','常规工时'),('overtime_hours','加班工时'),
              ('hours','生产工时'),('qty','总产量'),('good_rate','良品率'),('efficiency','生产效率'),
              ('utilization_rate','工时利用率'),('status','出勤状态')]
    for row_idx, (key, label) in enumerate(fields, 2):
        ws.cell(row=row_idx, column=1, value=label).border = bdr
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        for col_idx, day in enumerate(days, 2):
            val = day[key]
            if key in ('good_rate','efficiency','utilization_rate'): val = str(val)+'%' if val > 0 else '-'
            elif key in ('attendance_hours','normal_hours','overtime_hours','hours'): val = str(val)+'h' if val > 0 else '-'
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = bdr; cell.alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 14
    for col in range(2, len(days)+2):
        c_letter = chr(64+col) if col <= 26 else chr(64+col//26)+chr(64+col%26)
        ws.column_dimensions[c_letter].width = 12
    
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=name+'效率查询_'+date_from+'_'+date_to+'.xlsx')

@app.route('/api/reports/weekly/export')
@login_required
def api_report_weekly_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    import io
    from datetime import timedelta
    
    week_start = request.args.get('week_start', '')
    team_id = request.args.get('team_id', type=int)
    if not week_start: return jsonify({'error': u'\u8bf7\u9009\u62e9\u5468'}), 400
    
    start = datetime.strptime(week_start, '%Y-%m-%d')
    dates = [(start+timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    ph = ','.join(['?' for _ in dates])
    
    personnel_map = _get_personnel_map()
    week_end = (start + timedelta(days=6)).strftime('%Y-%m-%d')
    expanded = _expand_reports_with_attendance_note(week_start, week_end)
    report_data = {}
    for nm, rd in expanded.items():
        report_data[nm] = {'qty': rd['qty'], 'hours': rd['hours'], 'good_qty': rd['good_qty'], 'efficiency': rd['efficiency'], 'days': rd['report_count']}
    
    conn = get_connection(); c = conn.cursor()
    c.execute("SELECT p.name, SUM(a.work_hours), SUM(a.normal_hours), SUM(a.overtime_hours), COUNT(*) FROM attendance a INNER JOIN personnel p ON a.user_id=p.user_id WHERE a.work_date IN ("+ph+") GROUP BY p.name", dates)
    attend_data = {}
    for r in c.fetchall(): attend_data[r[0]] = {'hours': r[1] or 0, 'normal': r[2] or 0, 'ot': r[3] or 0, 'days': r[4] or 0}
    conn.close()
    
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = u'\u5468\u62a5\u8868'
    headers = [u'\u59d3\u540d',u'\u73ed\u7ec4',u'\u51fa\u52e4\u5929\u6570',u'\u6253\u5361\u5de5\u65f6',u'\u5e38\u89c4\u5de5\u65f6',u'\u52a0\u73ed\u5de5\u65f6',u'\u751f\u4ea7\u5de5\u65f6',u'\u603b\u4ea7\u91cf',u'\u826f\u54c1\u7387',u'\u5e73\u5747\u6548\u7387',u'\u52a0\u73ed\u5929\u6570',u'\u5de5\u65f6\u5229\u7528\u7387',u'\u65e5\u5747\u4ea7\u91cf']
    hfill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True, size=11)
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center'); cell.border = bdr
    
    row_num = 2
    for name, info in personnel_map.items():
        if team_id and info['team_id'] != team_id: continue
        rd = report_data.get(name, {'qty': 0, 'hours': 0, 'good_qty': 0, 'efficiency': 0, 'days': 0})
        ad = attend_data.get(name, {'hours': 0, 'normal': 0, 'ot': 0, 'days': 0})
        if rd['qty'] == 0 and ad['hours'] == 0: continue
        good_rate = round(rd['good_qty']/rd['qty']*100, 1) if rd['qty'] > 0 else 0
        att_hrs = ad['hours']; prod_hrs = round(rd['hours'], 1)
        util_rate = round(prod_hrs/att_hrs*100, 1) if att_hrs > 0 else 0
        daily_avg = round(rd['qty']/rd['days'], 1) if rd['days'] > 0 else 0
        vals = [name, info['department'], ad['days'], att_hrs, ad['normal'], ad['ot'], prod_hrs, rd['qty'],
                str(good_rate)+'%' if good_rate > 0 else '-',
                str(rd['efficiency'])+'%' if rd['efficiency'] > 0 else '-',
                rd['days'], str(util_rate)+'%' if util_rate > 0 else '-', daily_avg]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = bdr; cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    for col in range(1, len(headers)+1): ws.column_dimensions[chr(64+col)].width = 14
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=u'\u5468\u62a5\u8868_'+week_start+'.xlsx')


@app.route('/api/reports/monthly/export')
@login_required
def api_report_monthly_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from flask import send_file
    import io
    import calendar as cal
    
    month = request.args.get('month', '')
    team_id = request.args.get('team_id', type=int)
    if not month: return jsonify({'error': u'\u8bf7\u9009\u62e9\u6708\u4efd'}), 400
    
    date_from = month + '-01'
    y, m = int(month.split('-')[0]), int(month.split('-')[1])
    last_day = cal.monthrange(y, m)[1]
    date_to = month + '-' + str(last_day).zfill(2)
    
    conn = get_connection(); c = conn.cursor()
    personnel_map = _get_personnel_map()
    personnel_map = _get_personnel_map()
    expanded = _expand_reports_with_attendance_note(date_from, date_to)
    report_data = {}
    for nm, rd in expanded.items():
        report_data[nm] = {'qty': rd['qty'], 'hours': rd['hours'], 'good_qty': rd['good_qty'], 'efficiency': rd['efficiency'], 'days': rd['report_count']}
    
    conn = get_connection(); c = conn.cursor()
    c.execute("SELECT p.name, SUM(a.work_hours), SUM(a.normal_hours), SUM(a.overtime_hours), COUNT(*), SUM(CASE WHEN a.leave_type IS NOT NULL AND a.leave_type != '' THEN 1 ELSE 0 END) FROM attendance a INNER JOIN personnel p ON a.user_id=p.user_id WHERE a.work_date BETWEEN ? AND ? GROUP BY p.name", (date_from, date_to))
    attend_data = {}
    for r in c.fetchall(): attend_data[r[0]] = {'hours': r[1] or 0, 'normal': r[2] or 0, 'ot': r[3] or 0, 'days': r[4] or 0, 'leave': r[5] or 0}
    conn.close()
    hfont = Font(color='FFFFFF', bold=True, size=11)
    bdr = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center'); cell.border = bdr
    
    row_num = 2
    for name, info in personnel_map.items():
        if team_id and info['team_id'] != team_id: continue
        rd = report_data.get(name, {'qty': 0, 'hours': 0, 'good_qty': 0, 'efficiency': 0, 'days': 0})
        ad = attend_data.get(name, {'hours': 0, 'normal': 0, 'ot': 0, 'days': 0, 'leave': 0})
        if rd['qty'] == 0 and ad['hours'] == 0: continue
        good_rate = round(rd['good_qty']/rd['qty']*100, 1) if rd['qty'] > 0 else 0
        att_hrs = ad['hours']; prod_hrs = round(rd['hours'], 1)
        util_rate = round(prod_hrs/att_hrs*100, 1) if att_hrs > 0 else 0
        daily_avg = round(rd['qty']/rd['days'], 1) if rd['days'] > 0 else 0
        vals = [name, info['department'], ad['days'], att_hrs, ad['normal'], ad['ot'], prod_hrs, rd['qty'],
                str(good_rate)+'%' if good_rate > 0 else '-',
                str(rd['efficiency'])+'%' if rd['efficiency'] > 0 else '-',
                rd['days'], str(util_rate)+'%' if util_rate > 0 else '-', daily_avg, '0%', ad['leave']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = bdr; cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    for col in range(1, len(headers)+1): ws.column_dimensions[chr(64+col)].width = 14
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=u'\u6708\u62a5\u8868_'+month+'.xlsx')






@app.route('/material-alerts-page')
@login_required
def material_alerts_page():
    return render_template('material_alerts.html')

# ========== Material Alerts API ==========

def _check_material_alerts():
    """Scan for material alerts using BOM product hierarchy.

    Algorithm:
    1. Auto-close alerts for parent orders that already have progress > 0%.
    2. For each in_progress order, find the last substantive process (excluding 清洗/入库).
    3. Check if that last process has work reports (has started being reported).
    4. Use BOM to find parent product: bom.child_product_code = order's product_code -> bom.parent_product_code
    5. Find work orders with that parent product_code (status in_progress or pending).
       That work order IS the parent - no task-number filtering needed.
    6. Insert alert.
    """
    import re as _re

    conn = get_connection()
    c = conn.cursor()
    skip_kw = ['清洗', '入库']

    # Step 1: Auto-close alerts for parent orders that already started production
    c.execute("SELECT order_no, process_progress FROM work_orders "
              "WHERE process_progress IS NOT NULL AND process_progress != '' "
              "AND status IN ('in_progress','pending')")
    parent_started = set()
    for order_no, prog in c.fetchall():
        for step in prog.split('->'):
            step = step.strip()
            if not step:
                continue
            m = _re.match(r'^(.+?)[【（]', step)
            name = m.group(1).strip() if m else step
            if any(kw in name for kw in skip_kw):
                continue
            m2 = _re.search(r'【(\d+\.?\d*)/(\d+\.?\d*)】', step)
            if m2 and float(m2.group(1)) > 0:
                parent_started.add(order_no)
            break
    auto_closed = 0
    if parent_started:
        ph = ','.join(['?' for _ in parent_started])
        c.execute(
            "UPDATE material_alerts SET status='auto_closed', closed_at=datetime('now','localtime'), "
            "closed_by='system' WHERE status='pending' AND parent_order_no IN (" + ph + ")",
            list(parent_started)
        )
        auto_closed = c.rowcount

    # Step 2: Build work_reports index
    c.execute("SELECT order_no, process_name, COUNT(*), MAX(create_time), MAX(good_qty) "
              "FROM work_reports GROUP BY order_no, process_name")
    report_map = {}
    for r in c.fetchall():
        report_map[(r[0], r[1])] = {"count": r[2], "latest_time": r[3], "qty": r[4]}

    # Step 3: Find child trigger candidates
    c.execute("SELECT order_no, product_code, product_name, process_progress, status "
              "FROM work_orders WHERE process_progress IS NOT NULL AND process_progress != '' "
              "AND status = 'in_progress'")
    order_rows = c.fetchall()

    candidates = []
    for order_no, child_pc, child_pn, prog, st in order_rows:
        # Parse processes from progress string, exclude 清洗/入库
        last_sub = None
        for step in prog.split('->'):
            step = step.strip()
            if not step:
                continue
            m = _re.match(r'^(.+?)[【（]', step)
            name = m.group(1).strip() if m else step
            if any(kw in name for kw in skip_kw):
                continue
            last_sub = name
        if not last_sub:
            continue
        # Check if last substantive process has reports
        matched_info = None
        for rp_key, rp_info in report_map.items():
            if rp_key[0] == order_no and last_sub in rp_key[1]:
                matched_info = rp_info
                break
        if not matched_info or matched_info["count"] <= 0:
            continue
        candidates.append({
            'order_no': order_no,
            'product_code': child_pc,
            'product_name': child_pn,
            'trigger_process': last_sub,
            'trigger_qty': matched_info["qty"] or 0,
            'trigger_time': matched_info["latest_time"] or ''
        })

    if not candidates:
        conn.commit()
        conn.close()
        return auto_closed

    # Step 4: Build BOM child_product -> parent_product map
    c.execute("SELECT DISTINCT parent_product_code, child_product_code FROM bom")
    bom_child_to_parent = {}
    for parent_pc, child_pc in c.fetchall():
        if child_pc not in bom_child_to_parent:
            bom_child_to_parent[child_pc] = []
        bom_child_to_parent[child_pc].append(parent_pc)

    # Step 5: Build product_code -> work_orders index
    c.execute("SELECT order_no, product_code, product_name FROM work_orders "
              "WHERE status IN ('in_progress','pending')")
    pc_to_orders = {}
    for o_no, o_pc, o_pn in c.fetchall():
        if o_pc not in pc_to_orders:
            pc_to_orders[o_pc] = []
        pc_to_orders[o_pc].append((o_no, o_pn))

    # Step 6: Build order prefix -> order list index
    # Order prefix is the part before the first '-' (e.g. MO000489 from MO000489-005)
    def _order_prefix(on):
        idx = on.find('-')
        return on[:idx] if idx > 0 else on

    c.execute("SELECT order_no, product_code, product_name FROM work_orders "
              "WHERE status IN ('in_progress','pending')")
    all_active_orders = {}
    for o_no, o_pc, o_pn in c.fetchall():
        all_active_orders[o_no] = (o_pc, o_pn)

    # Step 7: Insert alerts - BOM product hierarchy + same order family
    c.execute("SELECT child_order_no, parent_order_no, trigger_process FROM material_alerts "
              "WHERE status NOT IN ('auto_closed')")
    existing = set((r[0], r[1], r[2]) for r in c.fetchall())

    new_count = 0
    for cand in candidates:
        child_order = cand['order_no']
        child_pc = cand['product_code']
        trigger_proc = cand['trigger_process']
        child_pfx = _order_prefix(child_order)

        # Find parent product codes from BOM
        parent_products = bom_child_to_parent.get(child_pc, [])
        if not parent_products:
            continue

        # Find parent orders: must share the same order prefix family
        for p_order_no, (p_pc, p_pn) in all_active_orders.items():
            if p_order_no == child_order:
                continue
            if p_order_no in parent_started:
                continue
            # Must be same order family (same prefix before '-')
            if _order_prefix(p_order_no) != child_pfx:
                continue
            # Must match a BOM parent product
            if p_pc not in parent_products:
                continue
            if (child_order, p_order_no, trigger_proc) in existing:
                continue
            try:
                c.execute(
                    """INSERT OR IGNORE INTO material_alerts
                        (child_order_no, child_product_code, child_product_name,
                         parent_order_no, parent_product_code, parent_product_name,
                         trigger_process, bom_qty, trigger_qty, trigger_time, status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending')""",
                    (
                        child_order, child_pc, cand['product_name'],
                        p_order_no, p_pc, p_pn,
                        trigger_proc, 0, cand['trigger_qty'], cand['trigger_time']
                    )
                )
                if c.rowcount > 0:
                    new_count += 1
                    existing.add((child_order, p_order_no, trigger_proc))
            except Exception:
                pass

    conn.commit()
    conn.close()
    return new_count + auto_closed
@app.route('/api/material-alerts/refresh', methods=['POST'])
@login_required
@planner_required
def api_material_alerts_refresh():
    """Trigger material alert scan."""
    new_count = _check_material_alerts()
    return jsonify({'ok': True, 'new_alerts': new_count})

@app.route('/api/material-alerts')
@login_required
def api_material_alerts():
    """Get material alerts with optional filters: status, team, date_from, date_to."""
    import re as _alert_re

    def _task_base(order_no):
        if not order_no:
            return ''
        s = str(order_no)
        m = _alert_re.match(r'^([A-Za-z]+0*)(\d+)(?:-\d+)?$', s)
        return m.group(1) + m.group(2) if m else s.split('-')[0]

    status = request.args.get('status', '')
    team = request.args.get('team', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    conn = get_connection()
    c = conn.cursor()

    # Build query with filters
    conditions = []
    params = []
    if status == 'closed':
        conditions.append("ma.status='closed'")
    elif status == 'all':
        pass  # no filter
    else:
        conditions.append("ma.status='pending'")
        conditions.append("ma.child_order_no NOT IN (SELECT order_no FROM work_orders WHERE status='completed')")
        conditions.append("ma.parent_order_no NOT IN (SELECT order_no FROM work_orders WHERE status='completed')")

    if date_from:
        conditions.append("ma.created_at >= ?")
        params.append(date_from + ' 00:00:00')
    if date_to:
        conditions.append("ma.created_at <= ?")
        params.append(date_to + ' 23:59:59')

    where = (' WHERE ' + ' AND '.join(conditions)) if conditions else ''

    c.execute("""SELECT ma.id, ma.child_order_no, ma.child_product_code, ma.child_product_name,
        ma.parent_order_no, ma.parent_product_code, ma.parent_product_name,
        ma.trigger_process, ma.bom_qty, ma.trigger_qty, ma.trigger_time, ma.status, ma.closed_by, ma.closed_at, ma.created_at
        FROM material_alerts ma""" + where + " ORDER BY ma.created_at DESC", params)

    alerts = []
    for r in c.fetchall():
        alerts.append({
            'id': r[0], 'child_order_no': r[1], 'child_product_code': r[2],
            'child_product_name': r[3], 'parent_order_no': r[4],
            'parent_product_code': r[5], 'parent_product_name': r[6],
            'trigger_process': r[7], 'bom_qty': r[8], 'trigger_qty': r[9], 'trigger_time': r[10], 'status': r[11],
            'closed_by': r[12], 'closed_at': r[13], 'created_at': r[14],
            'team_name': '', 'parent_progress': '', 'child_progress': ''
        })

    # Build process->team map
    process_team_map = {}
    c.execute("SELECT process_name, team_name FROM processes WHERE team_name IS NOT NULL AND team_name != ''")
    for r in c.fetchall():
        process_team_map[r[0]] = r[1]

    # Resolve team from PARENT order's first substantive process (exclude 清洗/入库)
    skip_kw = ['清洗', '入库']
    parent_team_cache = {}
    for alert in alerts:
        p_order = alert['parent_order_no']
        if p_order not in parent_team_cache:
            team_name = ''
            p_code = alert['parent_product_code']
            c.execute("SELECT process_list FROM process_routes WHERE product_code=? LIMIT 1", (p_code,))
            pr = c.fetchone()
            if pr and pr[0]:
                steps = [s.strip() for s in pr[0].split(',') if s.strip()]
                for step in steps:
                    skip = False
                    for kw in skip_kw:
                        if kw in step:
                            skip = True
                            break
                    if not skip:
                        team_name = process_team_map.get(step, '')
                        if not team_name:
                            for pn, tn in process_team_map.items():
                                if pn in step or step in pn:
                                    team_name = tn
                                    break
                        if not team_name:
                            # try stripping parenthetical suffix e.g. 焊接开关座（手动）-> 焊接开关座
                            base = re.sub(r'[（(].*?[）)]', '', step).strip()
                            if base:
                                team_name = process_team_map.get(base, '')
                                if not team_name:
                                    for pn, tn in process_team_map.items():
                                        if pn in base or base in pn:
                                            team_name = tn
                                            break
                        if team_name:
                            break
            parent_team_cache[p_order] = team_name
        alert['team_name'] = parent_team_cache[p_order]

    # Add parent order progress
    parent_progress_cache = {}
    for alert in alerts:
        p_order = alert['parent_order_no']
        if p_order not in parent_progress_cache:
            c.execute("SELECT process_progress FROM work_orders WHERE order_no=?", (p_order,))
            row = c.fetchone()
            parent_progress_cache[p_order] = row[0] if row and row[0] else ''
        alert['parent_progress'] = parent_progress_cache[p_order]

    # Add child order progress
    child_progress_cache = {}
    for alert in alerts:
        c_order = alert['child_order_no']
        if c_order not in child_progress_cache:
            c.execute('SELECT process_progress FROM work_orders WHERE order_no=?',(c_order,))
            row = c.fetchone()
            child_progress_cache[c_order] = row[0] if row and row[0] else ''
        alert['child_progress'] = child_progress_cache[c_order]

    # Filter by team if specified
    if team:
        alerts = [a for a in alerts if a['team_name'] == team]

    # Get BOM details for each parent order
    bom_details = {}
    for alert in alerts:
        parent_order = alert['parent_order_no']
        if parent_order not in bom_details:
            parent_code = alert['parent_product_code']
            c.execute("""SELECT child_product_code, child_product_name, quantity, process_team
                FROM bom WHERE parent_product_code=?""", (parent_code,))
            children = []
            for br in c.fetchall():
                c2 = conn.cursor()
                c2.execute("SELECT order_no, status, process_progress FROM work_orders WHERE product_code=? AND status IN ('in_progress','pending') LIMIT 1", (br[0],))
                wo = c2.fetchone()
                children.append({
                    'product_code': br[0], 'product_name': br[1],
                    'quantity': br[2], 'process_team': br[3],
                    'work_order_no': wo[0] if wo else None,
                    'work_order_status': wo[1] if wo else None,
                    'work_order_progress': wo[2] if wo else None,
                })
            bom_details[parent_order] = children
        alert['bom_children'] = bom_details.get(parent_order, [])

    conn.close()
    return jsonify({'data': alerts})
@app.route('/api/material-alerts/<int:alert_id>/close', methods=['POST'])
@login_required
def api_material_alert_close(alert_id):
    """Mark all alerts for the same parent order as closed."""
    user = session.get('user', {})
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT parent_order_no FROM material_alerts WHERE id=?", (alert_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'not found'}), 404
    parent_order_no = row[0]
    c.execute("UPDATE material_alerts SET status='closed', closed_by=?, closed_at=datetime('now','localtime') WHERE parent_order_no=? AND status='pending'",
              (user.get('display_name', ''), parent_order_no))
    closed = c.rowcount
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'closed': closed})

@app.route('/api/material-alerts/close-batch', methods=['POST'])
@login_required
def api_material_alerts_close_batch():
    """Close all alerts for the parent orders of selected alerts."""
    data = request.json
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400
    user = session.get('user', {})
    conn = get_connection()
    c = conn.cursor()
    placeholders = ','.join(['?' for _ in ids])
    c.execute('SELECT DISTINCT parent_order_no FROM material_alerts WHERE id IN (' + placeholders + ')', ids)
    parent_orders = [r[0] for r in c.fetchall() if r[0]]
    closed = 0
    if parent_orders:
        ph2 = ','.join(['?' for _ in parent_orders])
        c.execute('UPDATE material_alerts SET status=\'closed\', closed_by=?, closed_at=datetime(\'now\',\'localtime\') WHERE parent_order_no IN (' + ph2 + ') AND status=\'pending\'',
                  [user.get('display_name', '')] + parent_orders)
        closed = c.rowcount
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'closed': closed})
if __name__ == "__main__":
    init_database()
    # Initialize equipment status on server start (reset all to normal, then refresh from today's reports)
    try:
        _refresh_equipment_status()
        print("[equipment-status] Initial refresh on server start")
    except Exception as e:
        print(f"[equipment-status] Initial refresh error: {e}")
    app.run(debug=False, host="0.0.0.0", port=6002)





















