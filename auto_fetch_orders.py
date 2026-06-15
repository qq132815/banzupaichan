# -*- coding: utf-8 -*-
"""Auto-import work orders from MES /production/task page.
Aggregates task-level rows into order-level records with process_progress.
"""
import os, sys, time, sqlite3, re
from collections import defaultdict
from datetime import datetime

MES_URL = "https://web.ycmes.cn/"
FACTORY_CODE = "606999"
USERNAME = "gs8888"
PASSWORD = "256448"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
DB_PATH = os.path.join(BASE_DIR, "data", "production.db")


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def parse_task_export(filepath):
    """Parse MES task export Excel and aggregate into order-level records."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Row 1 = merged title, Row 2 = headers, Row 3+ = data
    headers = [str(c.value or "").strip() for c in ws[2]]

    # Build column index map
    col_map = {}
    field_aliases = {
        "task_no": ["任务编号", "工单编号"],
        "product_code": ["产品编号", "产品编码"],
        "product_type": ["产品类型"],
        "process_name": ["工序名称"],
        "task_progress": ["任务进度（%）", "任务进度"],
        "report_permission": ["报工权限"],
        "plan_qty": ["计划数"],
        "good_qty": ["良品数"],
        "bad_qty": ["不良品数"],
        "task_status": ["任务状态"],
        "priority": ["优先级"],
        "plan_start": ["计划开始时间"],
        "plan_end": ["计划结束时间"],
        "overdue": ["逾期状态"],
        "actual_start": ["实际开始时间"],
        "actual_end": ["实际结束时间"],
        "order_status": ["工单状态"],
        "related_doc": ["关联单据号"],
        "creator": ["创建人"],
        "create_time": ["创建时间"],
        "mfg_type": ["加工类型"],
        "order_plan_qty": ["工单计划数"],
    }

    for field, aliases in field_aliases.items():
        for alias in aliases:
            for i, h in enumerate(headers):
                if alias in h or h in alias:
                    col_map[field] = i
                    break
            if field in col_map:
                break

    if "task_no" not in col_map:
        print("  ERROR: Cannot find task_no column. Headers: %s" % headers[:10])
        wb.close()
        return []

    def get_val(row, field, default=""):
        idx = col_map.get(field, -1)
        if 0 <= idx < len(row) and row[idx] is not None:
            return str(row[idx]).strip()
        return default

    def get_num(row, field, default=0):
        idx = col_map.get(field, -1)
        if 0 <= idx < len(row) and row[idx] is not None:
            try:
                return float(row[idx])
            except:
                pass
        return default

    # Parse all tasks and group by parent order
    # Parent order = task_no without the last "-NNN" suffix
    tasks_by_order = defaultdict(list)
    order_info = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue

        task_no = get_val(row, "task_no")
        if not task_no or len(task_no) < 3:
            continue

        # Extract parent order no (e.g., "APO20260231-001" -> "APO20260231")
        parent_match = re.match(r'^(.+?)-\d+$', task_no)
        parent_order = parent_match.group(1) if parent_match else task_no

        process_name = get_val(row, "process_name")
        task_progress = get_val(row, "task_progress")
        plan_qty = get_num(row, "plan_qty")
        good_qty = get_num(row, "good_qty")
        priority = get_val(row, "priority", "P2")
        plan_end = get_val(row, "plan_end")
        order_status = get_val(row, "order_status", "执行中")
        product_code = get_val(row, "product_code")
        related_doc = get_val(row, "related_doc")

        # Parse progress percentage
        pct = 0
        pct_match = re.search(r'(\d+)', task_progress)
        if pct_match:
            pct = int(pct_match.group(1))

        task_info = {
            "task_no": task_no,
            "process_name": process_name,
            "progress_pct": pct,
            "plan_qty": plan_qty,
            "good_qty": good_qty,
        }
        tasks_by_order[parent_order].append(task_info)

        # Store order-level info (from first task seen)
        if parent_order not in order_info:
            order_info[parent_order] = {
                "product_code": product_code,
                "priority": priority,
                "plan_end": plan_end[:10] if plan_end and len(plan_end) >= 10 else "",
                "order_status": order_status,
                "related_doc": related_doc,
            }
        else:
            # Update with non-empty values
            info = order_info[parent_order]
            if not info["plan_end"] and plan_end:
                info["plan_end"] = plan_end[:10]
            if not info["product_code"] and product_code:
                info["product_code"] = product_code

    wb.close()

    # Build order-level records with process_progress
    records = []
    for parent_order, tasks in tasks_by_order.items():
        info = order_info.get(parent_order, {})

        # Get total quantity from the order (use max plan_qty seen, or first task's plan_qty)
        # Actually, for task-level data, each task has its own plan_qty
        # We use the first task's plan_qty as the order quantity
        total_qty = tasks[0]["plan_qty"] if tasks else 0

        # Build process_progress string
        # Format: process1【done/total】(pct%)->process2【done/total】(pct%)
        progress_parts = []
        total_good = 0
        for t in tasks:
            p_name = t["process_name"]
            p_pct = t["progress_pct"]
            p_qty = t["plan_qty"]
            p_good = t["good_qty"]

            # Calculate done/total for this process
            done = int(p_good) if p_good > 0 else int(p_qty * p_pct / 100) if p_pct > 0 else 0
            total = int(p_qty) if p_qty > 0 else int(total_qty)

            # Check mark for completed
            check = "\u2705" if p_pct >= 100 else ""
            progress_parts.append("%s%s[%d/%d](%d%%)" % (p_name, check, done, total, p_pct))
            total_good = max(total_good, p_good)

        process_progress = "->".join(progress_parts)

        # Determine status
        status = info.get("order_status", "执行中")
        if status == "已结束":
            status = "completed"
        elif status == "执行中":
            status = "in_progress"
        else:
            status = "pending"

        records.append({
            "order_no": parent_order,
            "product_code": info.get("product_code", ""),
            "product_name": info.get("product_code", ""),
            "quantity": total_qty,
            "completed_qty": total_good,
            "due_date": info.get("plan_end", ""),
            "priority": info.get("priority", "P2"),
            "status": status,
            "process_progress": process_progress,
            "source": "mes",
            "route_code": info.get("related_doc", ""),
        })

    return records


def save_to_db(records):
    """Save aggregated order records to database."""
    if not records:
        return 0, 0

    conn = get_connection()
    c = conn.cursor()

    # Ensure columns exist
    for col in ["process_progress", "source", "route_code"]:
        try:
            c.execute("ALTER TABLE work_orders ADD COLUMN %s TEXT" % col)
        except:
            pass

    inserted = updated = 0
    for r in records:
        c.execute("SELECT id FROM work_orders WHERE order_no=?", (r["order_no"],))
        existing = c.fetchone()

        if existing:
            c.execute("""UPDATE work_orders SET product_code=?, product_name=?, quantity=?,
                completed_qty=?, due_date=?, priority=?, status=?, process_progress=?, source=?, route_code=? WHERE order_no=?""",
                (r["product_code"], r["product_name"], r["quantity"], r["completed_qty"],
                 r["due_date"], r["priority"], r["status"], r["process_progress"],
                 r["source"], r["route_code"], r["order_no"]))
            updated += 1
        else:
            c.execute("""INSERT INTO work_orders
                (order_no, product_code, product_name, quantity, completed_qty, due_date, priority, status, process_progress, source, route_code)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (r["order_no"], r["product_code"], r["product_name"], r["quantity"], r["completed_qty"],
                 r["due_date"], r["priority"], r["status"], r["process_progress"],
                 r["source"], r["route_code"]))
            inserted += 1

    conn.commit()
    conn.close()
    return inserted, updated


def run_fetch_orders():
    """Login to MES, export tasks, aggregate and import as orders."""
    print("[%s] Fetching work orders from MES task page..." % datetime.now().strftime("%H:%M:%S"))

    from playwright.sync_api import sync_playwright

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=['--disable-blink-features=AutomationControlled', '--no-sandbox'])
        ctx = browser.new_context(accept_downloads=True, user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        page = ctx.new_page()
        page.add_init_script('Object.defineProperty(navigator, "webdriver", {get: () => undefined})')

        try:
            # Login
            for attempt in range(3):
                try:
                    page.goto(MES_URL, wait_until="networkidle", timeout=60000)
                    break
                except:
                    if attempt == 2:
                        raise
                    time.sleep(10)
            time.sleep(2)

            page.locator(u"text=\u8d26\u53f7\u767b\u5f55").first.click()
            time.sleep(1)
            page.locator(u"input[placeholder*=\u5de5\u5382]").first.fill(FACTORY_CODE)
            page.locator(u"input[placeholder*=\u7528\u6237\u540d]").first.fill(USERNAME)
            page.locator("input[type=password]").first.fill(PASSWORD)
            page.locator("input[type=password]").first.press("Enter")
            try:
                page.wait_for_url(lambda url: "login" not in url, timeout=15000)
            except Exception:
                pass
            time.sleep(3)
            print("  Login OK")

            # Navigate to task page (correct URL for work orders)
            page.goto(MES_URL.rstrip("/") + "/production/task", wait_until="networkidle", timeout=60000)
            time.sleep(3)
            print("  Task page loaded: %s" % page.url)

            # Click "执行中" filter first to get only active orders
            try:
                exec_btn = page.locator(u"text=\u6267\u884c\u4e2d").first
                exec_btn.click()
                time.sleep(3)
                print("  Filtered: 执行中")
            except:
                print("  Could not filter 执行中, using all")

            # Click export button and capture download
            print("  Exporting...")
            export_btn = None
            btns = page.locator("button").all()
            for btn in btns:
                try:
                    txt = (btn.text_content() or "").strip()
                    if u"\u5bfc\u51fa" in txt and u"\u6a21\u677f" not in txt:
                        export_btn = btn
                        break
                except:
                    pass

            if not export_btn:
                print("  ERROR: Export button not found")
                return

            with page.expect_download(timeout=300000) as dl_info:
                export_btn.click()
                print("  Waiting for download...")

            download = dl_info.value
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(DOWNLOAD_DIR, "mes_tasks_%s.xlsx" % ts)
            download.save_as(filepath)
            print("  Downloaded: %s" % filepath)

            # Parse and aggregate
            print("  Parsing task data...")
            records = parse_task_export(filepath)
            print("  Aggregated: %d orders from task data" % len(records))

            # Save to database
            inserted, updated = save_to_db(records)
            print("  DB: %d inserted, %d updated" % (inserted, updated))

            # Show sample
            if records:
                sample = records[0]
                print("  Sample: %s | %s | qty=%s | progress=%s" % (
                    sample["order_no"], sample["product_code"],
                    sample["quantity"], sample["process_progress"][:80]))

            # Cleanup old files (keep last 3)
            try:
                files = sorted([f for f in os.listdir(DOWNLOAD_DIR)
                    if f.startswith("mes_tasks_") and f.endswith(".xlsx")])
                for old in files[:-3]:
                    os.remove(os.path.join(DOWNLOAD_DIR, old))
            except:
                pass

            # Print import count for wrapper script
            print("IMPORTED:%d" % (inserted + updated))

        except Exception as e:
            print("Error: %s" % e)
            import traceback
            traceback.print_exc()
            pass  # screenshot disabled
        finally:
            browser.close()


if __name__ == "__main__":
    run_fetch_orders()
