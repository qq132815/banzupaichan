# -*- coding: utf-8 -*-
import openpyxl
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.db import get_connection


def _match_headers(ws, field_map):
    """Match column indices by header names.
    field_map: dict of {field_name: [possible_header_names]}
    Returns: (dict of {field_name: column_index}, headers_list) or raises ValueError
    Auto-detects header row: uses row 1 if it has 2+ non-empty values, else row 2.
    """
    # Auto-detect header row
    row1_vals = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    row1_count = sum(1 for v in row1_vals if v)
    if row1_count >= 2:
        headers = row1_vals
        header_row = 1
    else:
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[2]]
        header_row = 2
    # Store header_row for caller to use as min_row
    ws._import_header_row = header_row
    
    result = {}
    missing = []
    for field, aliases in field_map.items():
        found = False
        for alias in aliases:
            for i, h in enumerate(headers):
                if alias.lower() in h.lower() or h.lower() in alias.lower():
                    result[field] = i
                    found = True
                    break
            if found:
                break
        if not found:
            missing.append(field)
    
    if missing:
        raise ValueError("未找到以下列: " + ", ".join(missing) + "。表头: " + ", ".join(headers))
    return result, headers



def import_shipping_plan(file_path, plan_month=None):
    """Import shipping plan from Excel.
    Supports two formats:
    1. New template: 客户/项目/客户件号/广升件号/名称 + 1号~31号
    2. Legacy: product_code + date columns with datetime headers
    """
    import re as _re
    from datetime import date
    wb = openpyxl.load_workbook(file_path, data_only=True)
    conn = get_connection()
    c = conn.cursor()
    # Clear existing data for the month to allow re-import
    if plan_month:
        c.execute("DELETE FROM shipping_plan WHERE plan_month=?", (plan_month,))
    else:
        c.execute("DELETE FROM shipping_plan WHERE plan_month IS NULL")

    count = 0
    for ws in wb.worksheets:
        all_rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
            all_rows.append(list(row))
        if not all_rows:
            continue

        # Detect format: check if row 2 has "客户" in col A
        header_row = None
        for ri in range(min(3, len(all_rows))):
            row_vals = all_rows[ri]
            if row_vals and row_vals[0] and str(row_vals[0]).strip() in ('客户', '广升件号'):
                header_row = ri
                break

        if header_row is not None:
            # New template format
            headers = all_rows[header_row]
            day_cols = []
            for ci, h in enumerate(headers):
                if h:
                    m = _re.match(r'^(\d{1,2})\u53f7$', str(h).strip())
                    if m:
                        day_cols.append((ci, int(m.group(1))))

            if not day_cols:
                continue

            if not plan_month:
                fname = os.path.basename(file_path)
                m = _re.search(r'(\d{4})[.\-/](\d{1,2})\u6708', fname)
                if m:
                    plan_month = m.group(1) + '-' + m.group(2).zfill(2)
                else:
                    plan_month = date.today().strftime('%Y-%m')

            data_start = header_row + 2
            if data_start < len(all_rows) and all_rows[data_start] and all_rows[data_start][0]:
                val = str(all_rows[data_start][0] or '')
                if val.startswith('\u661f\u671f'):
                    data_start += 1

            current_customer = None
            current_project = None

            for ri in range(data_start, ws.max_row):
                row_vals_list = list(ws.iter_rows(min_row=ri+1, max_row=ri+1, values_only=True))
                if not row_vals_list or not row_vals_list[0]:
                    continue
                row = list(row_vals_list[0])

                col_a = row[0] if len(row) > 0 else None
                col_b = row[1] if len(row) > 1 else None
                col_c = row[2] if len(row) > 2 else None
                col_d = row[3] if len(row) > 3 else None
                col_e = row[4] if len(row) > 4 else None

                if col_a:
                    current_customer = str(col_a).strip()
                if col_b:
                    current_project = str(col_b).strip().replace('\n', ' ')

                product_code = str(col_c).strip() if col_c else None
                if not product_code:
                    continue

                gs_part_no = str(col_d).strip() if col_d else ''
                product_name = str(col_e).strip() if col_e else ''

                for ci, day_num in day_cols:
                    if ci >= len(row):
                        continue
                    qty = row[ci]
                    if qty is None:
                        continue
                    try:
                        qty_val = float(qty)
                    except (ValueError, TypeError):
                        continue
                    if qty_val <= 0:
                        continue

                    try:
                        year, month = int(plan_month[:4]), int(plan_month[5:7])
                        ship_date = date(year, month, day_num).strftime('%Y-%m-%d')
                    except Exception:
                        continue

                    c.execute(
                        "INSERT INTO shipping_plan (product_code, quantity, ship_date, plan_month, customer, project, customer_part_no, gs_part_no, product_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (product_code, qty_val, ship_date, plan_month,
                         current_customer or '', current_project or '',
                         product_code, gs_part_no, product_name)
                    )
                    count += 1
        else:
            # Legacy format: row 1 has datetime date headers
            headers = all_rows[0]
            dates = []
            for i, h in enumerate(headers):
                if hasattr(h, 'strftime'):
                    dates.append((i, h))
            if not dates:
                continue
            for ri in range(1, ws.max_row):
                row_vals_list = list(ws.iter_rows(min_row=ri+1, max_row=ri+1, values_only=True))
                if not row_vals_list or not row_vals_list[0]:
                    continue
                row = list(row_vals_list[0])
                if not row[0]:
                    continue
                product_code = str(row[0]).strip()
                for col_idx, ship_date in dates:
                    qty = row[col_idx] if col_idx < len(row) else None
                    if qty is None:
                        continue
                    try:
                        qty_val = float(qty)
                    except (ValueError, TypeError):
                        continue
                    if qty_val <= 0:
                        continue
                    date_str = ship_date.strftime('%Y-%m-%d')
                    c.execute(
                        "INSERT INTO shipping_plan (product_code, quantity, ship_date, plan_month, customer, project, customer_part_no, gs_part_no, product_name) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (product_code, qty_val, date_str,
                         ship_date.strftime('%Y-%m'), '', '', '', product_code, '')
                    )
                    count += 1

    conn.commit()
    conn.close()
    return count

def import_production_cycles(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM production_cycles")
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, '_import_header_row', 1) + 1, values_only=True):
        if not row or not row[0]:
            continue
        product_code = str(row[0]).strip()
        prod_days = row[1] if row[1] else 1
        lead_days = row[2] if row[2] else 0
        c.execute("INSERT INTO production_cycles (product_code, production_days, lead_days) VALUES (?, ?, ?)",
                  (product_code, prod_days, lead_days))
        count += 1
    conn.commit()
    conn.close()
    return count


def import_work_orders(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("ALTER TABLE work_orders ADD COLUMN process_progress TEXT")
    except:
        pass
    cols, headers = _match_headers(ws, {
        'order_no': ['工单编号', '工单号', '订单编号'],
        'product_code': ['产品编号', '产品编码', '件号'],
        'product_name': ['产品名称', '产品图', '品名'],
        'status': ['工单状态', '状态'],
        'process_progress': ['工单进度条', '工序进度', '进度条'],
        'quantity': ['计划数', '数量', '计划数量'],
        'priority': ['优先级'],
        'due_date': ['计划结束时间', '交期', '截止日期', '完工日期'],
        'completed_qty': ['完工数', '完成数', '已完工数'],
        'source': ['产品来源', '来源'],
        'route_code': ['关联单据', '关联编号'],
    })
    count = 0
    header_row = ws._import_header_row
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[cols['order_no']]:
            continue
        order_no = str(row[cols['order_no']] or '').strip()
        if not order_no:
            continue
        product_code = str(row[cols.get('product_code', -1)] or '').strip() if 'product_code' in cols and cols.get('product_code', -1) < len(row) else ''
        product_name = str(row[cols.get('product_name', -1)] or '').strip() if 'product_name' in cols and cols.get('product_name', -1) < len(row) else ''
        if not product_name:
            product_name = product_code
        quantity = 0
        if 'quantity' in cols and cols['quantity'] < len(row) and row[cols['quantity']]:
            try: quantity = float(row[cols['quantity']])
            except: pass
        priority = 'P2'
        if 'priority' in cols and cols['priority'] < len(row) and row[cols['priority']]:
            priority = str(row[cols['priority']]).strip()
        due_date = None
        if 'due_date' in cols and cols['due_date'] < len(row) and row[cols['due_date']]:
            val = row[cols['due_date']]
            if hasattr(val, 'strftime'):
                due_date = val.strftime('%Y-%m-%d')
            else:
                due_date = str(val).strip()[:10]
        status = 'pending'
        if 'status' in cols and cols['status'] < len(row) and row[cols['status']]:
            status = str(row[cols['status']]).strip()
        completed_qty = 0
        if 'completed_qty' in cols and cols['completed_qty'] < len(row) and row[cols['completed_qty']]:
            try: completed_qty = float(row[cols['completed_qty']])
            except: pass
        process_progress = ''
        if 'process_progress' in cols and cols['process_progress'] < len(row) and row[cols['process_progress']]:
            process_progress = str(row[cols['process_progress']]).strip()
        source = ''
        if 'source' in cols and cols['source'] < len(row) and row[cols['source']]:
            source = str(row[cols['source']]).strip()
        route_code = ''
        if 'route_code' in cols and cols['route_code'] < len(row) and row[cols['route_code']]:
            route_code = str(row[cols['route_code']]).strip()
        c.execute("INSERT OR REPLACE INTO work_orders (order_no, product_code, product_name, quantity, completed_qty, due_date, priority, status, source, route_code, process_progress) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                  (order_no, product_code, product_name, quantity, completed_qty, due_date, priority, status, source, route_code, process_progress))
        count += 1
    conn.commit()
    conn.close()
    return count


def import_processes(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM processes")
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, '_import_header_row', 1) + 1, values_only=True):
        if not row or not row[0]:
            continue
        process_code = str(row[0]).strip()
        process_name = str(row[1]).strip() if row[1] else ''
        team_name = str(row[2]).strip() if row[2] else None
        c.execute("INSERT OR REPLACE INTO processes (process_code, process_name, team_name) VALUES (?, ?, ?)",
                  (process_code, process_name, team_name))
        count += 1
    conn.commit()
    conn.close()
    return count


def import_bom(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM bom")
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, '_import_header_row', 1) + 1, values_only=True):
        if not row or not row[0]:
            continue
        parent_code = str(row[2]).strip() if row[2] else ''  # 父项产品编号
        parent_name = str(row[3]).strip() if row[3] else ''  # 父项产品名称
        child_code = str(row[6]).strip() if row[6] else ''   # 子项产品编号
        child_name = str(row[7]).strip() if row[7] else ''   # 子项产品名称
        
        # Skip if child code starts with "01." (purchased materials)
        if child_code.startswith('01.'):
            continue
            
        qty = abs(float(row[10])) if row[10] else 1  # 单位用量
        unit = str(row[9]).strip() if row[9] else ''  # 子项单位
        process_team = str(row[11]).strip() if row[11] else None  # 备注
        
        c.execute("INSERT INTO bom (parent_product_code, parent_product_name, child_product_code, child_product_name, quantity, unit, process_team) VALUES (?, ?, ?, ?, ?, ?, ?)",
                  (parent_code, parent_name, child_code, child_name, qty, unit, process_team))
        count += 1
    conn.commit()
    conn.close()
    return count


def import_process_routes(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    cols, headers = _match_headers(ws, {
        'code': ['工艺路线编号', '路线编码', 'route_code', '编码'],
        'name': ['工艺路线名称', '路线名称', 'route_name', '名称'],
        'processes': ['包含工序列表', '工序列表', 'process_list', '工序'],
        'product': ['产品编号', '产品编码', '成品件号', 'product_code', '产品'],
    })
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM process_routes")
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, '_import_header_row', 1) + 1, values_only=True):
        if not row:
            continue
        route_code = str(row[cols['code']]).strip() if cols['code'] < len(row) and row[cols['code']] else ''
        route_name = str(row[cols['name']]).strip() if cols['name'] < len(row) and row[cols['name']] else ''
        process_list = str(row[cols['processes']]).strip() if cols['processes'] < len(row) and row[cols['processes']] else ''
        product_code = str(row[cols['product']]).strip() if 'product' in cols and cols['product'] < len(row) and row[cols['product']] else None
        c.execute("INSERT INTO process_routes (route_code, route_name, product_code, process_list) VALUES (?, ?, ?, ?)",
                  (route_code, route_name, product_code, process_list))
        count += 1
    conn.commit()
    conn.close()
    return count


def import_equipment(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, name FROM teams")
    team_map = {r[1]: r[0] for r in c.fetchall()}
    sample = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    has_id_col = sample and len(sample) >= 3 and isinstance(sample[0], (int, float)) and isinstance(sample[2], str)
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, "_import_header_row", 1) + 1, values_only=True):
        if not row or not any(row):
            continue
        if has_id_col:
            eq_code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            eq_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            team_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        else:
            eq_code = str(row[0]).strip() if row[0] else ""
            eq_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            team_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if not eq_code:
            continue
        team_id = team_map.get(team_name)
        if team_id is None:
            for tn, tid in team_map.items():
                if tn in team_name or team_name in tn:
                    team_id = tid
                    break
        if team_id is None:
            try:
                team_id = int(float(team_name))
            except (ValueError, TypeError):
                team_id = 1
        equipment_type = "normal"
        c.execute("INSERT OR REPLACE INTO equipments (equipment_code, equipment_name, team_id, equipment_type, status, capacity_per_hour, location) VALUES (?, ?, ?, ?, ?, ?, ?)",
                  (eq_code, eq_name, team_id, equipment_type, "normal", 0, ""))
        count += 1
    conn.commit()
    conn.close()
    return count
def import_reports(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, '_import_header_row', 1) + 1, values_only=True):
        if not row or not row[0]:
            continue
        order_no = str(row[0]).strip() if row[0] else ''
        process_code = str(row[1]).strip() if row[1] else ''
        equipment_id = int(row[2]) if row[2] else None
        team_id = int(row[3]) if row[3] else None
        report_date = row[4].strftime('%Y-%m-%d') if hasattr(row[4], 'strftime') else str(row[4]) if row[4] else ''
        planned_qty = float(row[5]) if row[5] else 0
        actual_qty = float(row[6]) if row[6] else 0
        operator = str(row[7]).strip() if row[7] else ''
        c.execute("INSERT INTO reports (work_order_no, process_code, equipment_id, team_id, report_date, planned_qty, actual_qty, operator) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                  (order_no, process_code, equipment_id, team_id, report_date, planned_qty, actual_qty, operator))
        count += 1
    conn.commit()
    conn.close()
    return count

def import_equipments(file_path):
    """Import equipments from Excel file. Columns: id, equipment_code, equipment_name, team_name"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM equipments")
    c.execute("SELECT id, name FROM teams")
    team_map = {r[1]: r[0] for r in c.fetchall()}
    count = 0
    for row in ws.iter_rows(min_row=getattr(ws, '_import_header_row', 1) + 1, values_only=True):
        if not row or not row[0]:
            continue
        eq_code = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        eq_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        team_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        if not eq_code:
            continue
        team_id = team_map.get(team_name)
        if team_id is None:
            for tn, tid in team_map.items():
                if tn in team_name or team_name in tn:
                    team_id = tid
                    break
        if team_id is None:
            team_id = 1
        c.execute("INSERT INTO equipments (equipment_code, equipment_name, team_id, equipment_type, status, capacity_per_hour, location) VALUES (?,?,?,?,?,?,?)",
                  (eq_code, eq_name, team_id, 'normal', 'normal', 0, ''))
        count += 1
    conn.commit()
    conn.close()
    return count

def import_personnel(filepath):
    """Import personnel from Excel. Columns: 员工UserID, 姓名, 3级部门, 职位"""
    DEPT_MAP = {
        "前道班": "前段",
        "焊接班": "焊接",
        "扣压班": "扣压",
        "后道班": "装配包装",
    }
    import openpyxl
    from utils.db import get_connection, get_all_teams
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    cols, headers = _match_headers(ws, {
        'user_id': ['员工UserID', 'UserID', '工号', '员工ID'],
        'name': ['姓名', '员工姓名', '名字'],
        'department': ['3级部门', '部门', '班组', '班组名称'],
        'position': ['职位', '岗位', '职务'],
    })
    teams = get_all_teams()
    team_map = {t['name']: t['id'] for t in teams}
    conn = get_connection()
    c = conn.cursor()
    count = 0
    header_row = ws._import_header_row
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[cols['name']]:
            continue
        user_id = str(row[cols['user_id']] or '').strip() if 'user_id' in cols else ''
        name = str(row[cols['name']] or '').strip()
        dept_raw = str(row[cols['department']] or '').strip() if 'department' in cols else ''
        position = str(row[cols['position']] or '').strip() if 'position' in cols else ''
        if not name:
            continue
        dept_mapped = DEPT_MAP.get(dept_raw, dept_raw)
        team_id = team_map.get(dept_mapped)
        if team_id is None:
            for tn, tid in team_map.items():
                if tn in dept_mapped or dept_mapped in tn:
                    team_id = tid
                    break
        c.execute("INSERT OR REPLACE INTO personnel (user_id, name, department, position, team_id) VALUES (?, ?, ?, ?, ?)",
                  (user_id, name, dept_raw, position, team_id))
        count += 1
    conn.commit()
    conn.close()
    return count


def import_molds(filepath):
    """Import molds from Excel."""
    import openpyxl
    from utils.db import get_connection, get_all_teams
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    try:
        cols, headers = _match_headers(ws, {
            'mold_code': ['模具编号', '模具编码', '编号'],
            'mold_name': ['模具名称', '名称'],
            'mold_type': ['模具类型', '类型'],
            'product_code': ['产品编码', '产品编号', '关联产品'],
            'location': ['位置', '存放位置', '库位'],
            'remark': ['备注', '说明'],
        })
    except ValueError:
        cols, headers = _match_headers(ws, {
            'mold_code': ['模具编号', '模具编码', '编号'],
            'mold_name': ['模具名称', '名称'],
        })
    teams = get_all_teams()
    team_map = {t['name']: t['id'] for t in teams}
    conn = get_connection()
    c = conn.cursor()
    count = 0
    header_row = ws._import_header_row
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[cols['mold_code']]:
            continue
        mold_code = str(row[cols['mold_code']] or '').strip()
        mold_name = str(row[cols['mold_name']] or '').strip() if 'mold_name' in cols else ''
        if not mold_code:
            continue
        mold_type = str(row[cols.get('mold_type', -1)] or '').strip() if 'mold_type' in cols and cols.get('mold_type', -1) < len(row) else ''
        product_code = str(row[cols.get('product_code', -1)] or '').strip() if 'product_code' in cols and cols.get('product_code', -1) < len(row) else ''
        location = str(row[cols.get('location', -1)] or '').strip() if 'location' in cols and cols.get('location', -1) < len(row) else ''
        remark = str(row[cols.get('remark', -1)] or '').strip() if 'remark' in cols and cols.get('remark', -1) < len(row) else ''
        c.execute("INSERT OR REPLACE INTO molds (mold_code, mold_name, mold_type, product_code, status, location, remark) VALUES (?,?,?,?,?,?,?)",
                  (mold_code, mold_name, mold_type, product_code, 'normal', location, remark))
        count += 1
    conn.commit()
    conn.close()
    return count

def import_work_reports(filepath):
    """Import work reports from Excel."""
    import openpyxl
    from utils.db import get_connection
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    cols, headers = _match_headers(ws, {
        'report_qty': ['报工数', '报工数量'],
        'good_qty': ['报工良品数', '良品数'],
        'bad_qty': ['报工不良数', '不良数'],
        'report_unit': ['报工单位', '单位'],
        'good_rate': ['报工良品率', '良品率'],
        'operator': ['生产人员', '操作人', '报工人'],
        'start_time': ['报工开始时间', '开始时间'],
        'end_time': ['报工结束时间', '结束时间'],
        'approve_status': ['审批状态', '状态'],
        'approver': ['审批人'],
        'approve_time': ['审批时间'],
        'creator': ['创建人'],
        'create_time': ['报工创建时间', '创建时间'],
        'process_name': ['工序名称', '工序'],
        'order_no': ['工单编号', '工单号'],
        'product_code': ['产品编号', '产品编码'],
        'product_name': ['产品名称'],
        'related_no': ['关联单据号', '关联单据'],
        'equipment': ['设备机台', '设备'],
        'report_hours': ['报工时长', '时长'],
        'weld_count': ['焊点数量', '焊点'],
        'attendance_note': ['出勤人员备注', '备注'],
    })
    conn = get_connection()
    c = conn.cursor()
    count = 0
    header_row = ws._import_header_row
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        def gv(field):
            if field not in cols or cols[field] >= len(row):
                return ''
            v = row[cols[field]]
            return str(v).strip() if v is not None else ''
        def gf(field):
            if field not in cols or cols[field] >= len(row):
                return 0
            v = row[cols[field]]
            try: return float(v)
            except: return 0
        report_qty = gf('report_qty')
        if report_qty == 0 and not gv('order_no'):
            continue
        # Parse report_hours - formats: "30分钟", "6小时30分钟", "5小时", "2.5"
        report_hours_raw = gv('report_hours')
        report_hours = 0
        if report_hours_raw and report_hours_raw != '-':
            import re as _re
            hours_val = 0
            mins_val = 0
            hm = _re.search(r'([\d.]+)\s*小时', report_hours_raw)
            if hm:
                hours_val = float(hm.group(1))
            mm = _re.search(r'([\d.]+)\s*分钟', report_hours_raw)
            if mm:
                mins_val = float(mm.group(1))
            if hours_val > 0 or mins_val > 0:
                report_hours = round(hours_val + mins_val / 60, 2)
            else:
                m = _re.search(r'([\d.]+)', report_hours_raw)
                if m:
                    report_hours = float(m.group(1))
        # If no hours, calculate from start/end time
        if report_hours == 0:
            st = gv('start_time')
            et = gv('end_time')
            if st and et:
                try:
                    from datetime import datetime
                    fmt = '%Y-%m-%d %H:%M:%S'
                    t1 = datetime.strptime(st[:19], fmt)
                    t2 = datetime.strptime(et[:19], fmt)
                    report_hours = round((t2 - t1).total_seconds() / 3600, 2)
                except:
                    pass
        c.execute("""INSERT INTO work_reports
            (report_qty, good_qty, bad_qty, report_unit, good_rate, operator, start_time, end_time,
             approve_status, approver, approve_time, creator, create_time, process_name, order_no,
             product_code, product_name, related_no, equipment, report_hours, weld_count, attendance_note)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (report_qty, gf('good_qty'), gf('bad_qty'), gv('report_unit'), gv('good_rate'),
             gv('operator'), gv('start_time'), gv('end_time'), gv('approve_status'),
             gv('approver'), gv('approve_time'), gv('creator'), gv('create_time'),
             gv('process_name'), gv('order_no'), gv('product_code'), gv('product_name'),
             gv('related_no'), gv('equipment'), report_hours, gf('weld_count'), gv('attendance_note')))
        count += 1
    conn.commit()
    conn.close()
    return count

    """�����Ʒ����"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    
    # �ֶ�ӳ��
    field_map = {
        'product_code': ['��Ʒ����', '��Ʒ���', '����'],
        'product_name': ['��Ʒ����', '����'],
        'specifications': ['����ͺ�', '���'],
        'category': ['����', '��Ʒ����'],
        'customer': ['�ͻ�', '�ͻ�����'],
        'project': ['��Ŀ', '��Ŀ����'],
        'description': ['����', '��Ʒ����'],
        'status': ['״̬'],
        'image_url': ['ͼƬURL', 'ͼƬ����', 'ͼƬ']
    }
    
    try:
        col_map, headers = _match_headers(ws, field_map)
    except ValueError:
        conn.close()
        raise
    
    count = 0
    header_row = getattr(ws, '_import_header_row', 1)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        product_code = str(row[col_map['product_code']]).strip() if row[col_map['product_code']] else ''
        if not product_code:
            continue
        
        product_name = str(row[col_map['product_name']]).strip() if row[col_map['product_name']] else ''
        specifications = str(row[col_map['specifications']]).strip() if row[col_map.get('specifications')] else ''
        category = str(row[col_map['category']]).strip() if row[col_map.get('category')] else ''
        customer = str(row[col_map['customer']]).strip() if row[col_map.get('customer')] else ''
        project = str(row[col_map['project']]).strip() if row[col_map.get('project')] else ''
        description = str(row[col_map['description']]).strip() if row[col_map.get('description')] else ''
        status = str(row[col_map['status']]).strip() if row[col_map.get('status')] else 'active'
        image_url = str(row[col_map['image_url']]).strip() if row[col_map.get('image_url')] else ''
        
        # ����ͼƬ�������URL��
        image_path = ''
        if image_url:
            try:
                from app import download_product_image
                image_path = download_product_image(image_url, product_code) or ''
            except:
                image_path = ''
        
        # ʹ�� INSERT OR REPLACE �����»����
        c.execute("""INSERT OR REPLACE INTO products 
            (product_code, product_name, specifications, category, customer, project, 
             description, status, image_url, image_path, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))""",
            (product_code, product_name, specifications, category, customer, project,
             description, status, image_url, image_path))
        count += 1
    
    conn.commit()
    conn.close()
    return count

def import_product_definitions(file_path):
    """导入产品定义"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    conn = get_connection()
    c = conn.cursor()
    
    # 字段映射 - 匹配实际表头
    field_map = {
        'product_type': ['产品类型'],
        'product_code': ['产品编号', '产品编码', '编码'],
        'product_name': ['产品名称', '名称'],
        'specifications': ['产品规格', '规格型号', '规格'],
        'unit': ['单位'],
        'route_code': ['工艺路线'],
        'safety_stock_min': ['最小安全库存'],
        'safety_stock_max': ['最大安全库存'],
        'stock_qty': ['库存数量'],
        'source': ['产品来源'],
        'image_url': ['产品图', '图片URL', '图片链接', '图片'],
        'customer': ['客户', '客户名称'],
        'basket_capacity': ['每筐容量']
    }
    
    try:
        col_map, headers = _match_headers(ws, field_map)
    except ValueError:
        conn.close()
        raise
    
    count = 0
    header_row = getattr(ws, '_import_header_row', 1)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        product_code = str(row[col_map['product_code']]).strip() if row[col_map['product_code']] else ''
        if not product_code:
            continue
        
        product_name = str(row[col_map['product_name']]).strip() if row[col_map.get('product_name')] else ''
        product_type = str(row[col_map['product_type']]).strip() if row[col_map.get('product_type')] else ''
        specifications = str(row[col_map['specifications']]).strip() if row[col_map.get('specifications')] else ''
        unit = str(row[col_map['unit']]).strip() if row[col_map.get('unit')] else ''
        route_code = str(row[col_map['route_code']]).strip() if row[col_map.get('route_code')] else ''
        safety_stock_min = row[col_map.get('safety_stock_min')] if row[col_map.get('safety_stock_min')] else 0
        safety_stock_max = row[col_map.get('safety_stock_max')] if row[col_map.get('safety_stock_max')] else 0
        stock_qty = row[col_map.get('stock_qty')] if row[col_map.get('stock_qty')] else 0
        source = str(row[col_map.get('source')]).strip() if row[col_map.get('source')] else ''
        image_url = str(row[col_map['image_url']]).strip() if row[col_map.get('image_url')] else ''
        customer = str(row[col_map['customer']]).strip() if row[col_map.get('customer')] else ''
        basket_capacity = row[col_map.get('basket_capacity')] if row[col_map.get('basket_capacity')] else 0
        
        # 下载图片（如果有URL）
        image_path = ''
        if image_url:
            try:
                import requests
                from urllib.parse import urlparse
                
                # 创建图片目录
                images_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'images', 'products')
                os.makedirs(images_dir, exist_ok=True)
                
                # 生成文件名
                url_path = urlparse(image_url).path
                ext = os.path.splitext(url_path)[1] or '.jpg'
                safe_code = product_code.replace('/', '_').replace('\\', '_').replace(':', '_')
                filename = f"{safe_code}{ext}"
                filepath = os.path.join(images_dir, filename)
                
                # 下载图片
                response = requests.get(image_url, timeout=30)
                response.raise_for_status()
                
                # 保存图片
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                
                image_path = f"/static/images/products/{filename}"
            except Exception as e:
                print(f"下载图片失败 {product_code}: {e}")
                image_path = ''
        
        # 使用 INSERT OR REPLACE 来更新或插入
        c.execute("""INSERT OR REPLACE INTO products 
            (product_code, product_name, product_type, specifications, unit, route_code,
             safety_stock, stock_qty, source, image_url, image_path, customer, basket_capacity, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))""",
            (product_code, product_name, product_type, specifications, unit, route_code,
             safety_stock_min, stock_qty, source, image_url, image_path, customer, basket_capacity))
        count += 1
    
    conn.commit()
    conn.close()
    return count




